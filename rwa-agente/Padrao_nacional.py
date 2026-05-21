# PADRAO_NACIONAL_0105_4_CERT_MATCH_CORRIGIDO
# Base principal: V18E validada.
# Alteracao cirurgica:
# - adiciona rota CERTIFICADO DIGITAL antes do mesmo fluxo pos-login da V18E
# - mantem capturar_links_menu() da V18E
# - mantem ir_para_proxima_pagina() da V18E
# - mantem download direto por requests/cookies
# - força nomes controlados PDF/XML por pagina/nota para evitar sobrescrita pelo nome do portal
# - NÃO faz limpeza inicial da pasta da competência
# - AJUSTE RELATÓRIO: conta PDF/XML baixados na execução atual, não arquivos antigos da pasta

import os
import re
import time
import threading
import unicodedata
from difflib import SequenceMatcher
import openpyxl
from datetime import datetime, date
from dateutil.relativedelta import relativedelta
import calendar
import requests
import pyautogui
import pytesseract
import xml.etree.ElementTree as ET

from PIL import ImageGrab, ImageOps, ImageEnhance
from urllib.parse import urljoin, urlparse, unquote
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


import sys as _sys
import smtplib as _smtp_h
from email.mime.multipart import MIMEMultipart as _MMulti_h
from email.mime.text import MIMEText as _MText_h
_SCRIPT_DIR = os.path.dirname(os.path.abspath(_sys.executable if getattr(_sys, 'frozen', False) else __file__))

# =====================================================================
# CONFIGURAÇÕES POR CLIENTE — edite aqui a cada nova instalação
# =====================================================================
_NOME_CLIENTE    = "Roberto William de Angelo"               # (1) ← MUDE O NOME DO CLIENTE AQUI
# =====================================================================

PASTA_BASE = r"G:\Meu Drive\AUTOMACOES\PADRAO NACIONAL"
_hoje        = date.today()
_competencia = _hoje - relativedelta(months=1)
COMPETENCIA  = _competencia.strftime("%m%Y")
DATA_INICIAL = _competencia.replace(day=1).strftime("%d/%m/%Y")
DATA_FINAL   = _competencia.replace(day=calendar.monthrange(_competencia.year, _competencia.month)[1]).strftime("%d/%m/%Y")
URL_BASE = "https://www.nfse.gov.br"
URL_LOGIN = "https://www.nfse.gov.br/EmissorNacional/Login"

RELATORIO_NOME = f"Relatorio_Gerencial_Padrao_Nacional_{COMPETENCIA}.xlsx"
CONCLUSAO_NOME = f"CONCLUSAO_PADRAO_NACIONAL_{COMPETENCIA}.txt"
DIAGNOSTICO_NOME = f"DIAGNOSTICO_PADRAO_NACIONAL_{COMPETENCIA}.txt"
PASTA_DIAG_TECNICO = os.path.join(PASTA_BASE, "DIAGNOSTICO TECNICO")
_log_diag_pn = []  # Registro de erros para resumo do diagnóstico

# OCR certificado
pytesseract.pytesseract.tesseract_cmd = r"C:\Program Files\Tesseract-OCR\tesseract.exe"
os.environ["TESSDATA_PREFIX"] = r"C:\Program Files\Tesseract-OCR\tessdata"

pyautogui.FAILSAFE = False
pyautogui.PAUSE = 0.15

# =====================================================================
# COORDENADAS DO CERTIFICADO - LEITURA POR ARQUIVO EXTERNO
# Arquivo obrigatório ao lado do .py/.exe: COORDENADAS OCR.txt
# Se faltar, estiver vazio ou inválido, a automação trava antes do login
# e grava CONCLUSAO + RELATORIO_ERROS.
# =====================================================================

def _salvar_erro_coordenadas(mensagem, detalhes=None):
    detalhes = detalhes or []
    destinos = []

    for pasta in (_SCRIPT_DIR, PASTA_BASE):
        try:
            if pasta and os.path.isdir(pasta) and pasta not in destinos:
                destinos.append(pasta)
        except Exception:
            pass

    if not destinos:
        destinos = [_SCRIPT_DIR]

    caminho_coord = os.path.join(_SCRIPT_DIR, "COORDENADAS OCR.txt")

    linhas_conclusao = [
        "=" * 70,
        "CONCLUSAO — ERRO GRAVE DE CONFIGURACAO",
        "=" * 70,
        f"Competencia : {COMPETENCIA[:2]}/{COMPETENCIA[2:]}",
        f"Data/Hora   : {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}",
        f"Arquivo     : {caminho_coord}",
        "",
        mensagem,
        "",
        "STATUS FINAL: PROCESSO INTERROMPIDO ANTES DO INICIO.",
        "Motivo: arquivo de coordenadas ausente, incompleto ou invalido.",
        "Nenhum navegador foi aberto.",
        "Nenhum download foi iniciado.",
    ]

    if detalhes:
        linhas_conclusao.append("")
        linhas_conclusao.append("DETALHES:")
        for item in detalhes:
            linhas_conclusao.append(f"- {item}")

    linhas_erros = [
        "=" * 80,
        f"RELATÓRIO DE ERROS — PADRÃO NACIONAL — COMPETÊNCIA {COMPETENCIA[:2]}/{COMPETENCIA[2:]}",
        f"Gerado em: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}",
        "=" * 80,
        "",
        "ERRO GRAVE DE CONFIGURACAO",
        "Tipo     : COORDENADAS OCR OBRIGATORIAS",
        f"Arquivo  : {caminho_coord}",
        f"Erro     : {mensagem}",
        "",
        "Acao requerida:",
        "Corrigir o arquivo COORDENADAS OCR.txt antes de executar novamente.",
        "",
    ]

    if detalhes:
        linhas_erros.append("Detalhes:")
        for item in detalhes:
            linhas_erros.append(f"- {item}")
        linhas_erros.append("")

    linhas_erros.append(f"Total de erros: {max(1, len(detalhes))}")

    caminhos_salvos = []
    for pasta in destinos:
        try:
            caminho_conclusao = os.path.join(pasta, CONCLUSAO_NOME)
            with open(caminho_conclusao, "w", encoding="utf-8") as f:
                f.write("\n".join(linhas_conclusao))
            caminhos_salvos.append(caminho_conclusao)
        except Exception:
            pass

        try:
            caminho_erros = os.path.join(PASTA_DIAG_TECNICO, f"RELATORIO_ERROS_PADRAO_NACIONAL_{COMPETENCIA}.txt")
            os.makedirs(PASTA_DIAG_TECNICO, exist_ok=True)
            with open(caminho_erros, "w", encoding="utf-8") as f:
                f.write("\n".join(linhas_erros))
            caminhos_salvos.append(caminho_erros)
        except Exception:
            pass

    print()
    print("=" * 70)
    print("  ERRO GRAVE -- COORDENADAS OCR NAO CONFIGURADAS")
    print("=" * 70)
    print("  A automacao foi interrompida antes de abrir o navegador.")
    print(f"  Arquivo esperado: {caminho_coord}")
    print(f"  Motivo: {mensagem}")
    for item in detalhes:
        print(f"  - {item}")
    if caminhos_salvos:
        print("  Relatorios de erro salvos em:")
        for caminho in caminhos_salvos:
            print(f"  {caminho}")
    print("=" * 70)


def _ler_inteiro_config(dados, chave, erros):
    valor = dados.get(chave, "").strip()
    if not valor:
        erros.append(f"Campo vazio ou ausente: {chave}")
        return None
    try:
        numero = int(valor)
        if numero < 0:
            erros.append(f"Campo com numero negativo: {chave}={valor}")
            return None
        return numero
    except Exception:
        erros.append(f"Campo sem numero inteiro valido: {chave}={valor}")
        return None


def carregar_coordenadas_ocr():
    caminho = os.path.join(_SCRIPT_DIR, "COORDENADAS OCR.txt")

    if not os.path.exists(caminho):
        raise RuntimeError(f"Arquivo obrigatorio nao encontrado: {caminho}")

    dados = {}
    with open(caminho, "r", encoding="utf-8") as f:
        for linha in f:
            linha = linha.strip()
            if not linha or linha.startswith("=") or linha.startswith("#"):
                continue
            if "=" not in linha:
                continue
            chave, valor = linha.split("=", 1)
            dados[chave.strip().upper()] = valor.strip()

    erros = []
    campos = [
        "LINHA_1_X", "LINHA_1_Y",
        "LINHA_2_X", "LINHA_2_Y",
        "LINHA_3_X", "LINHA_3_Y",
        "BOTAO_OK_X", "BOTAO_OK_Y",
        "SUPERIOR_ESQUERDO_X", "SUPERIOR_ESQUERDO_Y",
        "SUPERIOR_DIREITO_X", "SUPERIOR_DIREITO_Y",
        "INFERIOR_ESQUERDO_X", "INFERIOR_ESQUERDO_Y",
        "INFERIOR_DIREITO_X", "INFERIOR_DIREITO_Y",
        "MAXIMO_ROLAGENS",
    ]

    valores = {chave: _ler_inteiro_config(dados, chave, erros) for chave in campos}

    if not erros:
        top_medio = abs(valores["SUPERIOR_ESQUERDO_Y"] - valores["SUPERIOR_DIREITO_Y"])
        bottom_medio = abs(valores["INFERIOR_ESQUERDO_Y"] - valores["INFERIOR_DIREITO_Y"])
        if top_medio > 25:
            erros.append("Os pontos superiores do OCR estao muito desalinhados entre si.")
        if bottom_medio > 25:
            erros.append("Os pontos inferiores do OCR estao muito desalinhados entre si.")

        ocr_left = min(valores["SUPERIOR_ESQUERDO_X"], valores["INFERIOR_ESQUERDO_X"])
        ocr_top = min(valores["SUPERIOR_ESQUERDO_Y"], valores["SUPERIOR_DIREITO_Y"])
        ocr_right = max(valores["SUPERIOR_DIREITO_X"], valores["INFERIOR_DIREITO_X"])
        ocr_bottom = max(valores["INFERIOR_ESQUERDO_Y"], valores["INFERIOR_DIREITO_Y"])

        if ocr_right <= ocr_left:
            erros.append("Area OCR invalida: direita menor ou igual a esquerda.")
        if ocr_bottom <= ocr_top:
            erros.append("Area OCR invalida: inferior menor ou igual ao superior.")
        if valores["MAXIMO_ROLAGENS"] <= 0:
            erros.append("MAXIMO_ROLAGENS precisa ser maior que zero.")

    if erros:
        raise RuntimeError("Arquivo COORDENADAS OCR.txt invalido.", erros)

    return {
        "LINHA_X": {
            1: valores["LINHA_1_X"],
            2: valores["LINHA_2_X"],
            3: valores["LINHA_3_X"],
        },
        "LINHA_Y": {
            1: valores["LINHA_1_Y"],
            2: valores["LINHA_2_Y"],
            3: valores["LINHA_3_Y"],
        },
        "OK_X": valores["BOTAO_OK_X"],
        "OK_Y": valores["BOTAO_OK_Y"],
        "OCR_LEFT": min(valores["SUPERIOR_ESQUERDO_X"], valores["INFERIOR_ESQUERDO_X"]),
        "OCR_TOP": min(valores["SUPERIOR_ESQUERDO_Y"], valores["SUPERIOR_DIREITO_Y"]),
        "OCR_RIGHT": max(valores["SUPERIOR_DIREITO_X"], valores["INFERIOR_DIREITO_X"]),
        "OCR_BOTTOM": max(valores["INFERIOR_ESQUERDO_Y"], valores["INFERIOR_DIREITO_Y"]),
        "MAX_ROLAGENS": valores["MAXIMO_ROLAGENS"],
    }


try:
    _COORD = carregar_coordenadas_ocr()
    LINHA_X = _COORD["LINHA_X"]
    LINHA_Y = _COORD["LINHA_Y"]
    OK_X = _COORD["OK_X"]
    OK_Y = _COORD["OK_Y"]
    OCR_LEFT = _COORD["OCR_LEFT"]
    OCR_TOP = _COORD["OCR_TOP"]
    OCR_RIGHT = _COORD["OCR_RIGHT"]
    OCR_BOTTOM = _COORD["OCR_BOTTOM"]
    MAX_ROLAGENS = _COORD["MAX_ROLAGENS"]
    ALTURA_LINHA_OCR = int((OCR_BOTTOM - OCR_TOP) / 3)
    print(f"[OK] Coordenadas OCR carregadas: {os.path.join(_SCRIPT_DIR, 'COORDENADAS OCR.txt')}")
except Exception as _e_coord:
    _detalhes_coord = []
    if len(getattr(_e_coord, "args", [])) > 1 and isinstance(_e_coord.args[1], list):
        _detalhes_coord = _e_coord.args[1]
    _salvar_erro_coordenadas(str(_e_coord.args[0] if getattr(_e_coord, "args", []) else _e_coord), _detalhes_coord)
    try:
        import smtplib as _smtp_coord
        from email.mime.multipart import MIMEMultipart as _MM_coord
        from email.mime.text import MIMEText as _MT_coord
        import os as _os_coord
        _dest  = _os_coord.environ.get("RWA_EMAIL_DESTINO",   "").strip()
        _conta = _os_coord.environ.get("RWA_EMAIL_CONTA",     "").strip()
        _senha = _os_coord.environ.get("RWA_EMAIL_SENHA_APP", "").strip()
        _agend = _os_coord.environ.get("RWA_EMAIL_AGENDADO",  "0") == "1"
        if _dest and _conta and _senha and _agend:
            from datetime import datetime as _dt_coord
            _comp = f"{COMPETENCIA[:2]}/{COMPETENCIA[2:]}" if 'COMPETENCIA' in dir() else "N/A"
            _hora = _dt_coord.now().strftime('%d/%m/%Y %H:%M:%S')
            _plain = (
                f"Prezado(a),\n\nA automação não pôde ser iniciada.\n\n"
                f"  Competência : {_comp}\n  Data/Hora   : {_hora}\n"
                f"  Tipo de erro: {type(_e_coord).__name__}\n\n"
                f"Verifique o computador onde a automação está instalada.\n\nRWA Soluções"
            )
            _msg_coord = _MM_coord("alternative")
            _msg_coord["From"]    = _conta
            _msg_coord["To"]      = _dest
            _msg_coord["Subject"] = f"RWA Soluções — NFS-e Padrão Nacional não iniciada — {_comp}"
            _msg_coord.attach(_MT_coord(_plain, "plain", "utf-8"))
            _srv_coord = _smtp_coord.SMTP("smtp.gmail.com", 587)
            _srv_coord.starttls()
            _srv_coord.login(_conta, _senha)
            _srv_coord.sendmail(_conta, _dest, _msg_coord.as_string())
            _srv_coord.quit()
            print("[EMAIL] Email de nao iniciada enviado (erro OCR).")
    except Exception:
        pass
    _sys.exit(1)


class PortalInstabilidadeException(Exception):
    pass


class NomeAmigavelInvalidoException(Exception):
    pass


def detectar_nome_amigavel_invalido(tema):
    """
    Detecta se o nome amigável tem letras isoladas separadas por espaço.
    Ex: 'A C A GOMES:43025975000187' -> inválido
        'ACA GOMES:43025975000187'   -> válido
    Retorna (True, sugestao) se inválido, (False, '') se válido.
    """
    if not tema:
        return False, ""
    parte_nome = tema.split(":")[0].strip()
    if re.search(r'(?<![A-Za-z])([A-Za-z])\s(?=[A-Za-z]\s|[A-Za-z]$)', parte_nome):
        sugestao_nome = re.sub(r'\b([A-Za-z])\s+(?=[A-Za-z]\b)', r'\1', parte_nome)
        partes = tema.split(":", 1)
        sugestao = sugestao_nome + (":" + partes[1] if len(partes) > 1 else "")
        return True, sugestao
    return False, ""


class LoginInvalidoException(Exception):
    pass


class CertificadoNaoLocalizadoException(Exception):
    pass


class CertificadoInterferenciaException(Exception):
    pass


def normalizar(valor):
    if valor is None:
        return ""
    txt = str(valor).strip()
    txt = unicodedata.normalize("NFKD", txt)
    txt = "".join(c for c in txt if not unicodedata.combining(c))
    txt = re.sub(r"\s+", " ", txt)
    return txt.upper()




def normalizar_certificado(txt):
    """Normaliza certificado/OCR removendo acento, espaço, pontuação e ruído visual."""
    if not txt:
        return ""
    txt = str(txt).upper()
    txt = txt.replace("€", "C")
    txt = txt.replace("¢", "C")
    txt = txt.replace("£", "E")
    txt = txt.replace("|", "I")
    txt = unicodedata.normalize("NFD", txt)
    txt = "".join(c for c in txt if unicodedata.category(c) != "Mn")
    txt = re.sub(r"[^A-Z0-9]", "", txt)
    return txt


def parece_certificado_correto(txt_ocr, tema_busca, minimo=0.75):
    """
    Comparação tolerante para certificado no popup nativo do Windows.

    Ordem correta dos parâmetros:
    - txt_ocr: texto lido pelo OCR na linha visível do popup
    - tema_busca: primeiros caracteres do nome amigável vindo da planilha

    Aceita variações comuns do OCR:
    'A C A GOMES', 'ACA GOMES', 'AC A GOMES', 'A CA GOMES'.
    """
    ocr_norm = normalizar_certificado(txt_ocr)
    busca_norm = normalizar_certificado(tema_busca)

    if not ocr_norm or not busca_norm:
        return False

    # Caminho forte: o tema limpo aparece dentro do OCR limpo.
    # Proteção: nunca aceitar leitura muito curta, como "O", "CA", "AC".
    # Isso evita falso positivo quando o OCR lê apenas uma letra que existe no nome da empresa.
    if len(ocr_norm) < 8 or len(busca_norm) < 8:
        return False

    if busca_norm in ocr_norm:
        return True

    if ocr_norm in busca_norm and len(ocr_norm) >= 12:
        return True

    # Caminho tolerante: OCR com pequeno ruído, letra sobrando ou espaço falho.
    score = SequenceMatcher(None, busca_norm, ocr_norm).ratio()
    return score >= minimo


def limpar_nome(nome):
    nome = str(nome).strip()
    nome = re.sub(r'[\\/:*?"<>|]', '', nome)
    return nome


def _driver_comunicacao_quebrada(msg):
    """Detecta quando o Selenium/ChromeDriver perdeu comunicação e não é seguro usar driver para diagnóstico."""
    txt = normalizar(str(msg or ""))
    sinais = [
        "HTTPCONNECTIONPOOL",
        "LOCALHOST",
        "MAX RETRIES EXCEEDED",
        "CONNECTION REFUSED",
        "CONNECTION RESET",
        "READ TIMED OUT",
        "CHROMEDRIVER",
        "DISCONNECTED",
        "NO SUCH WINDOW",
        "TARGET WINDOW ALREADY CLOSED",
        "INVALID SESSION ID",
        "RECEIVING MESSAGE FROM RENDERER",
        "GETHANDLEVERIFIER",
        "POS-LOGIN POR CERTIFICADO",
        "PÓS-LOGIN POR CERTIFICADO",
    ]
    return any(s in txt for s in sinais)


def _diagnostico_driver_seguro(driver, msg):
    """Retorna None quando o driver está quebrado, evitando travar screenshot/DOM."""
    if _driver_comunicacao_quebrada(msg):
        return None
    return driver



def esperar_site_estavel(driver, wait=None, timeout=35, descricao="site"):
    inicio = time.time()
    ultimo_erro = None

    while time.time() - inicio < timeout:
        try:
            estado = driver.execute_script("return document.readyState")
            body = driver.find_element(By.TAG_NAME, "body")
            texto = (body.text or "").strip()
            texto_norm = normalizar(texto)

            instavel = any(t in texto_norm for t in [
                "BAD GATEWAY", "SERVICE UNAVAILABLE", "GATEWAY TIMEOUT",
                "ERRO INTERNO", "SERVICO INDISPONIVEL", "SERVIÇO INDISPONÍVEL",
                "NAO FOI POSSIVEL CARREGAR", "NÃO FOI POSSÍVEL CARREGAR",
                "NAO FOI POSSIVEL EXIBIR", "NÃO FOI POSSÍVEL EXIBIR",
            ])

            if estado == "complete" and len(texto) > 20 and not instavel:
                return True

            ultimo_erro = f"readyState={estado}, texto={len(texto)}, instavel={instavel}"
        except Exception as e:
            ultimo_erro = str(e)

        time.sleep(0.5)

    raise Exception(f"Tempo excedido aguardando estabilidade do {descricao}. Último estado: {ultimo_erro}")


def esperar_clicavel_seguro(driver, by, seletor, timeout=35, descricao="elemento clicável"):
    wait_local = WebDriverWait(driver, timeout)
    esperar_site_estavel(driver, wait_local, timeout=timeout, descricao=descricao)
    return wait_local.until(EC.element_to_be_clickable((by, seletor)))

def pasta_empresa(codigo, nome, pasta_raiz=None):
    raiz = pasta_raiz if pasta_raiz else PASTA_BASE
    nome_limpo = limpar_nome(nome)
    codigo_txt = str(codigo).strip()
    nome_upper = nome_limpo.upper()

    if os.path.isdir(raiz):
        for item in os.listdir(raiz):
            caminho = os.path.join(raiz, item)
            if os.path.isdir(caminho):
                item_upper = item.upper()
                # Valida código E nome juntos
                prefixo_ok = item_upper.startswith(f"{codigo_txt}-") or item_upper.startswith(f"{codigo_txt} ")
                nome_ok = nome_upper[:15] in item_upper
                if prefixo_ok and nome_ok:
                    pasta = os.path.join(caminho, COMPETENCIA)
                    os.makedirs(pasta, exist_ok=True)
                    return pasta

    nome_pasta = f"{codigo_txt}-{nome_limpo}"
    caminho_empresa = os.path.join(raiz, nome_pasta)
    pasta = os.path.join(caminho_empresa, COMPETENCIA)
    os.makedirs(pasta, exist_ok=True)
    return pasta


def pasta_interna(pasta_mes, codigo):
    """Retorna (e cria) a subpasta interna {codigo}_{COMPETENCIA} dentro da pasta do mes.
    Somente o XML/ZIP prestado fica na raiz. Tudo mais vai aqui."""
    caminho = os.path.join(pasta_mes, f"{codigo}_{COMPETENCIA}")
    os.makedirs(caminho, exist_ok=True)
    return caminho


def nome_arquivo_do_header(content_disposition):
    if not content_disposition:
        return None

    match = re.search(r"filename\*=UTF-8''([^;]+)", content_disposition, re.IGNORECASE)
    if match:
        return unquote(match.group(1).strip().strip('"'))

    match = re.search(r'filename="?([^";]+)"?', content_disposition, re.IGNORECASE)
    if match:
        return unquote(match.group(1).strip().strip('"'))

    return None


def criar_sessao_logada(driver):
    sessao = requests.Session()

    for cookie in driver.get_cookies():
        sessao.cookies.set(
            cookie.get("name"),
            cookie.get("value"),
            domain=cookie.get("domain"),
            path=cookie.get("path", "/")
        )

    headers = {
        "User-Agent": driver.execute_script("return navigator.userAgent;"),
        "Referer": driver.current_url,
        "Accept": "*/*",
    }

    return sessao, headers



def baixar_conteudo_direto(driver, url, pasta, nome_debug="DOWNLOAD", timeout=60):
    url_absoluta = urljoin(URL_BASE, url)

    print(f"Baixando direto: {url_absoluta}")

    for _tentativa in range(1, 4):  # até 3 tentativas
        # Renova cookies a cada tentativa (resolve sessão expirada)
        sessao, headers = criar_sessao_logada(driver)
        resposta = sessao.get(url_absoluta, headers=headers, allow_redirects=True, timeout=timeout)

        # 503 — servidor indisponível: aguarda e retenta
        if resposta.status_code == 503:
            if _tentativa < 3:
                print(f"  [AVISO] HTTP 503 — tentativa {_tentativa}/3 — aguardando 10s...")
                time.sleep(10)
                continue
            raise Exception(f"Download direto falhou. Status HTTP: 503 (3 tentativas)")

        if resposta.status_code != 200:
            raise Exception(f"Download direto falhou. Status HTTP: {resposta.status_code}")

        conteudo = resposta.content or b""

        if len(conteudo) < 50:
            raise Exception("Download direto retornou arquivo vazio ou muito pequeno.")

        texto_inicio = conteudo[:300].lower()
        if b"<html" in texto_inicio or b"<!doctype html" in texto_inicio:
            if _tentativa < 3:
                print(f"  [AVISO] Portal retornou HTML (sessao expirada) — tentativa {_tentativa}/3 — renovando sessao...")
                time.sleep(5)
                continue
            debug_nome = f"ERRO_RETORNOU_HTML_{nome_debug}.html"
            caminho_debug = os.path.join(pasta, limpar_nome(debug_nome))
            with open(caminho_debug, "wb") as f:
                f.write(conteudo)
            raise Exception(f"Retornou HTML em vez de arquivo. Debug salvo: {caminho_debug}")

        return conteudo  # sucesso

    raise Exception("Download direto falhou após 3 tentativas.")


def salvar_conteudo_arquivo(conteudo, pasta, extensao, nome_padrao):
    nome_arquivo = limpar_nome(nome_padrao)

    if not nome_arquivo.lower().endswith(extensao.lower()):
        nome_arquivo += extensao

    caminho_final = os.path.join(pasta, nome_arquivo)

    if os.path.exists(caminho_final):
        os.remove(caminho_final)

    with open(caminho_final, "wb") as f:
        f.write(conteudo)

    print(f"OK - arquivo salvo: {caminho_final}")
    return caminho_final


def extrair_numero_nfse_do_xml(conteudo_xml):
    try:
        texto = conteudo_xml.decode("utf-8-sig", errors="ignore")
    except Exception:
        texto = str(conteudo_xml)

    tags_prioritarias = [
        "nNFSe", "Numero", "NumeroNfse", "NumeroNFSe", "NumeroNota",
        "numero", "numeroNfse", "numeroNFSe", "numeroNota"
    ]

    for tag in tags_prioritarias:
        padrao = rf"<[^>/]*{re.escape(tag)}[^>]*>\s*([0-9]+)\s*</[^>]+>"
        m = re.search(padrao, texto, re.IGNORECASE)
        if m:
            return str(int(m.group(1)))

    try:
        raiz = ET.fromstring(texto.encode("utf-8"))
        prioridade_norm = {normalizar(t) for t in tags_prioritarias}
        for el in raiz.iter():
            local = el.tag.split("}")[-1] if "}" in el.tag else el.tag
            valor = (el.text or "").strip()
            if normalizar(local) in prioridade_norm and valor.isdigit():
                return str(int(valor))
    except Exception:
        pass

    m = re.search(r'Id\s*=\s*["\']NFSe([0-9]+)["\']', texto, re.IGNORECASE)
    if m:
        digitos = m.group(1)
        return str(int(digitos[-15:])) if len(digitos) > 15 else str(int(digitos))

    return None


def extrair_dados_xml_bytes(conteudo_xml, numero_fallback="", status=""):
    """
    Extrai todos os campos do detalhamento a partir dos bytes do XML,
    enquanto ainda está na memória — antes de salvar e antes de zipar.
    Usa regex igual ao extrair_numero_nfse_do_xml, sem depender de namespace.
    """
    NS = "http://www.sped.fazenda.gov.br/nfse"

    def _n(tag): return f"{{{NS}}}{tag}"

    try:
        texto = conteudo_xml.decode("utf-8-sig", errors="ignore")
    except Exception:
        texto = str(conteudo_xml)

    dados = {
        "numero": numero_fallback,
        "cnpj": "", "cliente": "", "emissao": "",
        "cnae": "", "valor_nfse": 0.0, "deducao": 0.0, "v_desc_incond": 0.0, "valor_liq": 0.0,
        "valor_servico": 0.0, "iss": "", "aliquota": 0.0, "valor_iss": 0.0,
        "tp_ret_issqn": "", "base_calculo": 0.0, "chave": "",
        "status": status,
        # Tributação federal
        "op_simples":      "",
        "v_ret_cp":        0.0,
        "v_ret_irrf":      0.0,
        "v_ret_csll":      0.0,
        "v_bc_pis_cofins": 0.0,
        "p_aliq_pis":      0.0,
        "v_pis":           0.0,
        "p_aliq_cofins":   0.0,
        "v_cofins":        0.0,
        # Emitente (prestador) — usado na aba TOMADO
        "emit_cnpj": "",
        "emit_nome": "",
    }

    try:
        raiz = ET.fromstring(texto.encode("utf-8"))

        def _txt(el): return (el.text or "").strip() if el is not None else ""

        # NUMERO
        el = raiz.find(f".//{_n('nNFSe')}")
        if el is not None and _txt(el).isdigit():
            dados["numero"] = str(int(_txt(el)))

        # CHAVE — Id do infNFSe, só os dígitos após "NFS"
        inf_nfse = raiz.find(_n("infNFSe"))
        if inf_nfse is not None:
            id_val = inf_nfse.get("Id", "")
            m_chave = re.match(r"NFS(\d+)", id_val)
            if m_chave:
                dados["chave"] = m_chave.group(1)

        # EMISSAO — dhProc ou dhEmi, só a data dd.mm.aaaa
        for tag in ["dhProc", "dhEmi"]:
            el = raiz.find(f".//{_n(tag)}")
            if el is not None and _txt(el):
                raw = _txt(el)[:10]
                try:
                    from datetime import datetime as _dt
                    dados["emissao"] = _dt.strptime(raw, "%Y-%m-%d").strftime("%d.%m.%Y")
                except Exception:
                    dados["emissao"] = raw
                break

        # TOMADOR — CPF ou CNPJ + xNome dentro de <toma>
        toma = raiz.find(f".//{_n('toma')}")
        if toma is not None:
            cpf  = toma.find(_n("CPF"))
            cnpj = toma.find(_n("CNPJ"))
            dados["cnpj"]    = _txt(cpf) if cpf is not None else _txt(cnpj)
            dados["cliente"] = _txt(toma.find(_n("xNome")))

        # EMITENTE — CNPJ ou CPF + xNome dentro de <emit>
        # (usado na aba TOMADO e na coluna B "Inscricao" do gerencial)
        emit = raiz.find(f".//{_n('emit')}")
        if emit is not None:
            _cnpj_el = emit.find(_n("CNPJ"))
            _cpf_el  = emit.find(_n("CPF"))
            _cnpj_txt = _txt(_cnpj_el) if _cnpj_el is not None else ""
            _cpf_txt  = _txt(_cpf_el)  if _cpf_el  is not None else ""
            dados["emit_cnpj"] = _cnpj_txt if _cnpj_txt else _cpf_txt
            dados["emit_nome"] = _txt(emit.find(_n("xNome")))

        # CNAE — cTribNac: só o número inicial
        el_cnae = raiz.find(f".//{_n('cTribNac')}")
        if el_cnae is None:
            el_cnae = raiz.find(f".//{_n('xTribNac')}")
        if el_cnae is not None:
            m = re.match(r"(\d+)", _txt(el_cnae))
            dados["cnae"] = m.group(1) if m else _txt(el_cnae)

        # vDescIncond — desconto incondicionado
        v_desc_el = raiz.find(f".//{_n('vDescIncond')}")
        if v_desc_el is not None and _txt(v_desc_el):
            dados["v_desc_incond"] = float(_txt(v_desc_el))

        # vServ — DPS/infDPS/valores/vServPrest/vServ
        # G (valor_nfse) e I (valor_servico) ambos vem de <vServ>
        v_serv_el = raiz.find(f".//{_n('vServPrest')}/{_n('vServ')}")
        if v_serv_el is not None and _txt(v_serv_el):
            dados["valor_servico"] = float(_txt(v_serv_el))
            dados["valor_nfse"]    = float(_txt(v_serv_el))  # G = <vServ>

        # VALORES bloco infNFSe — vBC, vLiq, pAliqAplic, vISSQN
        inf    = raiz.find(_n("infNFSe")) or raiz
        valores = inf.find(_n("valores"))
        if valores is not None:
            vbc   = valores.find(_n("vBC"))
            vliq  = valores.find(_n("vLiq"))
            paliq = valores.find(_n("pAliqAplic"))
            viss  = valores.find(_n("vISSQN"))
            # K (base_calculo) = <vBC>
            if vbc is not None and _txt(vbc):
                dados["base_calculo"]  = float(_txt(vbc))
            # vLiq (valor liquido)
            if vliq is not None and _txt(vliq):
                dados["valor_liq"] = float(_txt(vliq))
            if paliq is not None and _txt(paliq): dados["aliquota"]  = float(_txt(paliq))
            if viss  is not None and _txt(viss):  dados["valor_iss"] = float(_txt(viss))

        # ISS — tribISSQN: 1=NORMAL / 2=RETIDO
        trib = raiz.find(f".//{_n('tribISSQN')}")
        if trib is not None:
            dados["iss"] = "NORMAL" if _txt(trib) == "1" else "RETIDO"

        # tpRetISSQN — 1=Não Retido / 2=Retido
        tp_ret = raiz.find(f".//{_n('tpRetISSQN')}")
        if tp_ret is not None:
            dados["tp_ret_issqn"] = "Não Retido" if _txt(tp_ret) == "1" else "Retido"

        # opSimpNac — regime Simples Nacional
        op_sn = raiz.find(f".//{_n('opSimpNac')}")
        if op_sn is not None:
            dados["op_simples"] = _txt(op_sn)

        # tribFed — retenções federais
        trib_fed = raiz.find(f".//{_n('tribFed')}")
        if trib_fed is not None:
            def _f(tag):
                el = trib_fed.find(_n(tag))
                return float(_txt(el)) if el is not None and _txt(el) else 0.0
            dados["v_ret_cp"]   = _f("vRetCP")
            dados["v_ret_irrf"] = _f("vRetIRRF")
            dados["v_ret_csll"] = _f("vRetCSLL")
            piscofins = trib_fed.find(_n("piscofins"))
            if piscofins is not None:
                def _fp(tag):
                    el = piscofins.find(_n(tag))
                    return float(_txt(el)) if el is not None and _txt(el) else 0.0
                dados["v_bc_pis_cofins"] = _fp("vBCPisCofins")
                dados["p_aliq_pis"]      = _fp("pAliqPis")
                dados["v_pis"]           = _fp("vPis")
                dados["p_aliq_cofins"]   = _fp("pAliqCofins")
                dados["v_cofins"]        = _fp("vCofins")

    except Exception:
        pass

    dados["v_desc_incond"]  = round(dados.get("v_desc_incond", 0.0), 2)
    dados["valor_nfse"]    = round(dados["valor_nfse"], 2)
    dados["valor_liq"]     = round(dados["valor_liq"], 2)
    dados["valor_servico"] = round(dados["valor_servico"], 2)
    dados["valor_iss"]     = round(dados["valor_iss"], 2)
    dados["base_calculo"]  = round(dados["base_calculo"], 2)
    dados["v_ret_cp"]        = round(dados["v_ret_cp"],        2)
    dados["v_ret_irrf"]      = round(dados["v_ret_irrf"],      2)
    dados["v_ret_csll"]      = round(dados["v_ret_csll"],      2)
    dados["v_bc_pis_cofins"] = round(dados["v_bc_pis_cofins"], 2)
    dados["p_aliq_pis"]      = round(dados["p_aliq_pis"],      4)
    dados["v_pis"]           = round(dados["v_pis"],           2)
    dados["p_aliq_cofins"]   = round(dados["p_aliq_cofins"],   4)
    dados["v_cofins"]        = round(dados["v_cofins"],        2)
    return dados


def baixar_arquivo_direto(driver, url, pasta, extensao, nome_padrao, timeout=60):
    conteudo = baixar_conteudo_direto(driver, url, pasta, nome_debug=nome_padrao, timeout=timeout)
    return salvar_conteudo_arquivo(conteudo, pasta, extensao, nome_padrao)

def contar_arquivos(pasta, extensao):
    return len([
        f for f in os.listdir(pasta)
        if f.lower().endswith(extensao.lower())
        and not f.lower().endswith(".crdownload")
        and not f.lower().endswith(".tmp")
    ])


def obter_total_registros(driver):
    texto = driver.find_element(By.TAG_NAME, "body").text
    match = re.search(r"Total\s+de\s+(\d+)\s+registros?", texto, re.IGNORECASE)
    if match:
        return int(match.group(1))

    html = driver.page_source
    match = re.search(r"Total\s+de\s+(\d+)\s+registros?", html, re.IGNORECASE)
    if match:
        return int(match.group(1))

    return None


def salvar_print_tela(driver, pasta, codigo, nome):
    nome_limpo = limpar_nome(nome)
    caminho_print = os.path.join(
        pasta,
        f"PRINT_TOTAL_REGISTROS_{codigo}_{nome_limpo}_{COMPETENCIA}.png"
    )

    try:
        try:
            driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
            time.sleep(1.2)

            total_width = driver.execute_script(
                "return Math.max(document.body.scrollWidth, document.documentElement.scrollWidth);"
            )
            total_height = driver.execute_script(
                "return Math.max(document.body.scrollHeight, document.documentElement.scrollHeight);"
            )

            total_width = max(1366, int(total_width or 1366))
            total_height = max(768, min(int(total_height or 768), 12000))

            driver.set_window_size(total_width, total_height)
            time.sleep(1.2)
            driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
            time.sleep(0.8)
        except Exception:
            pass

        driver.save_screenshot(caminho_print)
        print(f"OK - print salvo: {caminho_print}")

        try:
            driver.maximize_window()
            driver.execute_script("window.scrollTo(0, 0);")
            time.sleep(0.8)
        except Exception:
            pass

        return caminho_print
    except Exception as e:
        print(f"AVISO - não conseguiu salvar print: {e}")
        return ""


def esperar_tabela_atualizar(driver, texto_antigo=None, timeout=20):
    inicio = time.time()
    while time.time() - inicio < timeout:
        try:
            linhas = driver.find_elements(By.XPATH, "//table//tbody//tr")
            texto_atual = driver.find_element(By.TAG_NAME, "body").text
            if linhas and (texto_antigo is None or texto_atual != texto_antigo):
                return True
        except Exception:
            pass
        time.sleep(0.5)
    return False


# FUNCAO DA V18E VALIDADA - NAO RECRIADA
def capturar_links_menu(driver, wait, linha):
    try:
        driver.find_element(By.TAG_NAME, "body").send_keys(Keys.ESCAPE)
        time.sleep(0.3)
    except Exception:
        pass

    botao_opcoes = linha.find_element(By.CLASS_NAME, "icone-trigger")
    driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", botao_opcoes)
    time.sleep(0.3)
    driver.execute_script("arguments[0].click();", botao_opcoes)
    time.sleep(1.2)

    wait.until(
        EC.presence_of_element_located(
            (By.XPATH, "//a[contains(@href, '/EmissorNacional/Notas/Download/')]")
        )
    )

    links_visiveis = driver.execute_script("""
        const anchors = Array.from(document.querySelectorAll('a[href*="/EmissorNacional/Notas/Download/"]'));
        return anchors
            .filter(a => {
                const r = a.getBoundingClientRect();
                const s = window.getComputedStyle(a);
                return r.width > 0 && r.height > 0 && s.visibility !== 'hidden' && s.display !== 'none';
            })
            .map(a => a.href);
    """)

    url_xml = None
    url_pdf = None

    for href in links_visiveis:
        if "/Notas/Download/NFSe/" in href:
            url_xml = href
        if "/Notas/Download/DANFSe/" in href:
            url_pdf = href

    if not url_xml or not url_pdf:
        todos_links = driver.execute_script("""
            return Array.from(document.querySelectorAll('a[href*="/EmissorNacional/Notas/Download/"]'))
                .map(a => a.href);
        """)

        for href in reversed(todos_links):
            if not url_xml and "/Notas/Download/NFSe/" in href:
                url_xml = href
            if not url_pdf and "/Notas/Download/DANFSe/" in href:
                url_pdf = href

    if not url_xml:
        raise Exception("Não consegui capturar o link XML da linha.")

    if not url_pdf:
        raise Exception("Não consegui capturar o link PDF da linha.")

    print(f"OK - capturou link XML: {url_xml}")
    print(f"OK - capturou link PDF: {url_pdf}")

    return url_xml, url_pdf


# FUNCAO DA V18E VALIDADA - NAO RECRIADA
def ir_para_proxima_pagina(driver, pagina_atual):
    proxima = pagina_atual + 1
    texto_antes = driver.find_element(By.TAG_NAME, "body").text

    # Busca pelo href exato pg=N — evita ambiguidade com outros botões fa-angle-right
    xpaths = [
        f"//a[contains(@href,'pg={proxima}')]",
        f"//a[normalize-space()='{proxima}']",
        f"//button[normalize-space()='{proxima}']",
        f"//*[self::a or self::button or self::span][normalize-space()='{proxima}' and not(contains(@class,'disabled'))]",
    ]

    for xp in xpaths:
        try:
            candidatos = driver.find_elements(By.XPATH, xp)
            for el in candidatos:
                classe = (el.get_attribute("class") or "").lower()
                if el.is_displayed() and el.is_enabled() and "disabled" not in classe:
                    driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", el)
                    time.sleep(0.3)
                    driver.execute_script("arguments[0].click();", el)
                    print(f"OK - indo para página {proxima}")
                    esperar_tabela_atualizar(driver, texto_antigo=texto_antes, timeout=20)
                    time.sleep(2)
                    return True
        except Exception:
            pass

    return False


# =====================================================================
# CERTIFICADO DIGITAL - BASE V14 VALIDADA
# =====================================================================

def detectar_lang():
    if os.path.exists(r"C:\Program Files\Tesseract-OCR\tessdata\por.traineddata"):
        return "por"
    return "eng"


def ocr_linha(img_linha, lang):
    try:
        img_linha = img_linha.convert("L")
        img_linha = ImageOps.invert(img_linha)
        img_linha = ImageEnhance.Contrast(img_linha).enhance(2.0)
        w, h = img_linha.size
        img_linha = img_linha.resize((w * 2, h * 2))
        txt = pytesseract.image_to_string(img_linha, lang=lang, config="--psm 7")
        return normalizar(re.sub(r"\s+", " ", txt).strip())
    except Exception:
        return ""


def capturar_linhas_ocr(lang):
    resultado = {1: "", 2: "", 3: ""}
    try:
        img_full = ImageGrab.grab(bbox=(OCR_LEFT, OCR_TOP, OCR_RIGHT, OCR_BOTTOM))
        for n in range(1, 4):
            y0 = (n - 1) * ALTURA_LINHA_OCR
            y1 = n * ALTURA_LINHA_OCR
            img_l = img_full.crop((0, y0, img_full.width, y1))
            resultado[n] = ocr_linha(img_l, lang)
    except Exception as e:
        print(f"[OCR] Erro: {e}")
    return resultado



def _assinatura_tela_certificado():
    """Captura uma assinatura leve da área onde o popup/lista de certificados aparece."""
    try:
        img = ImageGrab.grab(bbox=(300, 170, 900, 430)).convert("L")
        img = img.resize((60, 26))
        if hasattr(img, "get_flattened_data"):
            return list(img.get_flattened_data())
        return list(img.getdata())
    except Exception:
        return None


def _diferenca_assinatura(a, b):
    if not a or not b or len(a) != len(b):
        return 0
    return sum(abs(x - y) for x, y in zip(a, b)) / len(a)


def _mouse_dentro_area_ocr(margem=5):
    try:
        x, y = pyautogui.position()
        return (OCR_LEFT - margem) <= x <= (OCR_RIGHT + margem) and (OCR_TOP - margem) <= y <= (OCR_BOTTOM + margem)
    except Exception:
        return False


def _afastar_mouse_area_ocr():
    """Evita que o próprio cursor da automação fique cobrindo a área lida pelo OCR."""
    try:
        x = OCR_RIGHT + 35
        y = OCR_BOTTOM + 35
        pyautogui.moveTo(x, y, duration=0)
        time.sleep(0.15)
    except Exception:
        pass


def _fechar_popup_certificado():
    """Fecha a janela nativa de certificado sem esperar timeout.

    O popup de certificado é nativo do Windows. Quando há interferência,
    o foco pode estar no popup, no Chrome, no CMD ou em outra janela.
    Por isso usamos ESC repetido e um pequeno intervalo, sem depender do Selenium.
    """
    try:
        for _ in range(4):
            pyautogui.press("esc")
            time.sleep(0.15)
    except Exception:
        pass


def _linha_indica_contexto_perdido(txt):
    """Detecta quando o OCR começou a ler a tela errada, e não a lista de certificados."""
    t = normalizar(txt)
    if not t:
        return False

    sinais_tela_errada = [
        "NFS-E PADRAO NACIONAL", "NFSE PADRAO NACIONAL", "PADRAO NACIONAL",
        "AUTOMACOES", "AUTOMAÇÕES", "COORDENADAS", "MEU DRIVE", "GOOGLE DRIVE",
        "ARQUIVO", "PASTA", "DOWNLOAD", "IMPORTACAO", "IMPORTAÇÃO",
        "RWA SOLUCOES", "RWA SOLUÇÕES", "EMISSOR NACIONAL",
        "CHROME", "GOOGLE", "WINDOWS", "EXPLORADOR DE ARQUIVOS",
        ".PY", "PY:", "DEPRECAT", "WARNING", "PROGRAM FILES", "WINDOWSAP",
        "TRUTH VALUE", "LEN(ELEM)", "RAIZ.FIND", "INFNFSE", "CMD", "POWERSHELL",
        "C:\\", "G:\\", "TESSERACT", "SELENIUM", "PYTHON",
    ]

    return any(sinal in t for sinal in sinais_tela_errada)


def _linha_parece_certificado(txt):
    """Valida se uma linha OCR ainda parece uma linha real da lista de certificados."""
    t = normalizar(txt)
    compacto = normalizar_certificado(txt)

    if not compacto or len(compacto) < 16:
        return False

    if _linha_indica_contexto_perdido(txt):
        return False

    # Lista de certificados normalmente contém o titular + AC/autoridade + serial.
    tem_ac_visual = " AC " in f" {t} " or compacto.startswith("AC") or "ACSAFEWEB" in compacto or "ACSOLUCAO" in compacto
    tem_numero_forte = bool(re.search(r"\d{8,14}", compacto))
    tem_autoridade = any(chave in compacto for chave in [
        "SAFEWEB", "SOLUCAO", "SOLUÇÃO", "SYNGULAR", "SERASA", "CERTISIGN",
        "VALID", "SOLUTI", "RFB", "MULTIPLA", "MULT", "DIGITAL", "ICPBRASIL",
        "ONLINE", "CERTIFICADORA",
    ])
    tem_serial = bool(re.search(r"[A-F0-9]{12,}", compacto))

    return tem_ac_visual and (tem_numero_forte or tem_autoridade or tem_serial)


def _qualidade_linhas_ocr(linhas):
    """Mede se a leitura continua dentro da lista real de certificados."""
    linhas = linhas or {}
    textos = [str(txt or "") for txt in linhas.values()]
    normalizadas = [normalizar_certificado(txt) for txt in textos]
    validas = [txt for txt in textos if _linha_parece_certificado(txt)]
    contexto_perdido = [txt for txt in textos if _linha_indica_contexto_perdido(txt)]
    vazias_ou_ruido = [txt for txt in normalizadas if len(txt) < 4]

    return {
        "validas": len(validas),
        "contexto_perdido": len(contexto_perdido),
        "ruido": len(vazias_ou_ruido),
        "texto_total": sum(len(txt) for txt in normalizadas),
        "amostra_contexto": contexto_perdido[0] if contexto_perdido else "",
    }


def _validar_leitura_certificado(linhas, assinatura_anterior=None, etapa="OCR_CERTIFICADO"):
    """
    Validação leve durante a leitura.

    Importante: esta função NÃO decide mais, no meio da busca, que o certificado
    não existe ou que houve interferência só porque uma rolagem veio ruim.
    A decisão principal é feita no FINAL, analisando o histórico completo do OCR.

    Aqui só mata imediatamente quando o mouse entra exatamente na área OCR,
    porque isso é interferência física direta e comprovada.
    """
    if _mouse_dentro_area_ocr():
        _fechar_popup_certificado()
        raise CertificadoInterferenciaException(
            f"CERT_INTERFERENCIA_HUMANA: mouse dentro da área OCR durante {etapa}."
        )

    if assinatura_anterior is not None:
        return _assinatura_tela_certificado()

    return _assinatura_tela_certificado()


def _analisar_historico_ocr_certificado(historico_ocr):
    """
    Decide, ao FINAL da busca, se a automação realmente percorreu a lista
    de certificados ou se se perdeu por interferência/janela sobreposta.

    Regra operacional:
    - Se leu linhas com cara real de certificado durante o percurso, e não leu
      tela errada, então é certificado não localizado.
    - Se leu CMD, .py, caminho, automação, Chrome/Excel/Windows ou texto fora
      do popup, então é possível interferência e a empresa deve ir para o fim da fila.
    - Se nunca leu certificado real, também trata como interferência/ambiente,
      porque as coordenadas provavelmente estavam sobre a tela errada.
    """
    total_linhas = 0
    linhas_certificado = 0
    linhas_contexto_perdido = 0
    linhas_legiveis_fora_contexto = 0
    amostra_contexto = ""
    amostra_fora_contexto = ""

    for linhas in historico_ocr or []:
        for txt in (linhas or {}).values():
            total_linhas += 1
            bruto = str(txt or "").strip()
            compacto = normalizar_certificado(bruto)

            if not bruto:
                continue

            if _linha_indica_contexto_perdido(bruto):
                linhas_contexto_perdido += 1
                if not amostra_contexto:
                    amostra_contexto = bruto
                continue

            if _linha_parece_certificado(bruto):
                linhas_certificado += 1
                continue

            # Texto grande, legível, mas sem cara de certificado.
            # Uma linha isolada pode ser OCR ruim de certificado; várias linhas indicam tela errada.
            if len(compacto) >= 18:
                linhas_legiveis_fora_contexto += 1
                if not amostra_fora_contexto:
                    amostra_fora_contexto = bruto

    if linhas_contexto_perdido > 0:
        return {
            "interferencia": True,
            "motivo": "OCR leu texto fora da lista de certificados",
            "amostra": amostra_contexto[:160],
            "certificados_lidos": linhas_certificado,
            "fora_contexto": linhas_contexto_perdido,
        }

    if linhas_certificado == 0:
        return {
            "interferencia": True,
            "motivo": "OCR não leu nenhuma linha com aparência de certificado",
            "amostra": amostra_fora_contexto[:160],
            "certificados_lidos": linhas_certificado,
            "fora_contexto": linhas_legiveis_fora_contexto,
        }

    # Se houve muita leitura legível que não parece certificado, mas também houve certificado,
    # considera interferência apenas quando o volume estranho for relevante.
    if linhas_legiveis_fora_contexto >= 6 and linhas_legiveis_fora_contexto > linhas_certificado:
        return {
            "interferencia": True,
            "motivo": "OCR leu muitas linhas legíveis sem aparência de certificado",
            "amostra": amostra_fora_contexto[:160],
            "certificados_lidos": linhas_certificado,
            "fora_contexto": linhas_legiveis_fora_contexto,
        }

    return {
        "interferencia": False,
        "motivo": "Histórico OCR percorreu lista real de certificados",
        "amostra": "",
        "certificados_lidos": linhas_certificado,
        "fora_contexto": linhas_legiveis_fora_contexto,
    }


def _finalizar_busca_certificado_nao_encontrado(historico_ocr, motivo="certificado não localizado"):
    analise = _analisar_historico_ocr_certificado(historico_ocr)

    print(
        "[CERT] Análise final OCR: "
        f"interferencia={analise.get('interferencia')} | "
        f"motivo={analise.get('motivo')} | "
        f"certificados_lidos={analise.get('certificados_lidos')} | "
        f"fora_contexto={analise.get('fora_contexto')} | "
        f"amostra={analise.get('amostra')}"
    )

    _fechar_popup_certificado()

    if analise.get("interferencia"):
        raise CertificadoInterferenciaException(
            "POSSÍVEL INTERFERÊNCIA OCR: "
            f"{analise.get('motivo')} | "
            f"certificados_lidos={analise.get('certificados_lidos')} | "
            f"fora_contexto={analise.get('fora_contexto')} | "
            f"amostra={analise.get('amostra')}"
        )

    print(
        "[CERT] Histórico OCR válido: lista real de certificados percorrida — "
        f"certificados_lidos={analise.get('certificados_lidos')}."
    )
    print(f"[CERT] {motivo} — tratando como certificado não localizado.")
    return False


def aguardar_popup_certificado_inteligente(timeout=25):
    """
    Espera inteligente do popup de certificado.

    REGRA OPERACIONAL DEFINITIVA:
    Se aparecer a mensagem "[CERT] Popup detectado por mudança de tela",
    a tentativa deve ser interrompida imediatamente e a empresa deve ir
    para o tratamento de reprocessamento/final da fila.

    O valor do diff fica apenas como diagnóstico no log.
    Não existe mais regra baseada em diff fixo.
    """
    print("[CERT] Aguardando popup do certificado aparecer...")
    base = _assinatura_tela_certificado()
    inicio = time.time()

    while time.time() - inicio < timeout:
        atual = _assinatura_tela_certificado()
        diff = _diferenca_assinatura(base, atual)

        if diff > 6:
            time.sleep(0.3)
            print(f"[CERT] Popup detectado por mudança de tela. diff={diff:.2f}")
            return True

        time.sleep(0.25)

    raise Exception("Popup de certificado não apareceu dentro do tempo limite.")

def thread_selecionar_certificado(tema, lang):
    tema_norm = normalizar(tema)
    tema_busca = tema_norm[:15].strip()

    aguardar_popup_certificado_inteligente(timeout=25)

    # FOCO DA LISTA — sequência original V5 restaurada.
    # Popup nativo do Windows: Selenium não enxerga. O clique físico foca a lista,
    # CTRL+HOME volta para o topo, e o segundo clique garante foco antes do loop.
    time.sleep(0.5)
    pyautogui.click(LINHA_X[1], LINHA_Y[1])
    time.sleep(0.5)
    pyautogui.hotkey("ctrl", "Home")
    time.sleep(0.5)
    pyautogui.click(LINHA_X[1], LINHA_Y[1])
    time.sleep(0.3)
    _afastar_mouse_area_ocr()
    _mouse_seguro_x, _mouse_seguro_y = pyautogui.position()

    # Verificação inicial: confirma que as coordenadas continuam lendo a lista.
    # Se alguém mexer no mouse/janela ou o AnyDesk alterar foco/renderização,
    # isso vira interferência e a empresa pode ir para a fila, sem confundir com certificado inexistente.
    assinatura_anterior = _assinatura_tela_certificado()
    historico_ocr = []
    if abs(pyautogui.position()[0] - _mouse_seguro_x) > 8 or abs(pyautogui.position()[1] - _mouse_seguro_y) > 8:
        _fechar_popup_certificado()
        raise CertificadoInterferenciaException(
            "CERT_INTERFERENCIA_HUMANA: mouse movimentado após abertura do popup do certificado."
        )

    leitura_inicial = capturar_linhas_ocr(lang)
    historico_ocr.append(leitura_inicial)
    print(f"[OCR] Verificacao inicial: {leitura_inicial}")
    assinatura_anterior = _validar_leitura_certificado(
        leitura_inicial,
        assinatura_anterior=assinatura_anterior,
        etapa="verificacao inicial"
    )

    print(f"[CERT] Buscando: '{tema_busca}'")

    encontrado = False
    linha_encontrada = None

    linhas_anteriores = None
    repeticoes = 0
    leituras_fracas = 0
    busca_inicio = time.time()
    limite_busca_segundos = min(45, max(20, int(MAX_ROLAGENS) * 2))

    for i in range(MAX_ROLAGENS):
        if time.time() - busca_inicio > limite_busca_segundos:
            print(f"[CERT] LIMITE DE BUSCA ATINGIDO ({limite_busca_segundos}s) — analisando histórico OCR")
            return _finalizar_busca_certificado_nao_encontrado(
                historico_ocr,
                motivo="limite de busca atingido"
            )
        if _mouse_dentro_area_ocr():
            _fechar_popup_certificado()
            raise CertificadoInterferenciaException(
                f"CERT_INTERFERENCIA_HUMANA: mouse entrou na área OCR antes da rolagem {i}."
            )

        if abs(pyautogui.position()[0] - _mouse_seguro_x) > 8 or abs(pyautogui.position()[1] - _mouse_seguro_y) > 8:
            _fechar_popup_certificado()
            raise CertificadoInterferenciaException(
                f"CERT_INTERFERENCIA_HUMANA: mouse movimentado durante OCR do certificado antes da rolagem {i}."
            )

        linhas = capturar_linhas_ocr(lang)
        historico_ocr.append(linhas)

        # DIAGNÓSTICO — mostra o que o OCR leu
        print(f"[OCR] Rolagem {i}: {linhas}")

        qualidade = _qualidade_linhas_ocr(linhas)
        if qualidade["validas"] == 0:
            leituras_fracas += 1
        else:
            leituras_fracas = 0

        # Não interrompe no meio por leitura fraca/contexto estranho.
        # Guarda no histórico e decide no final se foi certificado ausente ou interferência.
        assinatura_anterior = _validar_leitura_certificado(
            linhas,
            assinatura_anterior=assinatura_anterior,
            etapa=f"rolagem {i}"
        )

        # DETECÇÃO DE REPETIÇÃO (FIM DA LISTA)
        # Só considera fim da lista quando ainda existe leitura válida.
        # Se repetir lixo/vazio, é interferência, não certificado ausente.
        if linhas == linhas_anteriores:
            repeticoes += 1
        else:
            repeticoes = 0
            linhas_anteriores = linhas

        if repeticoes >= 2:
            print("[CERT] FIM DA LISTA DETECTADO — analisando histórico OCR")
            return _finalizar_busca_certificado_nao_encontrado(
                historico_ocr,
                motivo="fim da lista detectado"
            )

        for num, txt in linhas.items():
            if parece_certificado_correto(txt, tema_busca):
                print(f"[CERT] ENCONTRADO rolagem {i} linha {num}: '{txt[:60]}'")
                encontrado = True
                linha_encontrada = num
                break

        if encontrado:
            break

        pyautogui.press("down")
        time.sleep(0.18)
        _afastar_mouse_area_ocr()
        _mouse_seguro_x, _mouse_seguro_y = pyautogui.position()

    if not encontrado:
        print("[CERT] ERRO: certificado não encontrado — analisando histórico OCR")
        return _finalizar_busca_certificado_nao_encontrado(
            historico_ocr,
            motivo="certificado não encontrado após todas as rolagens"
        )

    x_clique = LINHA_X[linha_encontrada]
    y_clique = LINHA_Y[linha_encontrada]
    print(f"[CERT] Clicando linha {linha_encontrada} X={x_clique}, Y={y_clique}...")
    pyautogui.click(x_clique, y_clique)
    time.sleep(0.5)
    pyautogui.click(x_clique, y_clique)
    time.sleep(0.3)

    print(f"[CERT] Clicando OK ({OK_X}, {OK_Y})...")
    pyautogui.click(OK_X, OK_Y)
    time.sleep(3)

    print("[CERT] OK clicado! Certificado selecionado.")
    return True

def clicar_botao_certificado(driver, wait):
    xpaths = [
        "//*[self::button or self::a][contains(translate(normalize-space(.),'ABCDEFGHIJKLMNOPQRSTUVWXYZ','abcdefghijklmnopqrstuvwxyz'),'certificado')]",
        "//*[contains(translate(normalize-space(.),'ABCDEFGHIJKLMNOPQRSTUVWXYZ','abcdefghijklmnopqrstuvwxyz'),'acesso com certificado')]",
    ]

    for xpath in xpaths:
        try:
            el = wait.until(EC.presence_of_element_located((By.XPATH, xpath)))
            driver.execute_script("arguments[0].scrollIntoView({block:'center'});", el)
            time.sleep(0.5)
            wait.until(EC.element_to_be_clickable((By.XPATH, xpath)))
            driver.execute_script("arguments[0].click();", el)
            print("OK - botão certificado clicado.")
            return True
        except Exception:
            pass

    clicou = driver.execute_script("""
        const els = Array.from(document.querySelectorAll('button,a,img,div,span'));
        const alvo = els.find(el => {
            const txt = ((el.innerText||'')+(el.alt||'')+(el.title||'')+(el.ariaLabel||'')).toLowerCase();
            const r = el.getBoundingClientRect();
            const s = window.getComputedStyle(el);
            return r.width>0&&r.height>0&&s.visibility!=='hidden'&&s.display!=='none'&&txt.includes('certificado');
        });
        if(alvo){alvo.scrollIntoView({block:'center'});alvo.click();return true;}
        return false;
    """)

    if clicou:
        print("OK - botão certificado clicado via JS.")
        return True

    raise Exception("Botão de certificado não localizado.")


def login_com_senha(driver, wait, login, senha):
    print("Rota: LOGIN E SENHA")

    campo_login = wait.until(
        EC.presence_of_element_located((By.XPATH, "//input[@placeholder='CPF/CNPJ']"))
    )
    campo_login.click()

    for letra in str(login):
        campo_login.send_keys(letra)
        time.sleep(0.05)

    campo_senha = wait.until(
        EC.presence_of_element_located((By.XPATH, "//input[@type='password']"))
    )
    campo_senha.click()

    for letra in str(senha):
        campo_senha.send_keys(letra)
        time.sleep(0.05)

    wait.until(
        EC.element_to_be_clickable((By.XPATH, "//button[contains(., 'Entrar')]"))
    ).click()

    time.sleep(3)

    # Detecta mensagens de login inválido ANTES de esperar o site estabilizar
    try:
        texto_pagina = driver.find_element(By.TAG_NAME, "body").text.lower()
    except Exception:
        texto_pagina = driver.page_source.lower()

    mensagens_login_invalido = [
        "usuário informado deve ser um cpf",
        "usuario informado deve ser um cpf",
        "cpf(11 dígitos) ou cnpj(14 dígitos)",
        "cpf(11 digitos) ou cnpj(14 digitos)",
        "senha incorreta",
        "login incorreto",
        "usuário ou senha",
        "usuario ou senha",
        "login ou senha",
        "credenciais inválidas",
        "credenciais invalidas",
        "não confere",
        "nao confere",
        "inválido",
        "invalido",
        "acesso negado",
    ]

    if any(msg in texto_pagina for msg in mensagens_login_invalido):
        raise LoginInvalidoException(f"Login/senha inválido para o usuário '{login}'.")

    esperar_site_estavel(driver, wait, timeout=35, descricao="pós-login por senha")
    print("OK - login por senha concluído.")


def _encerrar_driver_sem_travar(driver, timeout=3):
    """Encerra o Selenium sem travar a fila e sem deixar Chrome aberto.

    Quando o popup nativo de certificado fica aberto, o driver.quit() pode travar.
    Nesse caso, usamos taskkill no processo do ChromeDriver com /T /F para matar
    também a árvore de processos do Chrome aberta por esta tentativa.
    """
    if driver is None:
        return

    pid_driver = None
    try:
        processo = getattr(getattr(driver, "service", None), "process", None)
        pid_driver = getattr(processo, "pid", None)
    except Exception:
        pid_driver = None

    def _quit():
        try:
            driver.quit()
        except Exception:
            pass

    try:
        _fechar_popup_certificado()
    except Exception:
        pass

    try:
        t_quit = threading.Thread(target=_quit, daemon=True)
        t_quit.start()
        t_quit.join(timeout=timeout)
        if not t_quit.is_alive():
            return

        print(f"  [AVISO] driver.quit demorou mais de {timeout}s — forçando encerramento do Chrome desta tentativa.")
    except Exception:
        pass

    # Se o quit travou, mata a árvore do ChromeDriver para não deixar abas/janelas abertas.
    if pid_driver:
        try:
            import subprocess
            subprocess.run(
                ["taskkill", "/PID", str(pid_driver), "/T", "/F"],
                stdout=subprocess.DEVNULL,
                stderr=subprocess.DEVNULL,
                timeout=4,
                check=False,
            )
            print(f"  [OK] Chrome/ChromeDriver encerrados à força. PID={pid_driver}")
            return
        except Exception as e:
            print(f"  [AVISO] taskkill por PID falhou: {e}")

    # Fallback: tenta matar diretamente o processo do serviço, se ainda existir.
    try:
        processo = getattr(getattr(driver, "service", None), "process", None)
        if processo:
            processo.kill()
            print("  [OK] Processo do ChromeDriver encerrado por fallback kill().")
    except Exception:
        pass


def login_com_certificado(driver, wait, tema, lang):
    print("Rota: CERTIFICADO DIGITAL")
    driver.maximize_window()

    try:
        driver.minimize_window()
        time.sleep(0.5)
        driver.maximize_window()

        driver.switch_to.window(driver.current_window_handle)
        driver.execute_script("window.focus();")
        driver.execute_script("window.moveTo(0,0);")
    except Exception:
        pass

    if not tema or str(tema).strip().upper() == "X":
        raise CertificadoNaoLocalizadoException("TEMA_CERTIFICADO vazio ou X — campo nao preenchido na planilha.")

    resultado_cert = {"ok": None}
    resultado_click = {"ok": None, "erro": ""}

    def executar_selecao_certificado():
        try:
            resultado_cert["ok"] = thread_selecionar_certificado(tema, lang)
        except CertificadoInterferenciaException as e:
            resultado_cert["ok"] = False
            resultado_cert["tipo_erro"] = "INTERFERENCIA_OCR"
            resultado_cert["erro"] = str(e)
        except Exception as e:
            resultado_cert["ok"] = False
            resultado_cert["tipo_erro"] = "CERTIFICADO"
            resultado_cert["erro"] = str(e)

    def executar_clique_certificado():
        try:
            clicar_botao_certificado(driver, wait)
            resultado_click["ok"] = True
        except Exception as e:
            resultado_click["ok"] = False
            resultado_click["erro"] = str(e)
            # Em alguns casos o Selenium fica bloqueado pelo popup nativo.
            # Se chegou aqui com erro, não é necessariamente falha: a janela pode ter aberto.
            print("OK - janela de certificado aberta; aguardando seleção segura.")

    t = threading.Thread(target=executar_selecao_certificado, daemon=True)
    t.start()

    # Importante: o clique que abre o popup também roda em thread.
    # Motivo: quando o popup nativo abre, o Selenium pode ficar bloqueado no clique
    # até a janela ser fechada. Se o OCR detectar interferência, a thread principal
    # precisa conseguir mandar a empresa para reprocessamento imediatamente.
    t_click = threading.Thread(target=executar_clique_certificado, daemon=True)
    t_click.start()

    print("Aguardando seleção do certificado...")
    inicio = time.time()
    timeout_total = 60

    while time.time() - inicio < timeout_total:
        if resultado_cert.get("ok") is not None:
            break
        time.sleep(0.2)

    if resultado_cert.get("ok") is False:
        if resultado_cert.get("tipo_erro") == "INTERFERENCIA_OCR":
            _fechar_popup_certificado()
            print("[CERT] Interferência OCR detectada — encerrando tentativa e enviando para tratamento de reprocessamento.")
            raise CertificadoInterferenciaException(
                f"INTERFERÊNCIA NO OCR DO CERTIFICADO: {resultado_cert.get('erro', tema)}"
            )
        raise CertificadoNaoLocalizadoException(
            f"CERTIFICADO NÃO LOCALIZADO: {resultado_cert.get('erro', tema)}"
        )

    if resultado_cert.get("ok") is not True:
        _fechar_popup_certificado()
        raise CertificadoInterferenciaException(
            "INTERFERÊNCIA NO OCR DO CERTIFICADO: seleção não retornou confirmação dentro do tempo limite."
        )

    try:
        esperar_site_estavel(driver, wait, timeout=45, descricao="pós-login por certificado")
    except Exception as e_pos_cert:
        msg_pos_cert = str(e_pos_cert)
        if _driver_comunicacao_quebrada(msg_pos_cert):
            raise PortalInstabilidadeException(
                "CONEXAO_CHROMEDRIVER_PERDIDA_POS_CERTIFICADO: " + msg_pos_cert
            )
        raise

    print("OK - login por certificado concluído.")

def extrair_numero_nfse_da_linha(linha):
    """
    Extrai o número da NFS-e diretamente do texto da linha da tabela.
    A coluna de número da nota contém apenas dígitos — pega o primeiro
    elemento da linha que seja somente numérico.
    Retorna o número como string, ou None se não encontrar.
    """
    try:
        celulas = linha.find_elements(By.TAG_NAME, "td")
        for cel in celulas:
            texto = (cel.text or "").strip()
            if texto.isdigit():
                return texto
    except Exception:
        pass
    return None


def processar_empresa_pos_login(driver, wait, codigo, nome, pasta_download, pasta_print=None):
    _pasta_print = pasta_print if pasta_print else pasta_download
    total_registros = None
    erros_empresa = []
    qtd_canceladas = 0
    qtd_analise = 0
    qtd_pdf_execucao = 0
    qtd_xml_execucao = 0
    xmls_execucao = []
    nfse_processadas = set()  # evita baixar mesma nota duas vezes (contingência)

    esperar_clicavel_seguro(
        driver,
        By.XPATH,
        "//a[contains(@href, '/EmissorNacional/Notas/Emitidas')]",
        timeout=40,
        descricao="menu NFS-e Emitidas"
    ).click()

    print("OK - entrou em Emitidas")
    esperar_site_estavel(driver, wait, timeout=35, descricao="tela de NFS-e Emitidas")

    campo_inicio = wait.until(EC.presence_of_element_located((By.ID, "datainicio")))
    campo_inicio.clear()
    campo_inicio.send_keys(DATA_INICIAL)

    campo_fim = wait.until(EC.presence_of_element_located((By.ID, "datafim")))
    campo_fim.clear()
    campo_fim.send_keys(DATA_FINAL)

    wait.until(
        EC.element_to_be_clickable((By.XPATH, "//button[contains(., 'Filtrar')]"))
    ).click()

    print("OK - filtrou")

    # Detecção rápida de "Nenhum registro encontrado" — não espera timeout cheio
    inicio_espera = time.time()
    nenhum_registro = False
    while time.time() - inicio_espera < 20:
        try:
            texto = driver.find_element(By.TAG_NAME, "body").text
            if "nenhum registro" in texto.lower():
                nenhum_registro = True
                break
            linhas = driver.find_elements(By.XPATH, "//table//tbody//tr")
            if linhas:
                break
        except Exception:
            pass
        time.sleep(0.5)

    esperar_site_estavel(driver, wait, timeout=20, descricao="resultado do filtro")

    total_registros = obter_total_registros(driver)

    if nenhum_registro or total_registros == 0:
        total_registros = 0
        print("Empresa sem movimento: nenhum download será realizado.")
        salvar_print_tela(driver, _pasta_print, codigo, f"{nome}_emitidas")
        return total_registros, qtd_canceladas, qtd_analise, erros_empresa, qtd_pdf_execucao, qtd_xml_execucao, xmls_execucao

    print(f"Total de registros no portal: {total_registros}")
    salvar_print_tela(driver, _pasta_print, codigo, f"{nome}_emitidas")

    pagina_atual = 1

    while True:
        linhas = driver.find_elements(By.XPATH, "//table//tbody//tr")
        print(f"\nPágina {pagina_atual} - notas na tela: {len(linhas)}")

        for i in range(len(linhas)):
            print(f"\nProcessando página {pagina_atual}, nota {i + 1}")

            try:
                linhas = driver.find_elements(By.XPATH, "//table//tbody//tr")
                linha = linhas[i]

                # Detecta status da nota — 4 tipos pelo HTML confirmado via DevTools
                # Normal: data-situacao="P100_GERADA"
                # Cancelada: tb-cancelada no src do img
                # Cancelamento em análise: tb-pendente.svg no src do img
                # Substituída: tb-subs.svg / data-situacao="P105_NFSE_SUBSTITUIDA"
                status_nota = ""
                try:
                    html_linha = linha.get_attribute("outerHTML") or ""
                    html_lower = html_linha.lower()
                    if "tb-cancelada" in html_lower or "nfs-e cancelada" in html_lower:
                        status_nota = "Cancelada"
                    elif "tb-pendente" in html_lower or "cancelamento sob análise" in html_lower or "cancelamento sob analise" in html_lower:
                        status_nota = "Cancelamento em análise"
                    elif "tb-subs" in html_lower or "p105_nfse_substituida" in html_lower or "substituição" in html_lower or "substituicao" in html_lower:
                        status_nota = "Substituída"
                except Exception:
                    pass

                if not status_nota:
                    try:
                        status_nota = driver.execute_script("""
                            const linha = arguments[0];
                            const imgs = linha.querySelectorAll('img');
                            for (const img of imgs) {
                                const src = (img.getAttribute('src') || '').toLowerCase();
                                const title = (img.getAttribute('title') || '').toLowerCase();
                                const orig = (img.getAttribute('data-original-title') || '').toLowerCase();
                                if (src.includes('tb-cancelada') || title.includes('cancelada') || orig.includes('cancelada')) return 'Cancelada';
                                if (src.includes('tb-pendente') || title.includes('análise') || orig.includes('análise') || orig.includes('analise')) return 'Cancelamento em análise';
                                if (src.includes('tb-subs') || title.includes('substitui') || orig.includes('substitui')) return 'Substituída';
                            }
                            return '';
                        """, linha) or ""
                    except Exception:
                        pass

                if status_nota == "Cancelada":
                    qtd_canceladas += 1
                    print(f"  [CANCELADA] Baixando XML para detalhamento.")
                elif status_nota == "Cancelamento em análise":
                    qtd_analise += 1
                    print(f"  [EM ANÁLISE] Baixando XML para detalhamento.")
                elif status_nota == "Substituída":
                    print(f"  [SUBSTITUÍDA] Baixando XML para detalhamento.")

                url_xml, url_pdf = capturar_links_menu(driver, wait, linha)

                conteudo_xml = baixar_conteudo_direto(
                    driver=driver,
                    url=url_xml,
                    pasta=pasta_download,
                    nome_debug=f"XML_p{pagina_atual}_{i + 1}"
                )

                num_nfse = extrair_numero_nfse_do_xml(conteudo_xml)

                if not num_nfse:
                    num_nfse = extrair_numero_nfse_da_linha(linha)
                    if num_nfse:
                        print(f"AVISO - número não veio pelo XML; usando número da linha: {num_nfse}")

                if num_nfse:
                    print(f"Número da NFS-e identificado pelo XML: {num_nfse}")
                    if num_nfse in nfse_processadas:
                        print(f"  [SKIP] Nota {num_nfse} já processada (contingência/duplicata) — ignorando.")
                        continue
                    nfse_processadas.add(num_nfse)
                    nome_base = f"PADRAO_NACIONAL_NFSE_{num_nfse}_{COMPETENCIA}"
                else:
                    print(f"AVISO - não extraiu número da NFS-e; usando fallback p{pagina_atual}_{i + 1}")
                    nome_base = f"PADRAO_NACIONAL_NFSE_p{pagina_atual}_{i + 1}_{COMPETENCIA}"

                # Notas normais: salva XML + PDF normalmente
                # Notas especiais (cancelada/análise): salva XML + PDF em pasta temporária separada
                if status_nota in ("Cancelada", "Cancelamento em análise"):
                    _status_norm = (status_nota.lower()
                                    .replace(' ', '_')
                                    .replace('ã', 'a').replace('á', 'a').replace('â', 'a').replace('à', 'a')
                                    .replace('é', 'e').replace('ê', 'e')
                                    .replace('í', 'i').replace('î', 'i')
                                    .replace('õ', 'o').replace('ó', 'o').replace('ô', 'o')
                                    .replace('ú', 'u').replace('û', 'u')
                                    .replace('ç', 'c'))
                    pasta_especial = os.path.join(pasta_download, f"_especiais_{_status_norm}")
                    os.makedirs(pasta_especial, exist_ok=True)
                    salvar_conteudo_arquivo(
                        conteudo=conteudo_xml,
                        pasta=pasta_especial,
                        extensao=".xml",
                        nome_padrao=f"{nome_base}.xml"
                    )
                    if url_pdf:
                        _pdf_ok = False
                        for _tent_pdf in range(1, 4):
                            try:
                                baixar_arquivo_direto(
                                    driver=driver,
                                    url=url_pdf,
                                    pasta=pasta_especial,
                                    extensao=".pdf",
                                    nome_padrao=f"{nome_base}.pdf",
                                    timeout=15
                                )
                                _pdf_ok = True
                                break
                            except Exception as e_pdf_esp:
                                if _tent_pdf < 3:
                                    print(f"  [AVISO] PDF nota especial tentativa {_tent_pdf}/3 falhou — retentando em 3s... ({e_pdf_esp})")
                                    time.sleep(3)
                                else:
                                    _msg_pdf_err = f"PDF nota especial {nome_base} nao baixado apos 3 tentativas: {e_pdf_esp}"
                                    print(f"  [ERRO] {_msg_pdf_err}")
                                    erros_empresa.append(_msg_pdf_err)
                                    _log_diag_pn.append({
                                        "fase": "PDF_ESPECIAL",
                                        "empresa": f"{codigo} - {nome}",
                                        "resultado": "FALHA_PDF_ESPECIAL",
                                        "observacao": _msg_pdf_err,
                                    })
                    dados_nota = extrair_dados_xml_bytes(conteudo_xml, numero_fallback=num_nfse or "", status=status_nota)
                    xmls_execucao.append(dados_nota)
                    try:
                        driver.find_element(By.TAG_NAME, "body").send_keys(Keys.ESCAPE)
                    except Exception:
                        pass
                    continue

                caminho_xml_salvo = salvar_conteudo_arquivo(
                    conteudo=conteudo_xml,
                    pasta=pasta_download,
                    extensao=".xml",
                    nome_padrao=f"{nome_base}.xml"
                )
                qtd_xml_execucao += 1
                dados_nota = extrair_dados_xml_bytes(conteudo_xml, numero_fallback=num_nfse or "", status=status_nota)
                xmls_execucao.append(dados_nota)

                baixar_arquivo_direto(
                    driver=driver,
                    url=url_pdf,
                    pasta=pasta_download,
                    extensao=".pdf",
                    nome_padrao=f"{nome_base}.pdf"
                )
                qtd_pdf_execucao += 1

                try:
                    driver.find_element(By.TAG_NAME, "body").send_keys(Keys.ESCAPE)
                except Exception:
                    pass

            except Exception as e:
                erro = f"Página {pagina_atual}, nota {i + 1}: {e}"
                print(f"ERRO - {erro}")
                erros_empresa.append(erro)

            time.sleep(1)

        if not ir_para_proxima_pagina(driver, pagina_atual):
            print("Última página ou próxima página não encontrada.")
            break

        pagina_atual += 1

    return total_registros, qtd_canceladas, qtd_analise, erros_empresa, qtd_pdf_execucao, qtd_xml_execucao, xmls_execucao


def calcular_valor_total_xmls(pasta):
    """Soma o <vLiq> de todos os XMLs PADRAO_NACIONAL_NFSE_*.xml da pasta."""
    NS = "http://www.sped.fazenda.gov.br/nfse"
    total = 0.0
    arquivos = [
        f for f in os.listdir(pasta)
        if f.upper().startswith("PADRAO_NACIONAL_NFSE_") and f.lower().endswith(".xml")
    ]
    for arq in arquivos:
        try:
            tree = ET.parse(os.path.join(pasta, arq))
            raiz = tree.getroot()
            # Tenta com namespace
            el = raiz.find(f".//{{{NS}}}valores/{{{NS}}}vLiq")
            if el is None:
                # Fallback sem namespace
                el = raiz.find(".//valores/vLiq")
            if el is not None and el.text:
                total += float(el.text.strip())
        except Exception:
            pass
    return round(total, 2)


def calcular_valor_total_xmls_execucao(xmls_execucao, filtro_status=None):
    """
    Soma o valor_servico dos dicionarios extraidos em memoria.

    filtro_status:
      - None      : soma TUDO (comportamento padrao - usado pelas Recebidas e pela
                    coluna D "Emitidas - Valor" do gerencial = total bruto)
      - "regular" : soma apenas notas NAO canceladas (Regular + Substituida +
                    Cancelamento em analise). Usado pela coluna F "Regular - Valor".
      - "cancelada": soma apenas notas com status "Cancelada" efetivo.
                    Usado pela coluna E "Canceladas - Valor".
    """
    total = 0.0
    for d in xmls_execucao:
        try:
            status = str(d.get("status", "") or "").strip()
            if filtro_status == "cancelada":
                if status == "Cancelada":
                    total += float(d.get("valor_servico", 0.0))
            elif filtro_status == "regular":
                if status != "Cancelada":
                    total += float(d.get("valor_servico", 0.0))
            else:
                total += float(d.get("valor_servico", 0.0))
        except Exception:
            pass
    return round(total, 2)


def gerar_detalhamento_nfse(pasta_download, nome_empresa, competencia,
                              xmls_emitidas=None, xmls_recebidas=None):
    """
    Gera arquivo com duas abas: PRESTADO e TOMADO.
    PRESTADO: notas emitidas — coluna CNPJ/CLIENTE = tomador (<toma>)
    TOMADO  : notas recebidas — coluna CNPJ/PRESTADOR = emitente (<emit>)
    """
    xmls_emitidas  = xmls_emitidas  or []
    xmls_recebidas = xmls_recebidas or []

    if not xmls_emitidas and not xmls_recebidas:
        return

    def _sort_key(d):
        try:
            from datetime import datetime as _dt
            return _dt.strptime(d.get("emissao", "01.01.1900"), "%d.%m.%Y")
        except Exception:
            return d.get("emissao", "")

    notas_prest = sorted(xmls_emitidas,  key=_sort_key)
    notas_toma  = sorted(xmls_recebidas, key=_sort_key)

    mm, aaaa   = competencia[:2], competencia[2:]
    nome_limpo = re.sub(r"\s+", "_", re.sub(r"[^\w\s\-]", "", nome_empresa).strip())
    caminho_saida = os.path.join(pasta_download,
                                 f"Detalhamento_de_NFSe_{nome_limpo}_{mm}{aaaa}.xlsx")

    # Regime do cliente — lido do primeiro XML prestado
    _REGIME_MAP = {
        "1": "NÃO OPTANTE",
        "2": "MEI — Microempreendedor Individual",
        "3": "Simples Nacional EPP",
        "4": "Simples Nacional Microempresa",
    }
    _op_sn  = notas_prest[0].get("op_simples", "") if notas_prest else ""
    _regime = _REGIME_MAP.get(_op_sn, "NÃO IDENTIFICADO")

    # Inscricao (CNPJ ou CPF) da EMPRESA dona do detalhamento — usado no titulo.
    # Em PRESTADO: a empresa eh o EMITENTE (emit_cnpj).
    # Em TOMADO  : a empresa eh o TOMADOR (cnpj).
    _inscricao_emp = ""
    if notas_prest:
        _inscricao_emp = notas_prest[0].get("emit_cnpj", "")
    elif notas_toma:
        _inscricao_emp = notas_toma[0].get("cnpj", "")

    def _fill(h):  return PatternFill("solid", fgColor="FF" + h)
    def _font(bold=False, cor="000000", size=9):
        return Font(bold=bold, color="FF" + cor, name="Calibri", size=size)
    def _borda():
        s = Side(style="thin", color="FFB0B0B0")
        return Border(left=s, right=s, top=s, bottom=s)
    def _alin(h="center"):
        return Alignment(horizontal=h, vertical="center")

    # ── Função interna: monta uma aba ────────────────────────────────────────
    def _render_sheet(ws, notas, tipo):
        eh_toma = (tipo == "tomado")

        # Linha 1 — título
        ws.merge_cells("A1:X1")
        # Inscricao aparece apenas em PRESTADO (a empresa eh a emitente).
        # Em TOMADO a empresa eh tomador e os prestadores variam por nota,
        # entao mostrar a inscricao no titulo confundiria.
        if eh_toma:
            ws["A1"] = (f"DETALHAMENTO DE NFS-e  —  {nome_empresa.upper()}  —  "
                        f"{mm}/{aaaa} - REGIME: {_regime}")
        else:
            ws["A1"] = (f"DETALHAMENTO DE NFS-e  —  {nome_empresa.upper()}  - {_inscricao_emp} —  "
                        f"{mm}/{aaaa} - REGIME: {_regime}")
        ws["A1"].fill      = _fill("1F4E78")
        ws["A1"].font      = _font(bold=True, cor="FFFFFF", size=11)
        ws["A1"].alignment = _alin("left")
        ws.row_dimensions[1].height = 22

        # Linha 2 — grupos
        label_a = "PRESTADOR" if eh_toma else "TOMADOR"
        grupos = [
            (label_a,              "A2", "B2", "2E75B6"),
            ("DOCUMENTO",          "C2", "J2", "1F6B75"),
            ("ISS",                "K2", "N2", "6C3483"),
            ("RETENÇÕES FEDERAIS", "O2", "W2", "1A5276"),
            ("NOTA",               "X2", "X2", "7D6608"),
        ]
        for label, ini, fim, cor in grupos:
            ws.merge_cells(f"{ini}:{fim}")
            c = ws[ini]
            c.value = label; c.fill = _fill(cor)
            c.font = _font(bold=True, cor="FFFFFF", size=10)
            c.alignment = _alin("center"); c.border = _borda()
        ws.row_dimensions[2].height = 18

        # Linha 3 — cabeçalhos
        label_b = "PRESTADOR" if eh_toma else "CLIENTE"
        cabecalhos = [
            "CNPJ", label_b,
            "EMISSÃO", "NÚMERO", "STATUS", "CNAE", "VALOR TOTAL NFSe", "DEDUÇÃO", "DESCONTOS\nINCONDICIONADOS", "VALOR SERVIÇO",
            "ISS", "BASE DE CALCULO", "ALÍQUOTA", "VALOR ISS",
            "INSS", "IRRF", "CSLL",
            "BASE PIS", "ALIQUOTA PIS", "PIS",
            "BASE COFINS", "ALIQUOTA COFINS", "COFINS",
            "CHAVE",
        ]
        larguras = [18, 35, 13, 10, 20, 10, 18, 12, 14, 16, 12, 16, 12, 14,
                    14, 14, 14, 14, 13, 14, 14, 15, 14, 40]
        for col, (cab, larg) in enumerate(zip(cabecalhos, larguras), 1):
            c = ws.cell(row=3, column=col, value=cab)
            # Coluna I (9) = DESCONTOS INCONDICIONADOS — fonte menor para o texto longo caber
            _size_header = 8 if col == 9 else 9
            c.fill = _fill("2E86C1"); c.font = _font(bold=True, cor="FFFFFF", size=_size_header)
            _alin_h = Alignment(horizontal="center", vertical="center", wrap_text=True) if "\n" in str(cab) else _alin("center")
            c.alignment = _alin_h; c.border = _borda()
            ws.column_dimensions[c.column_letter].width = larg
        ws.row_dimensions[3].height = 28
        ws.auto_filter.ref = "A3:X3"

        # Estilos de fundo
        fill_normal  = _fill("EBF5FB"); fill_alt     = _fill("F4F6F7")
        fill_cancel  = _fill("FADBD8"); fill_analise = _fill("FEF9E7")
        fill_substit = _fill("E8DAEF")

        total_g = total_i = 0.0
        n_regular = n_cancelada = n_analise = n_substit = 0

        monetarios    = {7, 8, 9, 10, 12, 14, 15, 16, 17, 18, 20, 21, 23}
        aliquota_col  = 13
        aliquota_cols = {19, 22}

        for i, d in enumerate(notas, start=4):
            status     = d.get("status", "")
            _vnfse     = d.get("valor_nfse", 0.0)
            _vliq      = d.get("valor_liq",  0.0)
            _cancelada = (status == "Cancelada")

            if status == "Cancelada":
                base_fill = fill_cancel;  n_cancelada += 1
            elif status == "Cancelamento em análise":
                base_fill = fill_analise; n_analise += 1
                total_g += round(_vnfse, 2); total_i += round(_vliq, 2)
            elif status == "Substituída":
                base_fill = fill_substit; n_substit += 1
                total_g += round(_vnfse, 2); total_i += round(_vliq, 2)
            else:
                base_fill = fill_alt if i % 2 == 0 else fill_normal
                n_regular += 1
                total_g += round(_vnfse, 2); total_i += round(_vliq, 2)

            # TOMADO: CNPJ/nome vêm do emitente (<emit>); PRESTADO: do tomador (<toma>)
            _cnpj    = d.get("emit_cnpj", d.get("cnpj",    "")) if eh_toma else d.get("cnpj",    "")
            _cliente = d.get("emit_nome", d.get("cliente", "")) if eh_toma else d.get("cliente", "")
            # Nota cancelada sem tomador no XML — portal nao fornece <toma>
            if _cancelada and not _cnpj and not _cliente:
                _cnpj    = ""
                _cliente = "Tomador e Intermediário não identificados"

            valores_linha = [
                _cnpj, _cliente,
                d.get("emissao", ""), d.get("numero", ""),
                "",  # STATUS col E — abaixo
                d.get("cnae", ""),
                0.0  if _cancelada else d.get("valor_nfse",      0.0),  # G
                0.00,                                                      # H
                0.0  if _cancelada else d.get("v_desc_incond",   0.0),  # I
                0.0  if _cancelada else d.get("valor_liq",        0.0),  # J
                ""   if _cancelada else d.get("tp_ret_issqn",     ""),   # K
                0.0  if _cancelada else d.get("base_calculo",     0.0),  # L
                0.0  if _cancelada else d.get("aliquota",         0.0),  # M
                0.0  if _cancelada else d.get("valor_iss",        0.0),  # N
                0.0  if _cancelada else d.get("v_ret_cp",         0.0),  # O
                0.0  if _cancelada else d.get("v_ret_irrf",       0.0),  # P
                0.0  if _cancelada else d.get("v_ret_csll",       0.0),  # Q
                0.0  if _cancelada else d.get("v_bc_pis_cofins",  0.0),  # R
                0.0  if _cancelada else d.get("p_aliq_pis",       0.0),  # S
                0.0  if _cancelada else d.get("v_pis",            0.0),  # T
                0.0  if _cancelada else d.get("v_bc_pis_cofins",  0.0),  # U
                0.0  if _cancelada else d.get("p_aliq_cofins",    0.0),  # V
                0.0  if _cancelada else d.get("v_cofins",         0.0),  # W
                d.get("chave", ""),                                        # X
            ]

            for col, val in enumerate(valores_linha, 1):
                c = ws.cell(row=i, column=col, value=val)
                c.fill = base_fill; c.font = _font(size=9); c.border = _borda()
                if col in monetarios:
                    c.number_format = 'R$ #,##0.00'; c.alignment = _alin("right")
                elif col == aliquota_col or col in aliquota_cols:
                    c.number_format = '0.00"%"'; c.alignment = _alin("center")
                elif col in (3, 4, 6, 11):
                    c.alignment = _alin("center")
                elif col == 24:
                    c.alignment = _alin("left"); c.font = _font(size=8)
                else:
                    c.alignment = _alin("left")

            # STATUS col E
            c_st = ws.cell(row=i, column=5)
            if status == "Cancelada":
                c_st.value = "✗  Cancelada"
                c_st.fill  = _fill("FADBD8")
                c_st.font  = Font(bold=True, color="FFC0392B", name="Calibri", size=9)
            elif status == "Cancelamento em análise":
                c_st.value = "⚠  Cancelamento em análise"
                c_st.fill  = _fill("FEF9E7")
                c_st.font  = Font(bold=True, color="FF9C6500", name="Calibri", size=9)
            elif status == "Substituída":
                c_st.value = "↺  Substituída"
                c_st.fill  = _fill("E8DAEF")
                c_st.font  = Font(bold=True, color="FF6C3483", name="Calibri", size=9)
            else:
                c_st.value = "✓"
                c_st.fill  = _fill("D5F5E3")
                c_st.font  = Font(bold=True, color="FF1E8449", name="Calibri", size=10)
            c_st.alignment = _alin("center"); c_st.border = _borda()
            ws.row_dimensions[i].height = 15

        # Rodapé
        lt = len(notas) + 4
        ws.merge_cells(f"A{lt}:F{lt}")
        _txt_tot = (f"TOTAL  |  Regular: {n_regular}  |  Substituídas: {n_substit}  |  "
                    f"Canceladas: {n_cancelada}  |  Cancelamento em análise: {n_analise}"
                    if n_substit else
                    f"TOTAL  |  Regular: {n_regular}  |  Canceladas: {n_cancelada}  |  "
                    f"Cancelamento em análise: {n_analise}")
        c_l = ws.cell(row=lt, column=1, value=_txt_tot)
        c_l.fill = _fill("F9E79F"); c_l.font = _font(bold=True, size=10)
        c_l.alignment = _alin("center"); c_l.border = _borda()

        p, u = 4, lt - 1

        def _tot_mon(col_num, formula):
            c = ws.cell(row=lt, column=col_num, value=formula)
            c.fill = _fill("D5F5E3"); c.font = _font(bold=True, size=10)
            c.number_format = 'R$ #,##0.00'; c.alignment = _alin("right"); c.border = _borda()

        def _tot_vaz(col_num):
            ws.cell(row=lt, column=col_num, value="").fill  = _fill("F9E79F")
            ws.cell(row=lt, column=col_num).border = _borda()

        _tot_mon(7,  round(total_g, 2))
        _tot_vaz(8)
        _tot_mon(9,  f"=SUMIFS(I{p}:I{u},E{p}:E{u},\"<>*Cancelada*\")")
        _tot_mon(10, round(total_i, 2))
        _tot_vaz(11)
        _tot_mon(12, f"=SUMIFS(L{p}:L{u},E{p}:E{u},\"<>*Cancelada*\")")
        _tot_vaz(13)
        _tot_mon(14, f"=SUMIFS(N{p}:N{u},E{p}:E{u},\"<>*Cancelada*\")")

        for cn, cl in {15:"O",16:"P",17:"Q",18:"R",19:None,20:"T",21:"U",22:None,23:"W"}.items():
            if cl:
                _tot_mon(cn, f"=SUMIFS({cl}{p}:{cl}{u},E{p}:E{u},\"<>*Cancelada*\")")
            else:
                _tot_vaz(cn)
        _tot_vaz(24)
        ws.row_dimensions[lt].height = 18

        # ISS Não Retido / Retido
        li1, li2 = lt + 1, lt + 2
        for li, txt, cor_fill, cor_font in [
            (li1, "Não Retido", "D5F5E3", "1E8449"),
            (li2, "Retido",     "FADBD8", "C0392B"),
        ]:
            c12 = ws.cell(row=li, column=13, value=txt)
            c12.fill = _fill(cor_fill); c12.font = _font(bold=True, size=10, cor=cor_font)
            c12.alignment = _alin("right"); c12.border = _borda()
            crit = f"\"Não Retido\"" if txt == "Não Retido" else f"\"Retido\""
            c13 = ws.cell(row=li, column=14,
                          value=f"=SUMIFS(N{p}:N{u},K{p}:K{u},{crit},E{p}:E{u},\"<>*Cancelada*\")")
            c13.fill = _fill(cor_fill); c13.font = _font(bold=True, size=10, cor=cor_font)
            c13.number_format = 'R$ #,##0.00'; c13.alignment = _alin("right"); c13.border = _borda()

        ws.freeze_panes = "A4"
    # ── fim _render_sheet ────────────────────────────────────────────────────

    wb = openpyxl.Workbook()
    ws_prest = wb.active
    ws_prest.title = "PRESTADO"
    _render_sheet(ws_prest, notas_prest, "prestado")

    ws_toma = wb.create_sheet("TOMADO")
    _render_sheet(ws_toma, notas_toma, "tomado")

    if os.path.exists(caminho_saida):
        try:
            os.remove(caminho_saida)
        except PermissionError:
            print(f"\n  *** ERRO: Feche o arquivo '{os.path.basename(caminho_saida)}' no Excel e execute novamente. ***\n")
            return
    try:
        wb.save(caminho_saida)
        print(f"  [DETALHAMENTO] {os.path.basename(caminho_saida)}")
    except Exception as e_d:
        print(f"  [AVISO] Detalhamento não salvo: {e_d}")



def criar_relatorio_gerencial(caminho_relatorio, linhas_relatorio):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Gerencial"

    # ── estilos base ──────────────────────────────────────────────────────────
    def _fill(hex6):  return PatternFill("solid", fgColor="FF" + hex6)
    def _font(color="222222", bold=False, size=9):
        return Font(color="FF" + color, bold=bold, size=size, name="Calibri")
    def _border():
        s = Side(style="thin", color="BFBFBF")
        return Border(left=s, right=s, top=s, bottom=s)
    def _aln(h="center", wrap=False):
        return Alignment(horizontal=h, vertical="center", wrap_text=wrap)

    header_fill  = _fill("1F4E78")
    prest_fill   = _fill("2E75B6")
    tomad_fill   = _fill("197A6E")
    row_fill_1   = _fill("FFFFFF")
    row_fill_2   = _fill("EBF3FB")
    green_fill   = _fill("E2F0D9")
    red_fill     = _fill("FCE4D6")
    yellow_fill  = _fill("FFF2CC")
    green_font   = _font("00B050", bold=True, size=11)
    red_font     = _font("C00000", bold=True, size=10)
    warning_font = _font("9C6500", bold=False, size=9)
    normal_font  = _font()

    # ── NOVO LAYOUT — 21 colunas (A..U) ───────────────────────────────────────
    #  A Codigo | B Inscricao | C Empresa
    #  EMITIDAS (B..M):
    #    D Emitidas Qtd | E Regular | F Canceladas
    #    G PDF | H XML | I Status PDF | J Status XML
    #    K Emitidas Valor | L Canceladas Valor | M Regular Valor
    #  RECEBIDAS (N..S):
    #    N Receb Qtd | O Receb Valor
    #    P PDF | Q Status PDF | R XML | S Status XML
    #  T Login/Senha | U Observacao
    cabecalhos = [
        "Código", "Inscrição", "Empresa",
        "Emitidas — Qtd", "Regular", "Canceladas",
        "PDF Baixados", "XML Baixados", "Status PDF", "Status XML",
        "Emitidas — Valor", "Canceladas — Valor", "Regular - Valor",
        "Recebidas — Qtd", "Recebidas — Valor",
        "PDF Baixados", "Status PDF", "XML Baixados", "Status XML",
        "Login/Senha", "Observação",
    ]
    larguras = [
        8, 18, 32,
        14, 10, 11,
        14, 14, 12, 12,
        17, 17, 17,
        16, 17,
        14, 12, 14, 12,
        14, 55,
    ]
    TOTAL_COLS = len(cabecalhos)  # 21

    # Cor leve verde para destacar a coluna "Regular" (qtd E e valor M) das emitidas
    regular_fill = _fill("EAF7DC")

    # ── linha 1: cabeçalhos de grupo EMITIDAS / RECEBIDAS ─────────────────────
    for col in range(1, TOTAL_COLS + 1):
        cel = ws.cell(row=1, column=col)
        cel.fill   = header_fill
        cel.border = _border()

    # EMITIDAS abrange colunas B..M (Inscrição até Status XML das emitidas)
    ws.merge_cells("B1:M1")
    c = ws["B1"]
    c.value     = "EMITIDAS"
    c.font      = Font(color="FFFFFFFF", bold=True, size=10, name="Calibri")
    c.fill      = prest_fill
    c.alignment = _aln("center")

    # RECEBIDAS abrange colunas N..S
    ws.merge_cells("N1:S1")
    c = ws["N1"]
    c.value     = "RECEBIDAS"
    c.font      = Font(color="FFFFFFFF", bold=True, size=10, name="Calibri")
    c.fill      = tomad_fill
    c.alignment = _aln("center")

    ws.row_dimensions[1].height = 18

    # ── linha 2: cabeçalhos de coluna ─────────────────────────────────────────
    for col, (cab, larg) in enumerate(zip(cabecalhos, larguras), 1):
        cel = ws.cell(row=2, column=col, value=cab)
        cel.fill      = header_fill
        cel.font      = Font(color="FFFFFFFF", bold=True, size=9, name="Calibri")
        cel.alignment = _aln("center", wrap=True)
        cel.border    = _border()
        ws.column_dimensions[cel.column_letter].width = larg

    ws.row_dimensions[2].height = 30

    # ── linhas de dados (início em 3) ─────────────────────────────────────────
    for idx, item in enumerate(linhas_relatorio, 3):
        observacao_upper = str(item.get("observacao", "")).upper()
        erro_acesso = any(t in observacao_upper for t in [
            "LOGIN/SENHA INCORRETO", "CERTIFICADO NÃO LOCALIZADO",
            "CERTIFICADO NAO LOCALIZADO", "CERTIFICADO VENCIDO",
        ])
        sem_movimento  = (
            item.get("emit_qtd") == "Sem movimento"
            and item.get("rec_qtd") == "Sem movimento"
            and not erro_acesso
        )
        eh_certificado = str(item.get("login_senha", "")) == "CERTIFICADO"
        eh_login_incorreto = str(item.get("login_senha", "")) == "Login incorreto"
        eh_divergente = (
            str(item.get("emit_pdf_st", "")) == "X"
            or str(item.get("emit_xml_st", "")) == "X"
            or str(item.get("rec_pdf_st",  "")) == "X"
            or str(item.get("rec_xml_st",  "")) == "X"
        ) and not erro_acesso and not eh_certificado and not eh_login_incorreto

        base_fill = row_fill_2 if idx % 2 == 0 else row_fill_1

        # Observação: encurtar texto de certificado + corrigir login
        obs_raw = str(item.get("observacao", ""))

        # LOGIN INCORRETO — prioridade máxima
        if str(item.get("login_senha", "")).strip().lower() == "login incorreto":
            obs_raw = "Login incorreto"
        elif "CERTIFICADO NÃO ENCONTRADO" in obs_raw or "CERTIFICADO NÃO LOCALIZADO" in obs_raw:
            obs_raw = "Certificado não localizado"
        elif eh_divergente:
            obs_raw = "Divergente — verificar"

        # Ordem das colunas do NOVO LAYOUT (A..U) — Qtd antes de Valor
        valores = [
            item.get("codigo", ""),               # A
            item.get("inscricao", ""),            # B
            item.get("empresa", ""),              # C
            item.get("emit_qtd", ""),             # D - Emitidas Qtd (total)
            item.get("emit_regular_qtd", ""),     # E - Regular Qtd          [VERDE LEVE]
            item.get("emit_cancel", ""),          # F - Canceladas Qtd
            item.get("emit_pdf", ""),             # G - PDF Baixados
            item.get("emit_xml", ""),             # H - XML Baixados
            item.get("emit_pdf_st", ""),          # I - Status PDF
            item.get("emit_xml_st", ""),          # J - Status XML
            item.get("emit_valor", ""),           # K - Emitidas Valor (total bruto)
            item.get("emit_valor_canc", ""),      # L - Canceladas Valor
            item.get("emit_regular_valor", ""),   # M - Regular Valor        [VERDE LEVE]
            item.get("rec_qtd", ""),              # N
            item.get("rec_valor", ""),            # O
            item.get("rec_pdf", ""),              # P
            item.get("rec_pdf_st", ""),           # Q
            item.get("rec_xml", ""),              # R
            item.get("rec_xml_st", ""),           # S
            item.get("login_senha", ""),          # T
            obs_raw,                              # U
        ]

        for col, val in enumerate(valores, 1):
            cel = ws.cell(row=idx, column=col, value=val)
            cel.border    = _border()
            # Empresa (C) e Observação (U) alinhadas à esquerda
            cel.alignment = _aln("left" if col in (3, 21) else "center")

            if sem_movimento:
                cel.fill = yellow_fill
                cel.font = warning_font
                # As colunas de emitidas (D..M) e recebidas (N..S) ficam amareladas
                if col in range(4, 20) and "Sem movimento" in str(cel.value or ""):
                    cel.value = "⚠ Sem movimento"
                continue

            if eh_certificado and col > 3:
                cel.fill = red_fill
                if col == 20:  # Login/Senha
                    cel.value = "Certificado"
                    cel.font  = _font("C00000", bold=False)
                else:
                    cel.font  = _font("C00000", bold=False)
                continue

            if eh_login_incorreto and col > 3:
                cel.fill = red_fill
                cel.font = _font("C00000", bold=False)
                continue

            if eh_divergente:
                cel.fill = PatternFill("solid", fgColor="FFD966")  # amarelo âmbar
                cel.font = _font("7D4E00", bold=True if col in (1, 3) else False)
                continue

            # Bloco emitidas (D..M) amarelado se sem movimento de emitidas
            emit_sem_mov = (item.get("emit_qtd") == "Sem movimento" and not erro_acesso)
            if emit_sem_mov and col in range(4, 14):
                cel.fill = yellow_fill
                cel.font = warning_font
                if "Sem movimento" in str(cel.value or ""):
                    cel.value = "⚠ Sem movimento"
                continue

            # Bloco recebidas (N..S) amarelado se sem movimento de recebidas
            rec_sem_mov = (item.get("rec_qtd") == "Sem movimento" and not erro_acesso)
            if rec_sem_mov and col in range(14, 20):
                cel.fill = yellow_fill
                cel.font = warning_font
                if "Sem movimento" in str(cel.value or ""):
                    cel.value = "⚠ Sem movimento"
                continue

            cel.fill = base_fill
            cel.font = normal_font

            # Formato monetário nas colunas de valor: K, L, M (emitidas), O (recebidas)
            if col in (11, 12, 13, 15) and isinstance(val, (int, float)):
                cel.number_format = 'R$ #,##0.00'

            # Status PDF/XML: I, J (emitidas), Q, S (recebidas)
            if col in (9, 10, 17, 19):
                if str(val) == "✓":
                    cel.fill = green_fill
                    cel.font = green_font
                elif str(val) == "X":
                    cel.fill = red_fill
                    cel.font = red_font

            # Canceladas qtd (col F = 6) — destaque vermelho quando tem cancelada
            if col == 6 and val and str(val) not in ("", "0", "Sem movimento"):
                try:
                    if int(val) > 0:
                        cel.fill = red_fill
                        cel.font = _font("C00000", bold=False)
                except Exception:
                    pass

            # Colunas Regular (qtd E=5 e valor M=13) — fundo verde leve para destacar
            # (aplicado por ultimo so quando nao recebeu outra cor especial)
            if col in (5, 13):
                # so aplica se nao tiver recebido ja uma cor de status (green/red)
                if cel.fill == base_fill or cel.fill is base_fill:
                    cel.fill = regular_fill

            # Login/Senha (col T = 20)
            if col == 20:
                if str(val) == "Login incorreto":
                    cel.fill = red_fill
                    cel.font = _font("C00000", bold=False)
                elif str(val) == "✓":
                    cel.fill = green_fill
                    cel.font = green_font

            # Observação (col U = 21)
            if col == 21:
                _login_ok = str(item.get("login_senha","")) not in ("Login incorreto","CERTIFICADO")
                _emit_ok  = str(item.get("emit_pdf_st","")) in ("✓","Sem movimento")
                _rec_ok   = str(item.get("rec_pdf_st",""))  in ("✓","Sem movimento")
                if obs_raw == "" and _login_ok and _emit_ok and _rec_ok:
                    cel.value     = "✓"
                    cel.fill      = green_fill
                    cel.font      = green_font
                    cel.alignment = _aln("center")
                elif "SEM MOVIMENTO" in obs_raw.upper():
                    cel.fill = yellow_fill
                    cel.font = warning_font
                elif obs_raw and ("ERRO" in obs_raw.upper() or "FALHA" in obs_raw.upper() or
                                  "CERTIFICADO" in obs_raw.upper() or "LOGIN" in obs_raw.upper()):
                    cel.fill = red_fill
                    cel.font = _font("C00000", bold=False)

        ws.row_dimensions[idx].height = 16

    ws.freeze_panes = "A3"
    ws.auto_filter.ref = f"A2:{ws.cell(2, ws.max_column).coordinate}"

    wb.save(caminho_relatorio)
    print(f"OK - relatório gerencial salvo: {caminho_relatorio}")


def zipar_arquivos_empresa(pasta, codigo, tipo):
    """
    Compacta os arquivos da empresa em 4 ZIPs separados por tipo e extensão.
    Nomes gerados:
        pdf_prestados_{cod}_{COMP}.zip   xml_prestados_{cod}_{COMP}.zip
        pdf_tomados_{cod}_{COMP}.zip     xml_tomados_{cod}_{COMP}.zip
    Após zipar, os arquivos individuais são deletados.
    tipo: 'prestados' ou 'tomados'
    """
    import zipfile

    prefixo = "PADRAO_NACIONAL_NFSE_" if tipo == "prestados" else "PADRAO_NACIONAL_RECEBIDA_"
    tipo_label = tipo  # 'prestados' ou 'tomados'

    for ext in (".pdf", ".xml"):
        ext_label = ext[1:]  # 'pdf' ou 'xml'
        arquivos = sorted([
            f for f in os.listdir(pasta)
            if f.startswith(prefixo) and f.lower().endswith(ext)
        ])
        if not arquivos:
            continue

        nome_zip = f"{ext_label}_{tipo_label}_{codigo}_{COMPETENCIA}.zip"
        caminho_zip = os.path.join(pasta, nome_zip)

        with zipfile.ZipFile(caminho_zip, "w", zipfile.ZIP_DEFLATED) as zf:
            for arq in arquivos:
                caminho_arq = os.path.join(pasta, arq)
                zf.write(caminho_arq, arq)

        # deleta arquivos individuais após zipar
        removidos = 0
        for arq in arquivos:
            try:
                os.remove(os.path.join(pasta, arq))
                removidos += 1
            except Exception as e_rm:
                print(f"  [AVISO] Não removeu {arq}: {e_rm}")

        print(f"  [ZIP] {nome_zip} — {len(arquivos)} arquivo(s) → {removidos} deletado(s)")


def processar_recebidas(driver, wait, codigo, nome, pasta_download):
    """
    Processa Notas Recebidas — mesma lógica das Emitidas.
    Arquivos nomeados com prefixo PADRAO_NACIONAL_RECEBIDA_.
    """
    total_registros  = None
    erros_empresa    = []
    qtd_canceladas   = 0
    qtd_pdf_execucao = 0
    qtd_xml_execucao = 0
    xmls_execucao    = []
    nfse_processadas = set()  # evita baixar mesma nota duas vezes (contingência)

    esperar_clicavel_seguro(
        driver,
        By.XPATH,
        "//a[contains(@href, '/EmissorNacional/Notas/Recebidas')]",
        timeout=40,
        descricao="menu NFS-e Recebidas"
    ).click()

    print("OK - entrou em Recebidas")
    esperar_site_estavel(driver, wait, timeout=35, descricao="tela de NFS-e Recebidas")

    campo_inicio = wait.until(EC.presence_of_element_located((By.ID, "datainicio")))
    campo_inicio.clear()
    campo_inicio.send_keys(DATA_INICIAL)

    campo_fim = wait.until(EC.presence_of_element_located((By.ID, "datafim")))
    campo_fim.clear()
    campo_fim.send_keys(DATA_FINAL)

    wait.until(
        EC.element_to_be_clickable((By.XPATH, "//button[contains(., 'Filtrar')]"))
    ).click()

    print("OK - filtrou Recebidas")

    inicio_espera = time.time()
    nenhum_registro = False
    while time.time() - inicio_espera < 20:
        try:
            texto = driver.find_element(By.TAG_NAME, "body").text
            if "nenhum registro" in texto.lower():
                nenhum_registro = True
                break
            linhas = driver.find_elements(By.XPATH, "//table//tbody//tr")
            if linhas:
                break
        except Exception:
            pass
        time.sleep(0.5)

    esperar_site_estavel(driver, wait, timeout=20, descricao="resultado filtro Recebidas")

    total_registros = obter_total_registros(driver)

    if nenhum_registro or total_registros == 0:
        total_registros = 0
        print("Sem notas recebidas no período.")
        salvar_print_tela(driver, pasta_download, codigo, f"{nome}_recebidas_sem_movimento")
        return total_registros, qtd_canceladas, erros_empresa, qtd_pdf_execucao, qtd_xml_execucao, xmls_execucao

    print(f"Total de notas recebidas no portal: {total_registros}")
    salvar_print_tela(driver, pasta_download, codigo, f"{nome}_recebidas")

    pagina_atual = 1

    while True:
        linhas = driver.find_elements(By.XPATH, "//table//tbody//tr")
        print(f"\nRecebidas — Página {pagina_atual} - notas na tela: {len(linhas)}")

        for i in range(len(linhas)):
            print(f"\nProcessando recebida — página {pagina_atual}, nota {i + 1}")

            try:
                linhas = driver.find_elements(By.XPATH, "//table//tbody//tr")
                linha = linhas[i]

                cancelada = False
                try:
                    html_linha = linha.get_attribute("outerHTML") or ""
                    if "NFS-e cancelada" in html_linha or "nfs-e cancelada" in html_linha.lower():
                        cancelada = True
                except Exception:
                    pass

                if not cancelada:
                    try:
                        elementos_cancelada = linha.find_elements(
                            By.XPATH,
                            ".//*[@title='NFS-e cancelada' or @title='NFS-e Cancelada' "
                            "or @title='Nota cancelada' or @title='Cancelada' "
                            "or contains(@title,'cancelada') or contains(@title,'Cancelada')]"
                        )
                        if elementos_cancelada:
                            cancelada = True
                    except Exception:
                        pass

                if not cancelada:
                    try:
                        cancelada = driver.execute_script("""
                            const linha = arguments[0];
                            const todos = linha.querySelectorAll('*');
                            for (const el of todos) {
                                const t = (el.getAttribute('title') || '').toLowerCase();
                                const a = (el.getAttribute('aria-label') || '').toLowerCase();
                                if (t.includes('cancelad') || a.includes('cancelad')) return true;
                            }
                            return false;
                        """, linha)
                    except Exception:
                        pass

                if cancelada:
                    qtd_canceladas += 1
                    print(f"  [CANCELADA] Nota recebida ignorada.")
                    try:
                        driver.find_element(By.TAG_NAME, "body").send_keys(Keys.ESCAPE)
                    except Exception:
                        pass
                    continue

                url_xml, url_pdf = capturar_links_menu(driver, wait, linha)

                conteudo_xml = baixar_conteudo_direto(
                    driver=driver,
                    url=url_xml,
                    pasta=pasta_download,
                    nome_debug=f"RECEBIDA_XML_p{pagina_atual}_{i + 1}"
                )

                num_nfse = extrair_numero_nfse_do_xml(conteudo_xml)

                if not num_nfse:
                    num_nfse = extrair_numero_nfse_da_linha(linha)
                    if num_nfse:
                        print(f"AVISO - número não veio pelo XML; usando número da linha: {num_nfse}")

                if num_nfse:
                    if num_nfse in nfse_processadas:
                        print(f"  [SKIP] Nota recebida {num_nfse} já processada (contingência/duplicata) — ignorando.")
                        continue
                    nfse_processadas.add(num_nfse)
                    nome_base = f"PADRAO_NACIONAL_RECEBIDA_{num_nfse}_{COMPETENCIA}"
                else:
                    nome_base = f"PADRAO_NACIONAL_RECEBIDA_p{pagina_atual}_{i + 1}_{COMPETENCIA}"

                caminho_xml_salvo = salvar_conteudo_arquivo(
                    conteudo=conteudo_xml,
                    pasta=pasta_download,
                    extensao=".xml",
                    nome_padrao=f"{nome_base}.xml"
                )
                qtd_xml_execucao += 1
                dados_nota = extrair_dados_xml_bytes(conteudo_xml, numero_fallback=num_nfse or "")
                xmls_execucao.append(dados_nota)

                baixar_arquivo_direto(
                    driver=driver,
                    url=url_pdf,
                    pasta=pasta_download,
                    extensao=".pdf",
                    nome_padrao=f"{nome_base}.pdf"
                )
                qtd_pdf_execucao += 1

                try:
                    driver.find_element(By.TAG_NAME, "body").send_keys(Keys.ESCAPE)
                except Exception:
                    pass

            except Exception as e:
                erro = f"Recebidas — Página {pagina_atual}, nota {i + 1}: {e}"
                print(f"ERRO - {erro}")
                erros_empresa.append(erro)

            time.sleep(1)

        if not ir_para_proxima_pagina(driver, pagina_atual):
            print("Recebidas — última página.")
            break

        pagina_atual += 1

    return total_registros, qtd_canceladas, erros_empresa, qtd_pdf_execucao, qtd_xml_execucao, xmls_execucao


def criar_driver(pasta_download):
    options = Options()
    options.add_argument("--disable-features=DownloadBubbleV2")
    options.add_argument("--safebrowsing-disable-download-protection")

    prefs = {
        "download.default_directory": pasta_download,
        "download.prompt_for_download": False,
        "download.directory_upgrade": True,
        "plugins.always_open_pdf_externally": True,
        "safebrowsing.enabled": False,
        "safebrowsing.disable_download_protection": True,
        "profile.default_content_setting_values.automatic_downloads": 1,
        "download_restrictions": 0,
    }

    options.add_experimental_option("prefs", prefs)

    driver = webdriver.Chrome(options=options)
    driver.maximize_window()
    return driver


def _filtrar_traceback_pn(tb_texto):
    """Mantém apenas linhas do script — remove stacktrace de bibliotecas."""
    linhas_out = []
    linhas = tb_texto.split("\n")
    script_kw = "servicos_padrao_nacional"
    capturar_codigo = False
    for linha in linhas:
        if re.match(r'^[A-Za-z][\w\.]*(?:Exception|Error|Warning)', linha.strip()):
            linhas_out.append(f"  {linha.strip()}")
            capturar_codigo = False
            continue
        if "The above exception" in linha:
            linhas_out.append(f"  {linha.strip()}")
            continue
        if 'File "' in linha and script_kw in linha:
            match = re.search(r'line (\d+), in (\S+)', linha)
            if match:
                linhas_out.append(f"  → {match.group(2)} (linha {match.group(1)})")
            capturar_codigo = True
            continue
        if capturar_codigo and linha.strip() and not linha.strip().startswith("File"):
            linhas_out.append(f"      {linha.strip()[:120]}")
            capturar_codigo = False
            continue
        capturar_codigo = False
    return "\n".join(linhas_out) if linhas_out else "  (traceback nao extraido)"


def registrar_diagnostico_pn(codigo, nome, tipo, fase, erro, driver=None, pasta=None,
                               tentativa=1, tema_cert="", status_atual=None, tempo_decorrido=None):
    """Registra erro detalhado por empresa no arquivo de diagnóstico."""
    import traceback as _tb
    caminho = os.path.join(PASTA_DIAG_TECNICO, DIAGNOSTICO_NOME)
    L = 90
    linhas = []
    linhas.append("=" * L)
    linhas.append(f"Data/Hora   : {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}")
    linhas.append(f"Empresa     : {codigo} - {nome}")
    linhas.append(f"Tipo acesso : {tipo}")
    linhas.append(f"Tentativa   : {tentativa}")
    linhas.append(f"Fase        : {fase}")
    if tipo == "CERTIFICADO" and tema_cert:
        linhas.append(f"Cert buscado: {tema_cert}")
    linhas.append("-" * L)
    if erro is not None:
        linhas.append(f"Erro tipo   : {type(erro).__name__}")
        linhas.append(f"Erro msg    : {str(erro)[:300]}")
    linhas.append("-" * L)
    linhas.append("Status ate o erro:")
    if status_atual:
        for campo, val in status_atual.items():
            linhas.append(f"  {campo:<24}: {val if val is not None else 'N/A'}")
    else:
        linhas.append("  (nao disponivel)")
    if tempo_decorrido is not None:
        m = int(tempo_decorrido) // 60
        s = int(tempo_decorrido) % 60
        linhas.append(f"Tempo ate o erro    : {m}min{s:02d}s")
    linhas.append("-" * L)
    if driver is not None:
        try:
            url = driver.current_url
            pagina = url.split("/")[-1].split("?")[0] if "/" in url else url
            linhas.append(f"Pagina no erro     : {pagina}")
            linhas.append(f"URL completa       : {url}")
        except Exception:
            linhas.append("URL no erro        : nao capturada")
        try:
            linhas.append(f"Titulo             : {driver.title}")
        except Exception:
            pass
        try:
            msgs = []
            elementos = driver.find_elements(By.XPATH,
                "//*[contains(@class,'erro') or contains(@class,'error') "
                "or contains(@class,'alert') or contains(@class,'mensagem') "
                "or contains(@class,'warning')]")
            for el in elementos:
                t = (el.text or "").strip()
                if t and 5 < len(t) < 500 and t not in msgs:
                    msgs.append(t)
            if not msgs:
                body = driver.find_element(By.TAG_NAME, "body").text or ""
                for termo in ["Bad Gateway", "Service Unavailable", "Gateway Timeout",
                              "Erro interno", "Indisponivel", "500", "503", "504"]:
                    if termo.lower() in body.lower():
                        idx_t = body.lower().find(termo.lower())
                        trecho = re.sub(r"\s+", " ", body[max(0,idx_t-20):idx_t+200]).strip()
                        msgs.append(trecho)
                        break
            linhas.append("-" * L)
            linhas.append("Erro portal (visivel):")
            if msgs:
                for msg in msgs[:3]:
                    linhas.append(f"  >> {msg}")
            else:
                linhas.append("  [nenhuma mensagem de erro encontrada]")
        except Exception:
            pass
    linhas.append("-" * L)
    linhas.append("Traceback (apenas script):")
    if erro is not None:
        linhas.append(_filtrar_traceback_pn(_tb.format_exc()))
    if pasta:
        linhas.append("-" * L)
        linhas.append(f"Pasta       : {pasta}")
    linhas.append("=" * L)
    linhas.append("")
    _log_diag_pn.append({
        "empresa":   f"{codigo} - {nome}",
        "tipo":      tipo,
        "fase":      fase,
        "erro_tipo": type(erro).__name__ if erro else "",
        "tentativa": tentativa,
    })
    try:
        with open(caminho, "a", encoding="utf-8") as f:
            f.write("\n".join(linhas))
        print(f"    [DIAGNOSTICO] Registrado em: {caminho}")
    except Exception as e_d:
        print(f"    [AVISO] Nao foi possivel salvar diagnostico: {e_d}")


def finalizar_diagnostico_pn():
    """Escreve bloco de resumo final no arquivo de diagnóstico."""
    if not _log_diag_pn:
        return
    caminho = os.path.join(PASTA_DIAG_TECNICO, DIAGNOSTICO_NOME)
    L = 90
    linhas = []
    linhas.append("")
    linhas.append("=" * L)
    linhas.append("  RESUMO FINAL DE ERROS — PADRAO NACIONAL")
    linhas.append("=" * L)
    linhas.append(f"  Total de registros : {len(_log_diag_pn)}")
    linhas.append("")
    por_fase = {}
    for e in _log_diag_pn:
        por_fase.setdefault(e["fase"], []).append(e["empresa"])
    linhas.append("  Por fase:")
    for f, emps in sorted(por_fase.items(), key=lambda x: -len(x[1])):
        linhas.append(f"    {f:<30}: {len(emps)}x")
        for emp in emps:
            linhas.append(f"      - {emp}")
    linhas.append("")
    por_tipo = {}
    for e in _log_diag_pn:
        por_tipo[e["erro_tipo"]] = por_tipo.get(e["erro_tipo"], 0) + 1
    linhas.append("  Por tipo de erro:")
    for tipo, qtd in sorted(por_tipo.items(), key=lambda x: -x[1]):
        linhas.append(f"    {tipo:<40}: {qtd}x")
    linhas.append("=" * L)
    linhas.append("")
    try:
        with open(caminho, "a", encoding="utf-8") as f:
            f.write("\n".join(linhas))
        print(f"\n[DIAGNOSTICO] Resumo final gravado em: {caminho}")
    except Exception as e_d:
        print(f"  [AVISO] Nao foi possivel salvar resumo diagnostico: {e_d}")


def gerar_relatorio_conclusao_pn(linhas_relatorio, tempo_total_s, total_original, tempos_empresa=None):
    """Gera CONCLUSAO_PADRAO_NACIONAL_MMAAAA.txt."""
    tempos_empresa  = tempos_empresa or {}
    agora           = datetime.now().strftime("%d/%m/%Y %H:%M:%S")
    competencia_fmt = f"{COMPETENCIA[:2]}/{COMPETENCIA[2:]}"
    total           = len(linhas_relatorio)
    total_mins      = int(tempo_total_s) // 60
    total_segs      = int(tempo_total_s) % 60
    media_s         = int(tempo_total_s / total) if total else 0
    media_mins      = media_s // 60
    media_s2        = media_s % 60

    def _fmt_t(s):
        s = int(s or 0)
        return f"{s // 60}min{s % 60:02d}s"

    def _st(v):
        v = str(v or "").strip()
        if v == "✓":             return "✓"
        if v == "Sem movimento": return "Sem movimento"
        if v == "X":             return "✗"
        return v or "N/A"

    def _fmt_doc(st, qtd):
        """Formata campo de documento: ✓ (N), Sem movimento ou ✗ (N)."""
        s = str(st or "").strip()
        if s == "Sem movimento": return "Sem movimento"
        if s == "✓":             return f"✓ ({qtd})"
        if s == "X":             return f"✗ ({qtd})"
        return str(qtd) if qtd else "—"

    def _cat(r):
        login    = str(r.get("login_senha", ""))
        emit_st  = str(r.get("emit_pdf_st", ""))
        rec_st   = str(r.get("rec_pdf_st",  ""))
        obs      = str(r.get("observacao",  "")).upper()
        if login == "Login incorreto":  return "LOGIN"
        if login == "CERTIFICADO":      return "CERT"
        if emit_st == "Sem movimento" and rec_st == "Sem movimento": return "SEM_MOV"
        if any(str(r.get(k,""))=="X" for k in ("emit_pdf_st","emit_xml_st","rec_pdf_st","rec_xml_st")):
            return "DIVERG"
        if login == "✓" and emit_st in ("✓","Sem movimento") and rec_st in ("✓","Sem movimento"):
            return "OK"
        if "INSTABILIDADE" in obs or "REPROCESSAR" in obs: return "INSTAVEL"
        return "ERRO"

    ok_lista       = [r for r in linhas_relatorio if _cat(r) == "OK"]
    diverg_lista   = [r for r in linhas_relatorio if _cat(r) == "DIVERG"]
    sem_mov_lista  = [r for r in linhas_relatorio if _cat(r) == "SEM_MOV"]
    login_lista    = [r for r in linhas_relatorio if _cat(r) == "LOGIN"]
    cert_lista     = [r for r in linhas_relatorio if _cat(r) == "CERT"]
    instavel_lista = [r for r in linhas_relatorio if _cat(r) == "INSTAVEL"]
    reprocess_lista= [r for r in linhas_relatorio
                      if "INSTABILIDADE" in str(r.get("observacao","")).upper()
                      or "REPROCESSAR"   in str(r.get("observacao","")).upper()]
    com_erros      = total_original - total

    L = 80
    linhas = []

    # ── CABECALHO ────────────────────────────────────────────────────────────
    linhas.append("=" * L)
    linhas.append("  NFS-e PADRAO NACIONAL  |  RELATORIO DE CONCLUSAO")
    linhas.append(f"  Competencia : {competencia_fmt}")
    linhas.append(f"  Gerado em   : {agora}")
    linhas.append(f"  Tempo total : {total_mins}min{total_segs:02d}s")
    linhas.append("=" * L)

    # ── RESUMO ───────────────────────────────────────────────────────────────
    linhas.append(f"RESUMO - Media/empresa    : {media_mins}min{media_s2:02d}s")
    linhas.append("-" * 60)
    linhas.append(f"  Total empresas                           : {total}")
    linhas.append(f"  Processamento concluido                  : {total_original}")
    linhas.append(f"  Processadas com sucesso com movimento    : {len(ok_lista)}")
    linhas.append(f"  Processadas com sucesso sem movimento    : {len(sem_mov_lista)}")
    linhas.append(f"  Certificado nao localizado               : {len(cert_lista)}")
    linhas.append(f"  Login ou senha incorreto                 : {len(login_lista)}")
    linhas.append(f"  Divergentes                              : {len(diverg_lista)}")
    linhas.append("-" * 60)
    linhas.append("Observacoes:")
    linhas.append(f"  Pasta nao informada                      : 0")
    linhas.append(f"  Com erros                                : {com_erros}")
    linhas.append("")

    # ── 1. EMPRESAS PROCESSADAS COM SUCESSO (COM MOVIMENTO) ──────────────────
    linhas.append("=" * L)
    linhas.append(f"EMPRESAS PROCESSADAS COM SUCESSO ({len(ok_lista)})")
    linhas.append("=" * L)
    for r in ok_lista:
        cod        = r.get("codigo", "")
        t          = tempos_empresa.get(str(cod), 0)
        canceladas = r.get("emit_cancel",  0) or 0
        analise    = r.get("emit_analise", 0) or 0
        linhas.append("")
        linhas.append(f"  Empresa                      : {cod} - {r.get('empresa','')}")
        linhas.append(f"  Emitidas PDF                 : {_fmt_doc(r.get('emit_pdf_st',''), r.get('emit_pdf',''))}")
        linhas.append(f"  Emitidas XML                 : {_fmt_doc(r.get('emit_xml_st',''), r.get('emit_xml',''))}")
        linhas.append(f"  Recebidas PDF                : {_fmt_doc(r.get('rec_pdf_st',''),  r.get('rec_pdf',''))}")
        linhas.append(f"  Recebidas XML                : {_fmt_doc(r.get('rec_xml_st',''),  r.get('rec_xml',''))}")
        if canceladas:
            linhas.append(f"  Canceladas                   : {canceladas}")
        if analise:
            linhas.append(f"  Cancelamento em analise      : {analise}")
        if t:
            linhas.append(f"  Tempo                        : {_fmt_t(t)}")
        linhas.append("  " + "-" * 60)

    # ── 2. EMPRESAS SEM MOVIMENTO — uma linha por empresa ────────────────────
    linhas.append("")
    linhas.append("=" * L)
    linhas.append(f"EMPRESAS PROCESSADAS COM SUCESSO SEM MOVIMENTO NA COMPETENCIA ({len(sem_mov_lista)})")
    linhas.append("=" * L)
    for r in sem_mov_lista:
        cod  = r.get("codigo", "")
        nome = r.get("empresa", "")
        t    = tempos_empresa.get(str(cod), 0)
        linha_sm = f"  {cod} - {nome}"
        linha_sm += f"  |  Observacao: Evidencia sem movimento salvo na pasta"
        if t:
            linha_sm += f"  -  {_fmt_t(t)}"
        linhas.append(linha_sm)

    # ── 3. CERTIFICADO NAO LOCALIZADO ────────────────────────────────────────
    linhas.append("")
    linhas.append("=" * L)
    linhas.append(f"EMPRESAS COM CERTIFICADO NAO LOCALIZADO ({len(cert_lista)})")
    linhas.append("=" * L)
    if cert_lista:
        for r in cert_lista:
            cod = r.get("codigo", "")
            t   = tempos_empresa.get(str(cod), 0)
            linhas.append("")
            linhas.append(f"  Empresa  : {cod} - {r.get('empresa','')}")
            linhas.append(f"  Motivo   : {r.get('observacao','Certificado nao localizado')}")
            if t: linhas.append(f"  Tempo    : {_fmt_t(t)}")
            linhas.append("  " + "-" * 60)
    else:
        linhas.append("NENHUM CERTIFICADO NAO LOCALIZADO REGISTRADO.")

    # ── 4. LOGIN OU SENHA INCORRETO ──────────────────────────────────────────
    linhas.append("")
    linhas.append("=" * L)
    linhas.append(f"EMPRESAS COM LOGIN OU SENHA INCORRETO ({len(login_lista)})")
    linhas.append("=" * L)
    if login_lista:
        for r in login_lista:
            cod = r.get("codigo", "")
            t   = tempos_empresa.get(str(cod), 0)
            linhas.append("")
            linhas.append(f"  Empresa  : {cod} - {r.get('empresa','')}")
            linhas.append(f"  Motivo   : {r.get('observacao','Login/senha incorretos')}")
            if t: linhas.append(f"  Tempo    : {_fmt_t(t)}")
            linhas.append("  " + "-" * 60)
    else:
        linhas.append("NENHUM LOGIN OU SENHA INCORRETO REGISTRADO.")

    # ── 5. PORTAL INSTAVEL ───────────────────────────────────────────────────
    linhas.append("")
    linhas.append("=" * L)
    if instavel_lista:
        linhas.append(f"EMPRESAS PORTAL INSTAVEL ({len(instavel_lista)})")
        linhas.append("=" * L)
        for r in instavel_lista:
            cod = r.get("codigo", "")
            t   = tempos_empresa.get(str(cod), 0)
            linhas.append("")
            linhas.append(f"  Empresa       : {cod} - {r.get('empresa','')}")
            linhas.append(f"  Emitidas PDF  : {_st(r.get('emit_pdf_st'))}")
            linhas.append(f"  Emitidas XML  : {_st(r.get('emit_xml_st'))}")
            linhas.append(f"  Recebidas PDF : {_st(r.get('rec_pdf_st'))}")
            linhas.append(f"  Recebidas XML : {_st(r.get('rec_xml_st'))}")
            obs = r.get("observacao", "")
            if obs: linhas.append(f"  Motivo        : {obs}")
            if t:   linhas.append(f"  Tempo         : {_fmt_t(t)}")
            linhas.append("  " + "-" * 60)
    else:
        linhas.append("EMPRESAS PORTAL INSTAVEL (0)")
        linhas.append("=" * L)
        linhas.append("NENHUMA INSTABILIDADE DE PORTAL REGISTRADA.")

    # ── 6. DIVERGENTES (quando houver) ───────────────────────────────────────
    if diverg_lista:
        linhas.append("")
        linhas.append("=" * L)
        linhas.append(f"EMPRESAS DIVERGENTES ({len(diverg_lista)})")
        linhas.append("=" * L)
        for r in diverg_lista:
            cod        = r.get("codigo", "")
            t          = tempos_empresa.get(str(cod), 0)
            canceladas = r.get("emit_cancel",  0) or 0
            analise    = r.get("emit_analise", 0) or 0
            linhas.append("")
            linhas.append(f"  Empresa                      : {cod} - {r.get('empresa','')}")
            linhas.append(f"  Emitidas PDF                 : {_fmt_doc(r.get('emit_pdf_st',''), r.get('emit_pdf',''))}")
            linhas.append(f"  Emitidas XML                 : {_fmt_doc(r.get('emit_xml_st',''), r.get('emit_xml',''))}")
            linhas.append(f"  Recebidas PDF                : {_fmt_doc(r.get('rec_pdf_st',''),  r.get('rec_pdf',''))}")
            linhas.append(f"  Recebidas XML                : {_fmt_doc(r.get('rec_xml_st',''),  r.get('rec_xml',''))}")
            if canceladas:
                linhas.append(f"  Canceladas                   : {canceladas}")
            if analise:
                linhas.append(f"  Cancelamento em analise      : {analise}")
            linhas.append(f"  Observacao                   : Divergencia apos reprocessamento — verificar")
            if t:
                linhas.append(f"  Tempo                        : {_fmt_t(t)}")
            linhas.append("  " + "-" * 60)

    # ── 7. REPROCESSADAS (quando houver) ─────────────────────────────────────
    if reprocess_lista:
        linhas.append("")
        linhas.append("=" * L)
        linhas.append(f"EMPRESAS REPROCESSADAS ({len(reprocess_lista)})")
        linhas.append("=" * L)
        for r in reprocess_lista:
            cod      = r.get("codigo", "")
            tem_erro = str(r.get("emit_pdf_st","")) == "X" or str(r.get("rec_pdf_st","")) == "X"
            situacao = ("Reprocessada — bloqueio persistiu, reprocessamento manual necessario"
                        if tem_erro else "Concluida com sucesso no reprocessamento")
            linhas.append("")
            linhas.append(f"  Empresa   : {cod} - {r.get('empresa','')}")
            linhas.append(f"  Situacao  : {situacao}")
            linhas.append("  " + "-" * 60)

    # ── RODAPE ───────────────────────────────────────────────────────────────
    linhas.append("")
    linhas.append("=" * L)
    linhas.append("  FIM DO RELATORIO DE CONCLUSAO  |  NFS-e PADRAO NACIONAL")
    linhas.append("=" * L)

    caminho = os.path.join(PASTA_BASE, CONCLUSAO_NOME)
    try:
        with open(caminho, "w", encoding="utf-8") as f:
            f.write("\n".join(linhas))
        print(f"\n[CONCLUSAO] Relatorio de conclusao salvo em:\n  {caminho}")
    except Exception as e_c:
        print(f"  [AVISO] Nao foi possivel salvar relatorio de conclusao: {e_c}")


def salvar_relatorio_erros(erros_globais):
    os.makedirs(PASTA_DIAG_TECNICO, exist_ok=True)
    """Gera arquivo TXT com todos os erros do processo para análise posterior."""
    caminho = os.path.join(PASTA_DIAG_TECNICO, f"RELATORIO_ERROS_PADRAO_NACIONAL_{COMPETENCIA}.txt")
    linhas = []
    linhas.append("=" * 80)
    linhas.append(f"RELATÓRIO DE ERROS — PADRÃO NACIONAL — COMPETÊNCIA {COMPETENCIA[:2]}/{COMPETENCIA[2:]}")
    linhas.append(f"Gerado em: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}")
    linhas.append("=" * 80)
    linhas.append("")

    if not erros_globais:
        linhas.append("Nenhum erro registrado. Processo concluído com sucesso.")
    else:
        for item in erros_globais:
            linhas.append(f"Empresa  : {item['codigo']} - {item['empresa']}")
            linhas.append(f"Tipo     : {item['tipo']}")
            linhas.append(f"Erro     : {item['erro']}")
            linhas.append("-" * 80)

    linhas.append("")
    linhas.append(f"Total de erros: {len(erros_globais)}")

    with open(caminho, "w", encoding="utf-8") as f:
        f.write("\n".join(linhas))

    print(f"\n[ERROS] Relatório de erros salvo em:\n  {caminho}")
    return caminho




def _salvar_erro_pastas_preflight(caminho_senhas, detalhes, pasta_relatorio=None):
    """
    Salva erro grave quando a coluna PASTA da SENHAS.xlsx não está 100% válida.
    A automação deve travar antes de abrir navegador se qualquer linha preenchida
    estiver sem pasta ou com pasta inexistente/inacessível.
    """
    destinos = []

    for pasta in (_SCRIPT_DIR, pasta_relatorio, PASTA_BASE):
        try:
            if pasta and os.path.isdir(pasta) and pasta not in destinos:
                destinos.append(pasta)
        except Exception:
            pass

    if not destinos:
        destinos = [_SCRIPT_DIR]

    mensagem = (
        "ERRO GRAVE: a coluna PASTA deve estar preenchida e valida em TODAS as linhas "
        "da planilha SENHAS.xlsx antes da automacao iniciar."
    )

    linhas_conclusao = [
        "=" * 70,
        "CONCLUSAO — ERRO GRAVE DE CONFIGURACAO",
        "=" * 70,
        f"Competencia : {COMPETENCIA[:2]}/{COMPETENCIA[2:]}",
        f"Data/Hora   : {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}",
        f"SENHAS.xlsx : {caminho_senhas}",
        "",
        mensagem,
        "",
        "STATUS FINAL: PROCESSO INTERROMPIDO ANTES DO INICIO.",
        "Motivo: existe pelo menos uma linha com PASTA vazia, inexistente ou inacessivel.",
        "Nenhum navegador foi aberto.",
        "Nenhum download foi iniciado.",
        "Nenhum arquivo foi salvo em pasta padrao por fallback.",
        "",
        "LINHAS COM ERRO:",
    ]

    for item in detalhes:
        linhas_conclusao.append(f"- {item}")

    linhas_erros = [
        "=" * 80,
        f"RELATÓRIO DE ERROS — PADRÃO NACIONAL — COMPETÊNCIA {COMPETENCIA[:2]}/{COMPETENCIA[2:]}",
        f"Gerado em: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}",
        "=" * 80,
        "",
        "ERRO GRAVE DE CONFIGURACAO",
        "Tipo     : PASTAS OBRIGATORIAS NA PLANILHA SENHAS",
        f"SENHAS   : {caminho_senhas}",
        f"Erro     : {mensagem}",
        "",
        "Acao requerida:",
        "Corrigir a coluna PASTA em TODAS as linhas preenchidas da planilha SENHAS.xlsx antes de executar novamente.",
        "",
        "Detalhes:",
    ]

    for item in detalhes:
        linhas_erros.append(f"- {item}")

    linhas_erros.append("")
    linhas_erros.append(f"Total de erros: {len(detalhes)}")

    caminhos_salvos = []
    for pasta in destinos:
        try:
            caminho_conclusao = os.path.join(pasta, CONCLUSAO_NOME)
            with open(caminho_conclusao, "w", encoding="utf-8") as f:
                f.write("\n".join(linhas_conclusao))
            caminhos_salvos.append(caminho_conclusao)
        except Exception:
            pass

        try:
            caminho_erros = os.path.join(PASTA_DIAG_TECNICO, f"RELATORIO_ERROS_PADRAO_NACIONAL_{COMPETENCIA}.txt")
            os.makedirs(PASTA_DIAG_TECNICO, exist_ok=True)
            with open(caminho_erros, "w", encoding="utf-8") as f:
                f.write("\n".join(linhas_erros))
            caminhos_salvos.append(caminho_erros)
        except Exception:
            pass

    print()
    print("=" * 70)
    print("  ERRO GRAVE -- PASTAS DA PLANILHA SENHAS NAO VALIDAS")
    print("=" * 70)
    print("  A automacao foi interrompida antes de abrir o navegador.")
    print("  Motivo: TODAS as linhas preenchidas precisam ter PASTA valida.")
    print(f"  SENHAS.xlsx lido: {caminho_senhas}")
    print("")
    for item in detalhes:
        print(f"  - {item}")
    if caminhos_salvos:
        print("  Relatorios de erro salvos em:")
        for caminho in caminhos_salvos:
            print(f"  {caminho}")
    print("=" * 70)


def _validar_todas_pastas_senhas(ws_senhas, caminho_senhas):
    """
    Valida a coluna PASTA antes de qualquer processamento.
    Regra: toda linha com CODIGO ou NOME preenchido deve ter PASTA preenchida,
    existente e acessivel. Se qualquer linha falhar, trava tudo.
    """
    erros = []
    primeira_pasta_valida = None

    for idx_linha, row in enumerate(ws_senhas.iter_rows(min_row=2, values_only=True), start=2):
        codigo = row[0] if len(row) > 0 else ""
        nome = row[1] if len(row) > 1 else ""

        # Linha totalmente vazia — ignora.
        if not codigo and not nome:
            continue

        pasta = ""
        if len(row) > 8 and row[8] is not None:
            pasta = str(row[8]).strip()
            if pasta in ("", "None"):
                pasta = ""

        identificacao = f"Linha {idx_linha} | {str(codigo).strip()} - {str(nome).strip()}"

        if not pasta:
            erros.append(f"{identificacao}: coluna PASTA vazia.")
            continue

        if not os.path.isdir(pasta):
            erros.append(f"{identificacao}: pasta informada nao existe ou nao esta acessivel: {pasta}")
            continue

        try:
            if not os.access(pasta, os.W_OK):
                erros.append(f"{identificacao}: pasta sem permissao de gravacao: {pasta}")
                continue
        except Exception:
            pass

        if primeira_pasta_valida is None:
            primeira_pasta_valida = pasta

    if erros:
        _salvar_erro_pastas_preflight(caminho_senhas, erros, pasta_relatorio=primeira_pasta_valida)
        raise Exception("ERRO GRAVE -- PASTAS DA PLANILHA SENHAS NAO VALIDAS")

    if not primeira_pasta_valida:
        detalhes = ["Nenhuma empresa preenchida foi encontrada na planilha SENHAS.xlsx."]
        _salvar_erro_pastas_preflight(caminho_senhas, detalhes, pasta_relatorio=None)
        raise Exception("ERRO GRAVE -- NENHUMA EMPRESA ENCONTRADA NA PLANILHA SENHAS")

    return primeira_pasta_valida

import queue   as _q_mod
import tkinter as _tk_mod

# ─────────────────────────────────────────────────────────────────────────────
# JANELA DE PROGRESSO — RWA Soluções
# ─────────────────────────────────────────────────────────────────────────────
class RWAProgressWindow:
    """Janela de progresso profissional — RWA Soluções."""

    def __init__(self):
        self._queue       = _q_mod.Queue()
        self._pronto      = threading.Event()
        self._fechado     = threading.Event()
        self._inicio      = None
        self._repro_fila  = 0   # total de empresas enviadas para reprocessamento
        self._repro_idx   = 0   # índice da empresa sendo reprocessada agora
        self._total       = 0   # total de empresas na lista
        self._thread = threading.Thread(target=self._run, daemon=True)
        self._thread.start()
        self._pronto.wait(timeout=4)

    def write(self, texto):
        if texto:
            self._queue.put(("LOG", texto))
            # Primeira passagem — EMPRESA X DE Y: cod - nome
            m = re.search(r'EMPRESA\s+(\d+)\s+DE\s+(\d+):\s+\S+\s+-\s+(.+)', texto)
            if m:
                self._queue.put(("PROG", int(m.group(1)), int(m.group(2)), m.group(3).strip()))
                return
            # Empresa enviada para fila de reprocessamento
            if re.search(r'\[REPROCESSAR\]|\[CONFRONTO\]', texto):
                self._queue.put(("REPRO_ADD",))
                return
            # Reprocessamento iniciado — REPROCESSAMENTO: X DE Y: cod - nome
            m2 = re.search(r'REPROCESSAMENTO:\s+(\d+)\s+DE\s+(\d+):\s+\S+\s+-\s+(.+)', texto)
            if m2:
                self._queue.put(("REPRO_START", int(m2.group(1)), int(m2.group(2)), m2.group(3).strip()))

    def flush(self):
        pass

    def finalizar(self):
        self._queue.put(("FIM",))
        self._fechado.wait()

    def _run(self):
        import time as _time

        root = _tk_mod.Tk()
        root.title("RWA Soluções — NFS-e Padrão Nacional")
        root.resizable(False, False)
        root.configure(bg="#13131f")

        W, H = 660, 600
        root.geometry(f"{W}x{H}+{(root.winfo_screenwidth()-W)//2}+{(root.winfo_screenheight()-H)//2}")

        _tk_mod.Frame(root, bg="#4f46e5", height=5).pack(fill="x")

        _hdr = _tk_mod.Frame(root, bg="#13131f")
        _hdr.pack(fill="x", padx=30, pady=(18, 0))

        _tk_mod.Label(_hdr, text="📄", font=("Arial", 22),
                      bg="#13131f", fg="#4f46e5").pack(side="left", padx=(0, 12))

        _col = _tk_mod.Frame(_hdr, bg="#13131f")
        _col.pack(side="left")
        _tk_mod.Label(_col, text="NFS-e Padrão Nacional", font=("Arial", 15, "bold"),
                      bg="#13131f", fg="#ffffff", anchor="w").pack(anchor="w")
        _tk_mod.Label(_col, text="Automação da NFS-e Padrão Nacional",
                      font=("Arial", 9), bg="#13131f", fg="#555577", anchor="w").pack(anchor="w")

        self._lbl_timer = _tk_mod.Label(_hdr, text="00:00:00", font=("Consolas", 13, "bold"),
                                         bg="#13131f", fg="#333355")
        self._lbl_timer.pack(side="right")

        _tk_mod.Frame(root, bg="#1e1e30", height=1).pack(fill="x", padx=30, pady=(14, 0))

        _card = _tk_mod.Frame(root, bg="#1a1a2e")
        _card.pack(fill="x", padx=30, pady=(14, 0))
        _tk_mod.Label(_card, text="EMPRESA ATUAL", font=("Arial", 7, "bold"),
                      bg="#1a1a2e", fg="#333366", anchor="w").pack(anchor="w", padx=14, pady=(10, 2))
        self._lbl_emp = _tk_mod.Label(_card, text="Iniciando...",
                                       font=("Arial", 11, "bold"),
                                       bg="#1a1a2e", fg="#aaaadd", anchor="w")
        self._lbl_emp.pack(anchor="w", padx=14, pady=(0, 10))

        _fp = _tk_mod.Frame(root, bg="#13131f")
        _fp.pack(fill="x", padx=30, pady=(14, 0))
        _topo = _tk_mod.Frame(_fp, bg="#13131f")
        _topo.pack(fill="x")
        _tk_mod.Label(_topo, text="PROGRESSO", font=("Arial", 7, "bold"),
                      bg="#13131f", fg="#333366").pack(side="left")
        self._lbl_pct = _tk_mod.Label(_topo, text="0 / 0",
                                       font=("Arial", 8, "bold"),
                                       bg="#13131f", fg="#4f46e5")
        self._lbl_pct.pack(side="right")
        _tk_mod.Frame(_fp, bg="#1e1e30", height=1).pack(fill="x", pady=(4, 6))
        self._canvas = _tk_mod.Canvas(_fp, height=10, bg="#1a1a2e",
                                       highlightthickness=0, bd=0)
        self._canvas.pack(fill="x")
        self._barra       = self._canvas.create_rectangle(0, 0, 0, 10, fill="#4f46e5", outline="")
        self._barra_shine = self._canvas.create_rectangle(0, 0, 0,  3, fill="#7c73f5", outline="")

        # Label de reprocessamento — aparece abaixo da barra quando necessário
        self._lbl_repro = _tk_mod.Label(
            root, text="",
            font=("Arial", 9, "bold"),
            bg="#13131f", fg="#f59e0b", anchor="w"
        )
        self._lbl_repro.pack(fill="x", padx=30, pady=(6, 0))

        _fl = _tk_mod.Frame(root, bg="#13131f")
        _fl.pack(fill="both", expand=True, padx=30, pady=(8, 0))
        _tk_mod.Label(_fl, text="LOG DE EXECUÇÃO", font=("Arial", 7, "bold"),
                      bg="#13131f", fg="#333366", anchor="w").pack(anchor="w")
        _tk_mod.Frame(_fl, bg="#1e1e30", height=1).pack(fill="x", pady=(4, 6))
        _fr_txt = _tk_mod.Frame(_fl, bg="#0d0d1a")
        _fr_txt.pack(fill="both", expand=True)
        self._txt = _tk_mod.Text(_fr_txt, bg="#0d0d1a", fg="#7777aa",
                                  font=("Consolas", 8), relief="flat",
                                  state="disabled", wrap="none",
                                  height=12,
                                  selectbackground="#2a2a4a")
        self._txt.tag_config("ok",    foreground="#22c55e")
        self._txt.tag_config("erro",  foreground="#ef4444")
        self._txt.tag_config("aviso", foreground="#f59e0b")
        self._txt.tag_config("info",  foreground="#6366f1")
        _sb = _tk_mod.Scrollbar(_fr_txt, bg="#1a1a2e", troughcolor="#13131f",
                                 command=self._txt.yview)
        self._txt.configure(yscrollcommand=_sb.set)
        _sb.pack(side="right", fill="y")
        self._txt.pack(fill="both", expand=True, padx=6, pady=6)

        _rod = _tk_mod.Frame(root, bg="#13131f")
        _rod.pack(fill="x", padx=30, pady=(10, 14))
        _tk_mod.Label(_rod, text="Desenvolvido por RWA Soluções",
                      font=("Arial", 7), bg="#13131f", fg="#252540").pack(side="left")
        self._btn = _tk_mod.Button(_rod, text="⏳  Processando...",
                                    font=("Arial", 9, "bold"),
                                    bg="#252538", fg="#444466",
                                    relief="flat", padx=18, pady=5,
                                    state="disabled", cursor="arrow")
        self._btn.pack(side="right")

        self._root  = root
        self._inicio = _time.time()
        self._pronto.set()
        root.after(50,   self._poll)
        root.after(1000, self._tick)
        root.mainloop()
        self._fechado.set()

    def _tick(self):
        import time as _time
        if self._inicio and not self._fechado.is_set():
            s = int(_time.time() - self._inicio)
            self._lbl_timer.configure(
                text=f"{s//3600:02d}:{(s%3600)//60:02d}:{s%60:02d}",
                fg="#4f46e5")
            self._root.after(1000, self._tick)

    def _poll(self):
        try:
            while True:
                item = self._queue.get_nowait()

                if item[0] == "LOG":
                    linha = item[1]
                    self._txt.configure(state="normal")
                    tag = ("ok"    if any(x in linha for x in ["[OK]", "✓", "sucesso"]) else
                           "erro"  if any(x in linha for x in ["[ERRO]", "✗", "BLOQUEIO"]) else
                           "aviso" if any(x in linha for x in ["[AVISO]", "REPROCESSAMENTO", "instável",
                                                                "[REPROCESSAR]", "[CONFRONTO]"]) else
                           "info"  if any(x in linha for x in ["[INFO]", "[LOGIN]", "[PDF]", "[TEMPO]"]) else
                           None)
                    self._txt.insert("end", linha, tag) if tag else self._txt.insert("end", linha)
                    self._txt.see("end")
                    self._txt.configure(state="disabled")

                elif item[0] == "PROG":
                    _, atual, total, nome = item
                    if total: self._total = total
                    pct    = int((atual / total) * 100) if total else 0
                    fill_w = int(598 * pct / 100)
                    self._canvas.coords(self._barra,       0, 0, fill_w, 10)
                    self._canvas.coords(self._barra_shine, 0, 0, fill_w,  3)
                    # Restaura cor azul caso tenha ficado âmbar de um repro anterior
                    self._canvas.itemconfig(self._barra,       fill="#4f46e5")
                    self._canvas.itemconfig(self._barra_shine, fill="#7c73f5")
                    if self._repro_fila > 0:
                        self._lbl_pct.configure(
                            text=f"{atual} / {total}  ({pct}%)  +  {self._repro_fila} repro",
                            fg="#4f46e5"
                        )
                    else:
                        self._lbl_pct.configure(text=f"{atual} / {total}  ({pct}%)", fg="#4f46e5")
                    self._lbl_emp.configure(text=nome, fg="#ffffff")

                elif item[0] == "REPRO_ADD":
                    self._repro_fila += 1
                    self._lbl_repro.configure(
                        text=f"⟳  {self._repro_fila} empresa(s) na fila de reprocessamento",
                        fg="#f59e0b"
                    )

                elif item[0] == "REPRO_START":
                    _, idx, total, nome = item
                    self._repro_idx = idx
                    self._lbl_emp.configure(
                        text=f"⟳  REPROCESSANDO ({self._repro_idx}/{self._repro_fila})  —  {nome}",
                        fg="#f59e0b"
                    )
                    self._lbl_repro.configure(
                        text=f"⟳  Reprocessando {self._repro_idx} de {self._repro_fila}  —  {nome}",
                        fg="#f59e0b"
                    )
                    # Barra vira âmbar durante o reprocessamento
                    self._canvas.itemconfig(self._barra,       fill="#f59e0b")
                    self._canvas.itemconfig(self._barra_shine, fill="#fbbf24")

                elif item[0] == "FIM":
                    self._canvas.coords(self._barra,       0, 0, 598, 10)
                    self._canvas.coords(self._barra_shine, 0, 0, 598,  3)
                    self._canvas.itemconfig(self._barra,       fill="#22c55e")
                    self._canvas.itemconfig(self._barra_shine, fill="#4ade80")
                    self._lbl_emp.configure(text="✅  Concluído com sucesso!", fg="#22c55e")
                    self._lbl_pct.configure(text="100%", fg="#22c55e")
                    self._lbl_timer.configure(fg="#22c55e")
                    if self._repro_fila > 0:
                        self._lbl_repro.configure(
                            text=f"✅  {self._total} empresa(s) processada(s) — concluído",
                            fg="#22c55e"
                        )
                    self._inicio = None
                    self._btn.configure(text="Concluído", state="normal",
                                        bg="#22c55e", fg="#0a0a0a",
                                        cursor="hand2", command=self._root.destroy)
        except _q_mod.Empty:
            pass
        if not self._fechado.is_set():
            self._root.after(50, self._poll)



def _render_html_email_rwa(tag_texto, tag_bg, tag_fg, titulo, intro, secoes, callout, intro_destaque=False):
    """Monta HTML inline-friendly do email RWA Soluções.

    secoes: lista de tuplas (titulo_secao_ou_None, [(label, valor, destacar_vermelho), ...])
    callout: tuple (bg, borda, fg, texto)
    intro_destaque: se True, aplica negrito + verde sobrio + icone na frase do intro.
    """
    partes_secoes = []
    for titulo_secao, linhas in secoes:
        if titulo_secao:
            partes_secoes.append(
                f'<tr><td style="padding:14px 0 4px 0;font-family:Arial,Helvetica,sans-serif;'
                f'font-size:10px;color:#6B6B6B;letter-spacing:1px;text-transform:uppercase;'
                f'font-weight:bold;">{titulo_secao}</td></tr>'
            )
        partes_secoes.append('<tr><td style="padding:0;"><table width="100%" cellpadding="0" cellspacing="0" border="0" style="border-collapse:collapse;">')
        for idx, item in enumerate(linhas):
            label, valor, destacar = item
            eh_ultima = (idx == len(linhas) - 1)
            borda = "" if eh_ultima else "border-bottom:1px solid #EFEFEF;"
            cor_valor = "#A32D2D" if destacar else "#1a1a1a"
            partes_secoes.append(
                f'<tr>'
                f'<td style="padding:8px 0;{borda}font-family:Arial,Helvetica,sans-serif;'
                f'font-size:13px;color:#6B6B6B;width:55%;">{label}</td>'
                f'<td style="padding:8px 0;{borda}font-family:Arial,Helvetica,sans-serif;'
                f'font-size:13px;color:{cor_valor};text-align:right;font-weight:bold;word-break:break-word;overflow-wrap:anywhere;">{valor}</td>'
                f'</tr>'
            )
        partes_secoes.append('</table></td></tr>')

    html_secoes = "".join(partes_secoes)
    callout_bg, callout_borda, callout_fg, callout_texto = callout

    _intro_cor   = "#1F6B43" if intro_destaque else "#1a1a1a"
    _intro_peso  = "bold"    if intro_destaque else "normal"
    _intro_icone = "✓ "      if intro_destaque else ""

    return (
        f'<!DOCTYPE html PUBLIC "-//W3C//DTD XHTML 1.0 Transitional//EN" '
        f'"http://www.w3.org/TR/xhtml1/DTD/xhtml1-transitional.dtd">'
        f'<html><head><meta charset="utf-8"/><meta name="viewport" content="width=device-width, initial-scale=1.0"/></head>'
        f'<body style="margin:0;padding:0;background-color:#f4f4f4;-webkit-text-size-adjust:100%;-ms-text-size-adjust:100%;">'
        f'<table width="100%" cellpadding="0" cellspacing="0" border="0" style="width:100%;background-color:#f4f4f4;padding:12px 0;border-collapse:collapse;">'
        f'<tr><td align="center" style="padding:0 10px;box-sizing:border-box;">'
        f'<table width="100%" cellpadding="0" cellspacing="0" border="0" '
        f'style="width:100%;max-width:600px;background-color:#ffffff;border:1px solid #e0e0e0;border-radius:6px;overflow:hidden;box-sizing:border-box;border-collapse:separate;">'
        # ── header escuro ──
        f'<tr><td style="background-color:#0F1B2D;padding:18px 16px;box-sizing:border-box;">'
        f'<div style="font-family:Arial,Helvetica,sans-serif;font-size:11px;color:#7FB3E0;'
        f'letter-spacing:1.5px;font-weight:bold;margin-bottom:8px;">RWA SOLUÇÕES</div>'
        f'<div style="display:inline-block;background-color:{tag_bg};color:{tag_fg};'
        f'font-family:Arial,Helvetica,sans-serif;font-size:10px;padding:3px 8px;'
        f'border-radius:3px;font-weight:bold;letter-spacing:0.5px;margin-bottom:10px;">{tag_texto}</div>'
        f'<div style="font-family:Arial,Helvetica,sans-serif;font-size:20px;color:#FFFFFF;'
        f'font-weight:bold;line-height:26px;margin-top:4px;word-break:break-word;overflow-wrap:anywhere;">{titulo}</div>'
        f'<div style="font-family:Arial,Helvetica,sans-serif;font-size:13px;color:#B5D4F4;'
        f'line-height:18px;margin-top:2px;word-break:break-word;overflow-wrap:anywhere;">NFS-e Padrão Nacional</div>'
        f'</td></tr>'
        # ── corpo ──
        f'<tr><td style="padding:16px;box-sizing:border-box;">'
        f'<div style="font-family:Arial,Helvetica,sans-serif;font-size:14px;color:{_intro_cor};'
        f'line-height:1.5;margin-bottom:12px;font-weight:{_intro_peso};">{_intro_icone}{intro}</div>'
        f'<table width="100%" cellpadding="0" cellspacing="0" border="0" style="border-collapse:collapse;">'
        f'{html_secoes}'
        f'</table>'
        # ── callout ──
        f'<table width="100%" cellpadding="0" cellspacing="0" border="0" style="margin-top:14px;border-collapse:collapse;">'
        f'<tr><td style="background-color:{callout_bg};border-left:3px solid {callout_borda};'
        f'padding:10px 12px;font-family:Arial,Helvetica,sans-serif;font-size:12px;'
        f'color:{callout_fg};line-height:1.5;word-break:break-word;overflow-wrap:anywhere;">{callout_texto}</td></tr>'
        f'</table>'
        # ── rodapé ──
        f'<table width="100%" cellpadding="0" cellspacing="0" border="0" style="margin-top:16px;border-top:1px solid #EFEFEF;">'
        f'<tr><td style="padding-top:10px;font-family:Arial,Helvetica,sans-serif;font-size:12px;color:#888;">'
        f'<span style="color:#0F1B2D;font-weight:bold;">RWA Soluções</span><br/>'
        f'<span style="font-size:11px;">Automação fiscal para escritórios contábeis</span>'
        f'</td></tr></table>'
        f'</td></tr>'
        f'</table>'
        f'</td></tr></table>'
        f'</body></html>'
    )

def _enviar_email_rwa(conta, senha, destino, assunto, corpo_html, corpo_plain):
    """Envia email multipart/alternative (HTML + plain) via Gmail SMTP."""
    msg = _MMulti_h("alternative")
    msg["From"]    = conta
    msg["To"]      = destino
    msg["Subject"] = assunto
    msg.attach(_MText_h(corpo_plain, "plain", "utf-8"))
    msg.attach(_MText_h(corpo_html,  "html",  "utf-8"))
    srv = _smtp_h.SMTP("smtp.gmail.com", 587)
    srv.starttls()
    srv.login(conta, senha)
    srv.sendmail(conta, destino, msg.as_string())
    srv.quit()

def _render_html_email_interrupcao(tag_texto, tag_bg, tag_fg, titulo, intro, secoes, callout, intro_destaque=False):
    """Monta HTML inline-friendly do email RWA Soluções — Interrupção."""
    partes_secoes = []
    for titulo_secao, linhas in secoes:
        if titulo_secao:
            partes_secoes.append(
                f'<tr><td style="padding:14px 0 4px 0;font-family:Arial,Helvetica,sans-serif;'
                f'font-size:10px;color:#6B6B6B;letter-spacing:1px;text-transform:uppercase;'
                f'font-weight:bold;">{titulo_secao}</td></tr>'
            )
        partes_secoes.append('<tr><td style="padding:0;"><table width="100%" cellpadding="0" cellspacing="0" border="0" style="border-collapse:collapse;">')
        for idx, item in enumerate(linhas):
            label, valor, destacar = item
            eh_ultima = (idx == len(linhas) - 1)
            borda = "" if eh_ultima else "border-bottom:1px solid #EFEFEF;"
            cor_valor = "#A32D2D" if destacar else "#1a1a1a"
            partes_secoes.append(
                f'<tr>'
                f'<td style="padding:8px 0;{borda}font-family:Arial,Helvetica,sans-serif;'
                f'font-size:13px;color:#6B6B6B;width:55%;">{label}</td>'
                f'<td style="padding:8px 0;{borda}font-family:Arial,Helvetica,sans-serif;'
                f'font-size:13px;color:{cor_valor};text-align:right;font-weight:bold;word-break:break-word;overflow-wrap:anywhere;">{valor}</td>'
                f'</tr>'
            )
        partes_secoes.append('</table></td></tr>')
    html_secoes = "".join(partes_secoes)
    callout_bg, callout_borda, callout_fg, callout_texto = callout
    _intro_cor   = "#A32D2D" if intro_destaque else "#1a1a1a"
    _intro_peso  = "bold"    if intro_destaque else "normal"
    _intro_icone = "⚠ "      if intro_destaque else ""
    return (
        f'<!DOCTYPE html PUBLIC "-//W3C//DTD XHTML 1.0 Transitional//EN" '
        f'"http://www.w3.org/TR/xhtml1/DTD/xhtml1-transitional.dtd">'
        f'<html><head><meta charset="utf-8"/>'
        f'<meta name="color-scheme" content="light"/>'
        f'<meta name="supported-color-schemes" content="light"/>'
        f'</head>'
        f'<body style="margin:0;padding:0;background-color:#f4f4f4;" bgcolor="#f4f4f4">'
        f'<table width="100%" cellpadding="0" cellspacing="0" border="0" bgcolor="#f4f4f4" style="background-color:#f4f4f4;padding:20px 0;">'
        f'<tr><td align="center" style="padding:0 10px;box-sizing:border-box;">'
        f'<table width="100%" cellpadding="0" cellspacing="0" border="0" bgcolor="#ffffff" '
        f'style="width:100%;max-width:600px;background-color:#ffffff;border:1px solid #e0e0e0;border-radius:6px;overflow:hidden;box-sizing:border-box;border-collapse:separate;">'
        f'<tr><td bgcolor="#0F1B2D" style="background-color:#0F1B2D;padding:20px 18px;">'
        f'<div style="font-family:Arial,Helvetica,sans-serif;font-size:11px;color:#7FB3E0;'
        f'letter-spacing:1.5px;font-weight:bold;margin-bottom:8px;">RWA SOLUÇÕES</div>'
        f'<div style="display:inline-block;background-color:{tag_bg};color:{tag_fg};'
        f'font-family:Arial,Helvetica,sans-serif;font-size:10px;padding:3px 8px;'
        f'border-radius:3px;font-weight:bold;letter-spacing:0.5px;margin-bottom:10px;">{tag_texto}</div>'
        f'<div style="font-family:Arial,Helvetica,sans-serif;font-size:20px;color:#FFFFFF;'
        f'font-weight:bold;line-height:26px;margin-top:4px;word-break:break-word;overflow-wrap:anywhere;">{titulo}</div>'
        f'<div style="font-family:Arial,Helvetica,sans-serif;font-size:13px;color:#B5D4F4;'
        f'line-height:18px;margin-top:2px;word-break:break-word;overflow-wrap:anywhere;">NFS-e Padrão Nacional</div>'
        f'</td></tr>'
        f'<tr><td bgcolor="#ffffff" style="padding:18px;background-color:#ffffff;">'
        f'<div style="font-family:Arial,Helvetica,sans-serif;font-size:14px;color:{_intro_cor};'
        f'line-height:1.5;margin-bottom:12px;font-weight:{_intro_peso};">{_intro_icone}{intro}</div>'
        f'<table width="100%" cellpadding="0" cellspacing="0" border="0" style="border-collapse:collapse;">'
        f'{html_secoes}'
        f'</table>'
        f'<table width="100%" cellpadding="0" cellspacing="0" border="0" style="margin-top:14px;border-collapse:collapse;">'
        f'<tr><td bgcolor="{callout_bg}" style="background-color:{callout_bg};border-left:3px solid {callout_borda};'
        f'padding:10px 12px;font-family:Arial,Helvetica,sans-serif;font-size:12px;'
        f'color:{callout_fg};line-height:1.5;word-break:break-word;overflow-wrap:anywhere;">{callout_texto}</td></tr>'
        f'</table>'
        f'<table width="100%" cellpadding="0" cellspacing="0" border="0" style="margin-top:16px;border-top:1px solid #EFEFEF;">'
        f'<tr><td style="padding-top:10px;font-family:Arial,Helvetica,sans-serif;font-size:12px;color:#888;">'
        f'<span style="color:#0F1B2D;font-weight:bold;">RWA Soluções</span><br/>'
        f'<span style="font-size:11px;">Automação fiscal para escritórios contábeis</span>'
        f'</td></tr></table>'
        f'</td></tr>'
        f'</table>'
        f'</td></tr></table>'
        f'</body></html>'
    )

def _enviar_email_interrupcao(conta, senha, destino, assunto, corpo_html, corpo_plain):
    """Envia email multipart/alternative (HTML + plain) via Gmail SMTP."""
    msg = _MMulti_h("alternative")
    msg["From"]    = conta
    msg["To"]      = destino
    msg["Subject"] = assunto
    msg.attach(_MText_h(corpo_plain, "plain", "utf-8"))
    msg.attach(_MText_h(corpo_html,  "html",  "utf-8"))
    srv = _smtp_h.SMTP("smtp.gmail.com", 587)
    srv.starttls()
    srv.login(conta, senha)
    srv.sendmail(conta, destino, msg.as_string())
    srv.quit()

def _disparar_email_interrupcao():
    """Dispara email quando automação é interrompida."""
    try:
        _rwa_email_destino  = os.environ.get("RWA_EMAIL_DESTINO",   "").strip()
        _rwa_email_conta    = os.environ.get("RWA_EMAIL_CONTA",     "").strip()
        _rwa_email_senha    = os.environ.get("RWA_EMAIL_SENHA_APP", "").strip()
        _rwa_email_agendado = os.environ.get("RWA_EMAIL_AGENDADO",  "0") == "1"
        if not (_rwa_email_destino and _rwa_email_conta and _rwa_email_senha and _rwa_email_agendado):
            print(
                "[EMAIL] Email de nao iniciada NAO enviado: "
                f"destino={'OK' if _rwa_email_destino else 'VAZIO'}, "
                f"conta={'OK' if _rwa_email_conta else 'VAZIA'}, "
                f"senha_app={'OK' if _rwa_email_senha else 'VAZIA'}, "
                f"agendado={_rwa_email_agendado}"
            )
            return
        _comp_fmt = f"{COMPETENCIA[:2]}/{COMPETENCIA[2:]}"
        _data_int = datetime.now().strftime('%d/%m/%Y %H:%M:%S')
        _tempo_decorrido_s = time.time() - _tempo_inicio_processo
        _mins_int = int(_tempo_decorrido_s) // 60
        _segs_int = int(_tempo_decorrido_s) % 60
        _tempo_str = f"{_mins_int}min{_segs_int:02d}s"
        _total_processadas = len(linhas_relatorio) if 'linhas_relatorio' in globals() else 0
        _total_definidas   = total_empresas if 'total_empresas' in globals() else 0
        _motivo = "Interrupção do usuário"
        _sep = "  " + "-" * 32
        _corpo_int_plain = (
            f"Prezado(a),\n\n"
            f"A automação NFS-e Padrão Nacional foi interrompida antes da conclusão.\n\n"
            f"  Competência       : {_comp_fmt}\n"
            f"  Interrupção       : {_data_int}\n"
            f"  Tempo decorrido   : {_tempo_str}\n\n"
            f"{_sep}\n"
            f"  Total de empresas        : {_total_definidas}\n"
            f"  Empresas processadas     : {_total_processadas}\n"
            f"  Motivo                   : {_motivo}\n"
            f"{_sep}\n\n"
            f"Verifique o computador onde a automação está instalada.\n"
            f"Os relatórios parciais foram salvos na pasta de competência.\n\n"
            f"RWA Soluções"
        )
        _corpo_int_html = _render_html_email_interrupcao(
            tag_texto="INTERROMPIDA",
            tag_bg="#DC3545", tag_fg="#FFE0E0",
            titulo="Automação encerrada",
            intro="A automação foi interrompida antes da conclusão.",
            secoes=[
                ("Informações da execução", [
                    ("Competência",          _comp_fmt,        False),
                    ("Data/Hora",            _data_int,        False),
                    ("Empresas processadas", f"{_total_processadas} de {_total_definidas}", False),
                    ("Tempo decorrido",      _tempo_str,       False),
                    ("Motivo",               _motivo,          True),
                ]),
            ],
            callout=(
                "#FAEEDA", "#BA7517", "#633806",
                "Verifique o computador onde a automação está instalada. Os relatórios parciais foram salvos na pasta de competência.",
            ),
        )
        _enviar_email_interrupcao(
            _rwa_email_conta, _rwa_email_senha, _rwa_email_destino,
            f"RWA Soluções — NFS-e Padrão Nacional interrompida — {_comp_fmt}",
            _corpo_int_html, _corpo_int_plain,
        )
        print("[EMAIL] Email de interrupcao enviado.")
    except Exception as _e_int:
        print(f"[EMAIL] Falha ao enviar email de interrupcao: {_e_int}")

def _disparar_email_parou():
    """Dispara email quando automação para no meio (por qualquer motivo)."""
    try:
        _rwa_email_destino  = os.environ.get("RWA_EMAIL_DESTINO",   "").strip()
        _rwa_email_conta    = os.environ.get("RWA_EMAIL_CONTA",     "").strip()
        _rwa_email_senha    = os.environ.get("RWA_EMAIL_SENHA_APP", "").strip()
        if not (_rwa_email_destino and _rwa_email_conta and _rwa_email_senha):
            return
        _comp_fmt = f"{COMPETENCIA[:2]}/{COMPETENCIA[2:]}"
        _data_parada = datetime.now().strftime('%d/%m/%Y %H:%M:%S')
        _total_processadas = len(linhas_relatorio) if 'linhas_relatorio' in globals() else 0
        _total_definidas   = total_empresas if 'total_empresas' in globals() else 0
        _corpo_plain = (
            f"Prezado(a),\n\n"
            f"A automação foi interrompida antes da conclusão.\n\n"
            f"  Competência         : {_comp_fmt}\n"
            f"  Data/Hora parada    : {_data_parada}\n"
            f"  Empresas processadas: {_total_processadas} de {_total_definidas}\n\n"
            f"Verifique o computador onde a automação está instalada.\n"
            f"Os relatórios parciais foram salvos na pasta de competência.\n\n"
            f"RWA Soluções"
        )
        _corpo_html = _render_html_email_rwa(
            tag_texto="INTERROMPIDA",
            tag_bg="#DC3545", tag_fg="#FFE0E0",
            titulo="Automação interrompida",
            intro="A automação foi interrompida antes da conclusão.",
            secoes=[
                (None, [
                    ("Competência",          _comp_fmt,        False),
                    ("Data/Hora parada",     _data_parada,     False),
                    ("Empresas processadas", f"{_total_processadas} de {_total_definidas}", False),
                ]),
            ],
            callout=(
                "#FEE2E2", "#DC3545", "#991B1B",
                "Verifique o computador onde a automação está instalada. Os relatórios parciais foram salvos na pasta de competência.",
            ),
        )
        _enviar_email_rwa(
            _rwa_email_conta, _rwa_email_senha, _rwa_email_destino,
            f"RWA Soluções — NFS-e Padrão Nacional interrompida — {_comp_fmt}",
            _corpo_html, _corpo_plain,
        )
        print("[EMAIL] Email de parada enviado.")
    except Exception as _e_parou:
        print(f"[EMAIL] Falha ao enviar email de parada: {_e_parou}")

def _disparar_email_nao_iniciada(exception_obj):
    """Dispara email quando automação não consegue iniciar."""
    try:
        _rwa_email_destino  = os.environ.get("RWA_EMAIL_DESTINO",   "").strip()
        _rwa_email_conta    = os.environ.get("RWA_EMAIL_CONTA",     "").strip()
        _rwa_email_senha    = os.environ.get("RWA_EMAIL_SENHA_APP", "").strip()
        _rwa_email_agendado = os.environ.get("RWA_EMAIL_AGENDADO",  "0") == "1"
        if not (_rwa_email_destino and _rwa_email_conta and _rwa_email_senha and _rwa_email_agendado):
            return
        _tipo_erro = type(exception_obj).__name__
        _data_erro = datetime.now().strftime('%d/%m/%Y %H:%M:%S')
        try:
            _comp_fmt = f"{COMPETENCIA[:2]}/{COMPETENCIA[2:]}"
        except:
            _comp_fmt = "N/A"
        _corpo_plain = (
            f"Prezado(a),\n\n"
            f"A automação não pôde ser iniciada.\n\n"
            f"  Competência : {_comp_fmt}\n"
            f"  Data/Hora   : {_data_erro}\n"
            f"  Tipo de erro: {_tipo_erro}\n\n"
            f"Verifique o computador onde a automação está instalada.\n"
            f"Consulte os logs de diagnóstico para mais detalhes.\n\n"
            f"RWA Soluções"
        )
        _corpo_html = _render_html_email_rwa(
            tag_texto="NÃO INICIADA",
            tag_bg="#DC2626", tag_fg="#FEE2E2",
            titulo="Automação não iniciada",
            intro="Prezado(a), a automação não pôde ser iniciada.",
            secoes=[
                (None, [
                    ("Competência",  _comp_fmt,    False),
                    ("Data/Hora",    _data_erro,   False),
                    ("Tipo de erro", _tipo_erro,   True),
                ]),
            ],
            callout=(
                "#FEE2E2", "#DC2626", "#991B1B",
                "Verifique o computador onde a automação está instalada. Consulte os logs de diagnóstico para mais detalhes.",
            ),
        )
        _enviar_email_rwa(
            _rwa_email_conta, _rwa_email_senha, _rwa_email_destino,
            f"RWA Soluções — NFS-e Padrão Nacional não iniciada — {_comp_fmt}",
            _corpo_html, _corpo_plain,
        )
        os.environ["RWA_EMAIL_NAO_INICIADA_ENVIADO"] = "1"
        print("[EMAIL] Email de nao iniciada enviado.")
    except Exception as _e_ninc:
        print(f"[EMAIL] Falha ao enviar email de nao iniciada: {_e_ninc}")

def main():
    import sys as _sys
    _pw = RWAProgressWindow()
    _stdout_original = _sys.stdout
    _sys.stdout = _pw
    try:
        _main_corpo(_pw, _sys, _stdout_original)
    except BaseException as _e_main:
        # Se a falha ocorreu antes do loop real de empresas, é NÃO INICIADA.
        # Se já começou a processar empresa, o launcher tratará como INTERROMPIDA.
        try:
            if os.environ.get("RWA_PROCESSAMENTO_INICIADO", "0") != "1":
                _disparar_email_nao_iniciada(_e_main)
        except Exception as _e_email_main:
            print(f"[EMAIL] Falha ao tentar disparar email de nao iniciada dentro do main(): {_e_email_main}")
        raise
    finally:
        _sys.stdout = _stdout_original
        _pw.finalizar()


def _main_corpo(_pw, _sys, _stdout_original):
    competencia_fmt = f"{COMPETENCIA[:2]}/{COMPETENCIA[2:]}"

    print("=" * 70)
    print(f"  AUTOMAÇÃO NFS-e — PADRÃO NACIONAL")
    print(f"  Competência : {competencia_fmt}")
    print(f"  Iniciado em : {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}")
    print("=" * 70)

    lang = detectar_lang()
    erros_globais = []

    _rwa_paths_pn = {}
    try:
        import json as _json_pn
        _rwa_paths_file_pn = os.path.join(os.environ.get("LOCALAPPDATA", ""), "RWA_AUTOMACOES", "config", "paths.json")
        if os.path.exists(_rwa_paths_file_pn):
            _rwa_paths_pn = _json_pn.loads(open(_rwa_paths_file_pn, encoding="utf-8").read())
    except Exception:
        pass
    _caminho_senhas = _rwa_paths_pn.get("senhas_pn") or os.path.join(_SCRIPT_DIR, "SENHAS PADRAO NACIONAL.xlsx")
    if not os.path.exists(_caminho_senhas):
        _arquivos_xlsx = [f for f in os.listdir(_SCRIPT_DIR) if f.lower().endswith(".xlsx")]
        print()
        print("=" * 70)
        print("  ERRO -- PLANILHA NAO ENCONTRADA")
        print("=" * 70)
        print("  Arquivo esperado : SENHAS.xlsx")
        print("  Pasta verificada : " + _SCRIPT_DIR)
        if _arquivos_xlsx:
            print("  Arquivos .xlsx encontrados na pasta:")
            for _arq in _arquivos_xlsx:
                print("    -> " + _arq)
            print("  Renomeie o arquivo correto para SENHAS.xlsx e execute novamente.")
        else:
            print("  Nenhum arquivo .xlsx encontrado na pasta.")
            print("  Coloque o arquivo SENHAS.xlsx na mesma pasta do script.")
        print("=" * 70)
        raise Exception(f"PLANILHA SENHAS PADRAO NACIONAL.xlsx nao encontrada em: {_caminho_senhas}")
    wb_senhas = openpyxl.load_workbook(_caminho_senhas)
    ws_senhas = wb_senhas.active

    todas_empresas = []
    linhas_ignoradas = []
    for row in ws_senhas.iter_rows(min_row=2, values_only=True):
        codigo = row[0] if len(row) > 0 else ""
        nome   = row[1] if len(row) > 1 else ""
        tipo   = normalizar(row[3]) if len(row) > 3 else ""
        if not codigo and not nome:
            continue
        # Tem código mas tipo inválido ou ausente — registra no relatório
        if tipo not in ("LOGIN E SENHA", "CERTIFICADO"):
            obs_ignorada = (
                "TIPO DE ACESSO NÃO PREENCHIDO — Preencher coluna 'TIPO DE ACESSO' "
                "com 'LOGIN E SENHA' ou 'CERTIFICADO'"
                if not tipo else
                f"TIPO DE ACESSO INVÁLIDO: '{tipo}' — Usar 'LOGIN E SENHA' ou 'CERTIFICADO'"
            )
            linhas_ignoradas.append({
                "codigo":      str(codigo).strip(),
                "empresa":     str(nome).strip() if nome else "(sem nome)",
                "emit_qtd":    "Login incorreto",
                "emit_valor":  "Login incorreto",
                "emit_pdf":    "Login incorreto",
                "emit_pdf_st": "Login incorreto",
                "emit_xml":    "Login incorreto",
                "emit_xml_st": "Login incorreto",
                "rec_qtd":     "Login incorreto",
                "rec_valor":   "Login incorreto",
                "rec_pdf":     "Login incorreto",
                "rec_pdf_st":  "Login incorreto",
                "rec_xml":     "Login incorreto",
                "rec_xml_st":  "Login incorreto",
                "emit_cancel": "Login incorreto",
                "login_senha": "Login incorreto",
                "observacao":  obs_ignorada,
            })
            print(f"  [IGNORADA] {codigo} - {nome}: {obs_ignorada}")
            continue
        todas_empresas.append(row)

    total_empresas = len(todas_empresas)

    if total_empresas <= 0:
        raise Exception("Nenhuma empresa valida encontrada na planilha SENHAS PADRAO NACIONAL.xlsx.")

    # VALIDAÇÃO GRAVE PRÉ-BROWSER:
    # A automação NÃO pode iniciar se qualquer linha preenchida da SENHAS.xlsx
    # estiver sem PASTA ou com pasta inválida/inacessível.
    global PASTA_BASE
    pasta_global = _validar_todas_pastas_senhas(ws_senhas, _caminho_senhas)
    PASTA_BASE = pasta_global

    # MARCO CENTRAL PARA O LAUNCHER:
    # a partir daqui a automação passou pelas validações graves e vai iniciar o loop real.
    # Qualquer erro antes disso deve ser NÃO INICIADA; qualquer erro depois disso é INTERROMPIDA.
    os.environ["RWA_PROCESSAMENTO_INICIADO"] = "1"

    print(f"  Empresas    : {total_empresas}")
    print(f"  Pasta base  : {pasta_global}")
    print("  Pastas      : OK — todas as linhas preenchidas foram validadas antes do inicio")
    if linhas_ignoradas:
        print(f"  Ignoradas   : {len(linhas_ignoradas)} (tipo inválido/ausente — veja relatório)")
    print("=" * 70)

    # Já insere ignoradas no início do relatório
    linhas_relatorio = list(linhas_ignoradas)

    _tempo_inicio_processo = time.time()
    _log_diagnostico = []   # diagnóstico CMD por empresa
    _tempos_empresa_pn = {}  # tempo por empresa para CONCLUSAO
    _fase_pn           = "INICIO"  # fase atual para diagnóstico
    _email_50_enviado  = False     # garante disparo único do email de 50%

    # ── Helpers de email RWA (identidade visual unificada) ───────────────────

    # ── Email de início — enviado aqui, quando o processamento realmente começa ──
    try:
        _rwa_email_destino  = os.environ.get("RWA_EMAIL_DESTINO",   "").strip()
        _rwa_email_agendado = os.environ.get("RWA_EMAIL_AGENDADO",  "0") == "1"
        _rwa_email_conta    = os.environ.get("RWA_EMAIL_CONTA",     "").strip()
        _rwa_email_senha    = os.environ.get("RWA_EMAIL_SENHA_APP", "").strip()

        if _rwa_email_destino and _rwa_email_agendado and _rwa_email_conta and _rwa_email_senha:
            _comp_fmt = f"{COMPETENCIA[:2]}/{COMPETENCIA[2:]}"
            _ini_dt   = datetime.now().strftime('%d/%m/%Y %H:%M:%S')
            # ── plain text (fallback) ──
            _corpo_ini_plain = (
                f"Prezado(a),\n\n"
                f"A automação NFS-e Padrão Nacional foi iniciada conforme agendado.\n\n"
                f"Competência    : {_comp_fmt}\n"
                f"Início         : {_ini_dt}\n"
                f"Total empresas : {total_empresas}\n\n"
                f"Acompanhe o andamento no computador onde a automação está sendo executada.\n\n"
                f"RWA Soluções"
            )
            # ── HTML (visual) ──
            _corpo_ini_html = _render_html_email_rwa(
                tag_texto="INICIADA",
                tag_bg="#185FA5", tag_fg="#E6F1FB",
                titulo="Automação iniciada",
                intro="Prezado(a), a automação foi iniciada conforme agendado.",
                secoes=[
                    (None, [
                        ("Competência",       _comp_fmt,           False),
                        ("Início",            _ini_dt,             False),
                        ("Total de empresas", str(total_empresas), False),
                    ]),
                ],
                callout=(
                    "#F4F8FB", "#185FA5", "#0C447C",
                    "Acompanhe o andamento no computador onde a automação está sendo executada.",
                ),
            )
            _enviar_email_rwa(
                _rwa_email_conta, _rwa_email_senha, _rwa_email_destino,
                f"RWA Soluções — NFS-e Padrão Nacional iniciada — {_comp_fmt}",
                _corpo_ini_html, _corpo_ini_plain,
            )
            print("[EMAIL] Email de inicio enviado.")
    except Exception as _e_mail:
        print(f"[EMAIL] Falha ao enviar email de inicio: {_e_mail}")

    fila = [{"row": row, "tentativa": 1, "idx_original": idx}
            for idx, row in enumerate(todas_empresas, 1)]

    try:
        while fila:
            item = fila.pop(0)
            row       = item["row"]
            tentativa = item["tentativa"]
            idx       = item["idx_original"]

            codigo = row[0] if len(row) > 0 else ""
            nome   = row[1] if len(row) > 1 else ""
            tipo   = normalizar(row[3]) if len(row) > 3 else ""
            login  = str(row[4]).strip() if len(row) > 4 and row[4] is not None else ""
            senha  = str(row[5]).strip() if len(row) > 5 and row[5] is not None else ""
            tema   = str(row[6]).strip() if len(row) > 6 and row[6] is not None else ""
            _pasta_col = str(row[8]).strip() if len(row) > 8 and row[8] and str(row[8]).strip() not in ("", "None") else ""
            if not _pasta_col or not os.path.isdir(_pasta_col):
                raise Exception(f"ERRO GRAVE: pasta da linha ficou invalida apos validacao inicial: {_pasta_col}")
            pasta_raiz = _pasta_col

            pasta_download = pasta_empresa(codigo, nome, pasta_raiz)
            pasta_int = pasta_interna(pasta_download, codigo)
            _tomados_raw = str(row[7]).strip().upper() if len(row) > 7 and row[7] is not None else "SIM"
            tomados = "NAO" if _tomados_raw in ("NAO", "NÃO", "N", "NO") else "SIM"

            print(f"\n{'=' * 70}")
            if tentativa == 1:
                print(f"  EMPRESA {idx} DE {total_empresas}: {codigo} - {nome}")
            else:
                print(f"  REPROCESSAMENTO: {idx} DE {total_empresas}: {codigo} - {nome}")
            print(f"  Tipo  : {tipo}")
            print(f"  Pasta : {pasta_download}")
            print(f"{'=' * 70}")

            driver = None
            wait = None
            total_registros = None
            qtd_canceladas  = 0
            qtd_analise     = 0
            qtd_pdf_execucao = 0
            qtd_xml_execucao = 0
            xmls_execucao = []
            reprocessar     = False
            _tempo_inicio_empresa = time.time()
            _resultado_empresa = "OK"
            _fase_pn = "INICIO"

            try:
                # ── VALIDAÇÃO PRÉ-BROWSER ─────────────────────────────────────────────
                if not nome or str(nome).strip() == "":
                    raise LoginInvalidoException(
                        "NOME DA EMPRESA NÃO PREENCHIDO"
                    )
                if tipo == "LOGIN E SENHA":
                    digits_login = re.sub(r"\D", "", login)
                    if not login or login.upper() == "X":
                        raise LoginInvalidoException(
                            "LOGIN NÃO PREENCHIDO"
                        )
                    if len(digits_login) not in (11, 14):
                        _nd = len(digits_login)
                        if _nd == 13:
                            _hint = (
                                f"LOGIN INCOMPLETO — {_nd} dígito(s) informado(s), esperado 14 (CNPJ). "
                                f"O Excel removeu o zero à esquerda ao salvar o CNPJ como número. "
                                f"Na planilha SENHAS.xlsx, formate a coluna LOGIN como TEXTO e "
                                f"adicione o zero: '{('0' + str(login)).strip()}'"
                            )
                        elif _nd == 10:
                            _hint = (
                                f"LOGIN INCOMPLETO — {_nd} dígito(s) informado(s), esperado 11 (CPF). "
                                f"O Excel removeu o zero à esquerda ao salvar o CPF como número. "
                                f"Na planilha SENHAS.xlsx, formate a coluna LOGIN como TEXTO e "
                                f"adicione o zero: '{('0' + str(login)).strip()}'"
                            )
                        else:
                            _hint = (
                                f"LOGIN INCOMPLETO — {_nd} dígito(s) informado(s), "
                                f"esperado 11 (CPF) ou 14 (CNPJ). "
                                f"Verifique o campo LOGIN na planilha SENHAS.xlsx."
                            )
                        raise LoginInvalidoException(_hint)
                    if not senha or senha.upper() == "X":
                        raise LoginInvalidoException(
                            "SENHA NÃO PREENCHIDA"
                        )
                if tipo == "CERTIFICADO":
                    if not tema or tema.upper() == "X":
                        raise CertificadoNaoLocalizadoException(
                            "NOME AMIGÁVEL VAZIO — Preencher o campo Nome Amigável "
                            "no repositório de certificados (certmgr.msc) e atualizar planilha"
                        )
                    # Não validar letras isoladas antes do browser.
                    # O popup + OCR tolerante resolve nomes como A C A GOMES sem bloquear empresa legítima.

                driver = criar_driver(pasta_download)
                wait = WebDriverWait(driver, 25)

                driver.get(URL_LOGIN)
                _fase_pn = "LOGIN"
                esperar_site_estavel(driver, wait, timeout=40, descricao="tela inicial de login")

                if tipo == "LOGIN E SENHA":
                    login_com_senha(driver, wait, login, senha)
                elif tipo == "CERTIFICADO":
                    _fase_pn = "OCR_CERTIFICADO"
                    login_com_certificado(driver, wait, tema, lang)

                # ── PRESTADAS ─────────────────────────────────────────────────────
                _fase_pn = "EMITIDAS"
                total_registros, qtd_canceladas, qtd_analise, erros_empresa, qtd_pdf_execucao, qtd_xml_execucao, xmls_execucao = processar_empresa_pos_login(
                    driver, wait, codigo, nome, pasta_download, pasta_print=pasta_int
                )

                qtd_pdf = qtd_pdf_execucao
                qtd_xml = qtd_xml_execucao

                # Valores das emitidas — 3 totais diferentes:
                #   D = total bruto (com canceladas) — continua igual ao gerencial antigo
                #   E = soma so das canceladas efetivas
                #   F = soma das nao-canceladas (Regular + Substituida + Em analise)
                valor_total      = calcular_valor_total_xmls_execucao(xmls_execucao)
                valor_canceladas = calcular_valor_total_xmls_execucao(xmls_execucao, filtro_status="cancelada")
                valor_regular    = calcular_valor_total_xmls_execucao(xmls_execucao, filtro_status="regular")

                # Quantidade de regulares = total - canceladas efetivas
                # (Em analise entra como regular conforme regra definida)
                qtd_regular = (total_registros or 0) - (qtd_canceladas or 0)

                # Inscricao (CNPJ ou CPF) do emitente — pega do primeiro XML processado
                inscricao = ""
                for _d_emit in xmls_execucao:
                    _ic = _d_emit.get("emit_cnpj", "")
                    if _ic:
                        inscricao = _ic
                        break

                # Move PDFs prestados da pasta raiz para pasta_int (XMLs ficam na raiz para importação)
                import shutil as _shutil_mover
                for _arq_prest in os.listdir(pasta_download):
                    if _arq_prest.startswith("PADRAO_NACIONAL_NFSE_") and _arq_prest.lower().endswith(".pdf"):
                        _shutil_mover.move(
                            os.path.join(pasta_download, _arq_prest),
                            os.path.join(pasta_int, _arq_prest)
                        )

                # ZIP XML prestados — fica na raiz (para importação do sistema)
                zipar_arquivos_empresa(pasta_download, codigo, "prestados")
                # ZIP PDF prestados — vai para pasta_int
                zipar_arquivos_empresa(pasta_int, codigo, "prestados")

                # ZIP notas especiais (canceladas / em análise / substituídas) — destino pasta_int
                import zipfile as _zipfile
                for _status_label, _pasta_nome in [
                    ("cancelada",               "_especiais_cancelada"),
                    ("cancelamento_em_analise", "_especiais_cancelamento_em_analise"),
                    ("substituida",             "_especiais_substitu_da"),
                ]:
                    _pasta_esp = os.path.join(pasta_download, _pasta_nome)
                    if not os.path.isdir(_pasta_esp):
                        continue
                    _arquivos_esp = sorted([f for f in os.listdir(_pasta_esp) if f.lower().endswith((".xml", ".pdf"))])
                    if not _arquivos_esp:
                        continue
                    _nome_zip = f"{_status_label}_{codigo}_{COMPETENCIA}.zip"
                    _caminho_zip = os.path.join(pasta_int, _nome_zip)
                    with _zipfile.ZipFile(_caminho_zip, "w", _zipfile.ZIP_DEFLATED) as _zf:
                        for _arq in _arquivos_esp:
                            _zf.write(os.path.join(_pasta_esp, _arq), _arq)
                    import shutil
                    shutil.rmtree(_pasta_esp, ignore_errors=True)
                    print(f"  [ZIP] {_nome_zip} — {len(_arquivos_esp)} arquivo(s)")

                # Limpeza de pastas antigas com acento (resíduos de versões anteriores)
                import shutil as _shutil_clean
                for _pasta_antiga in [
                    "_especiais_cancelamento_em_análise",
                    "xml_especiais_cancelamento_em_análise",
                    "_especiais_cancelamento_em_an_lise",
                ]:
                    _p_antiga = os.path.join(pasta_download, _pasta_antiga)
                    if os.path.isdir(_p_antiga):
                        _shutil_clean.rmtree(_p_antiga, ignore_errors=True)
                        print(f"  [LIMPEZA] Pasta antiga removida: {_pasta_antiga}")

                # ── RECEBIDAS ─────────────────────────────────────────────────────
                _fase_pn = "RECEBIDAS"
                if tomados == "SIM":
                    total_rec, canceladas_rec, erros_rec, pdf_rec, xml_rec, xmls_rec = processar_recebidas(
                        driver, wait, codigo, nome, pasta_int
                    )
                    valor_rec = calcular_valor_total_xmls_execucao(xmls_rec)
                    # ZIP tomadas — de pasta_int
                    zipar_arquivos_empresa(pasta_int, codigo, "tomados")
                else:
                    print("  [INFO] Recebidas dispensadas (TOMADOS=NAO na planilha).")
                    total_rec = 0
                    canceladas_rec = 0
                    erros_rec = []
                    pdf_rec = 0
                    xml_rec = 0
                    xmls_rec = []
                    valor_rec = 0.0

                print("\nConferência da empresa:")
                print(f"Total portal (prestadas): {total_registros}")
                print(f"Cancelamento em analise : {qtd_analise}")
                print(f"Canceladas  (prestadas) : {qtd_canceladas}")
                print(f"PDFs prestadas baixados : {qtd_pdf}")
                print(f"XMLs prestadas baixados : {qtd_xml + qtd_canceladas + qtd_analise}")
                print(f"Valor total prestadas   : R$ {valor_total:.2f}")
                print(f"Total portal (recebidas): {total_rec}")
                print(f"PDFs recebidas baixados : {pdf_rec}")
                print(f"XMLs recebidas baixados : {xml_rec}")
                print(f"Valor total recebidas   : R$ {valor_rec:.2f}")

                validas = (total_registros or 0) - qtd_canceladas - qtd_analise

                if total_registros == 0:
                    pdf_status = "Sem movimento"
                    xml_status = "Sem movimento"
                else:
                    pdf_status = "✓" if (qtd_pdf + qtd_canceladas + qtd_analise) == total_registros else "X"
                    xml_status = "✓" if (qtd_xml + qtd_canceladas + qtd_analise) == total_registros else "X"

                rec_validas = (total_rec or 0) - (canceladas_rec or 0)
                if total_rec == 0 and not valor_rec:
                    rec_pdf_st = "Sem movimento"
                    rec_xml_st = "Sem movimento"
                elif total_rec == 0 and valor_rec:
                    # Portal não retornou contagem mas há valor — critério: baixou arquivo?
                    rec_pdf_st = "✓" if pdf_rec > 0 else "X"
                    rec_xml_st = "✓" if xml_rec > 0 else "X"
                else:
                    rec_pdf_st = "✓" if pdf_rec == rec_validas else "X"
                    rec_xml_st = "✓" if xml_rec == rec_validas else "X"

                # Determina situação geral e observação
                sem_emit = total_registros == 0
                sem_rec  = (total_rec or 0) == 0 and not valor_rec
                if sem_emit and sem_rec:
                    obs = "SEM MOVIMENTO NO PERÍODO"
                    resultado_diag = "SEM MOVIMENTO"
                    obs_diag = ""
                elif sem_emit and not sem_rec:
                    obs = "SOMENTE RECEBIDAS NO MÊS"
                    resultado_diag = "SOMENTE RECEBIDAS NO MÊS"
                    obs_diag = f"Recebidas: {pdf_rec} PDF / {xml_rec} XML / R$ {valor_rec:.2f}"
                elif not sem_emit and sem_rec:
                    obs = "SOMENTE EMITIDAS NO MÊS"
                    if qtd_canceladas > 0:
                        obs += f" | {qtd_canceladas} nota(s) cancelada(s) ignorada(s)"
                    resultado_diag = "OK"
                    obs_diag = f"Emitidas: {qtd_pdf} PDF / {qtd_xml} XML / R$ {valor_total:.2f}"
                else:
                    obs = f"{qtd_canceladas} nota(s) cancelada(s) ignorada(s)" if qtd_canceladas > 0 else ""
                    resultado_diag = "OK"
                    obs_diag = (
                        f"Emitidas: {qtd_pdf} PDF / {qtd_xml} XML / R$ {valor_total:.2f} | "
                        f"Recebidas: {pdf_rec} PDF / {xml_rec} XML / R$ {valor_rec:.2f}"
                    )

                # ── CONFRONTO QUANTIDADE — reprocessamento se não bateu ──────────
                # Emitidas: qtd_xml baixados vs total_registros (descontando canceladas)
                # Recebidas: xml_rec baixados vs total_rec (descontando canceladas_rec)
                validas = (total_registros or 0) - qtd_canceladas - qtd_analise
                emit_incompleta = (total_registros or 0) > 0 and qtd_xml < validas
                rec_incompleta  = (total_rec or 0) > 0 and xml_rec < rec_validas

                if (emit_incompleta or rec_incompleta) and tentativa == 1:
                    _faltam_emit = validas - qtd_xml if emit_incompleta else 0
                    _faltam_rec  = rec_validas - xml_rec if rec_incompleta else 0
                    _msg_inc = []
                    if emit_incompleta:
                        _msg_inc.append(f"Emitidas: {qtd_xml} baixadas / {validas} esperadas — faltam {_faltam_emit}")
                    if rec_incompleta:
                        _msg_inc.append(f"Recebidas: {xml_rec} baixadas / {rec_validas} esperadas — faltam {_faltam_rec}")
                    _msg_inc_str = " | ".join(_msg_inc)
                    print(f"  [CONFRONTO] Quantidade não bateu — enviando para o final da fila.")
                    print(f"  [CONFRONTO] {_msg_inc_str}")
                    erros_globais.append({"codigo": codigo, "empresa": nome, "tipo": tipo, "erro": f"DOWNLOAD INCOMPLETO (tentativa 1): {_msg_inc_str}"})
                    _log_diagnostico.append({
                        "ordem": idx, "empresa": f"{codigo} - {nome}",
                        "resultado": "DOWNLOAD INCOMPLETO", "observacao": _msg_inc_str,
                    })
                    registrar_diagnostico_pn(codigo, nome, tipo, "CONFRONTO_QUANTIDADE",
                                              Exception(_msg_inc_str),
                                              driver=driver, pasta=pasta_download,
                                              tentativa=tentativa, tema_cert=tema,
                                              tempo_decorrido=round(time.time()-_tempo_inicio_empresa,1),
                                              status_atual={
                                                  "PDF Emitidas":  pdf_status,
                                                  "XML Emitidas":  xml_status,
                                                  "PDF Recebidas": rec_pdf_st,
                                                  "XML Recebidas": rec_xml_st,
                                              })
                    fila.append({"row": row, "tentativa": 2, "idx_original": idx})
                    reprocessar = True

                else:
                    linhas_relatorio.append({
                        "codigo":             codigo,
                        "inscricao":          inscricao,
                        "empresa":            nome,
                        "emit_qtd":           "Sem movimento" if sem_emit else (total_registros if total_registros is not None else ""),
                        "emit_valor":         "Sem movimento" if sem_emit else (f"R$ {valor_total:.2f}" if valor_total else ""),
                        "emit_valor_canc":    "Sem movimento" if sem_emit else (f"R$ {valor_canceladas:.2f}" if valor_canceladas else ""),
                        "emit_regular_valor": "Sem movimento" if sem_emit else (f"R$ {valor_regular:.2f}" if valor_regular else ""),
                        "emit_regular_qtd":   "Sem movimento" if sem_emit else (qtd_regular if qtd_regular else ""),
                        "emit_pdf":           "Sem movimento" if sem_emit else (qtd_pdf + qtd_canceladas + qtd_analise),
                        "emit_pdf_st":        "Sem movimento" if sem_emit else pdf_status,
                        "emit_xml":           "Sem movimento" if sem_emit else (qtd_xml + qtd_canceladas + qtd_analise),
                        "emit_xml_st":        "Sem movimento" if sem_emit else xml_status,
                        "rec_qtd":            "Sem movimento" if sem_rec else (total_rec if total_rec is not None else ""),
                        "rec_valor":          "Sem movimento" if sem_rec else (f"R$ {valor_rec:.2f}" if valor_rec else ""),
                        "rec_pdf":            "Sem movimento" if sem_rec else pdf_rec,
                        "rec_pdf_st":         "Sem movimento" if sem_rec else rec_pdf_st,
                        "rec_xml":            "Sem movimento" if sem_rec else xml_rec,
                        "rec_xml_st":         "Sem movimento" if sem_rec else rec_xml_st,
                        "emit_cancel":        qtd_canceladas if qtd_canceladas else "",
                        "emit_analise":       qtd_analise if qtd_analise else "",
                        "login_senha":        "✓",
                        "observacao":         obs,
                    })

                print("\nEmpresa finalizada")

                # ── DETALHAMENTO POR NOTA ──────────────────────────────────────────
                # Salva o Detalhamento direto na raiz da competência (ex: 2-TEO TECH\042026\)
                # em vez da subpasta interna {codigo}_{COMPETENCIA}. Facilita acesso do
                # contador e da conferência Padrão Nacional.
                gerar_detalhamento_nfse(
                    pasta_download = pasta_download,
                    nome_empresa   = nome,
                    competencia    = COMPETENCIA,
                    xmls_emitidas  = xmls_execucao,
                    xmls_recebidas = xmls_rec,
                )

                # ── GERENCIAL PARCIAL — salva após cada empresa ────────────────────
                try:
                    caminho_parcial = os.path.join(PASTA_BASE, f"PARCIAL_{RELATORIO_NOME}")
                    criar_relatorio_gerencial(caminho_parcial, linhas_relatorio)
                    print(f"  [GERENCIAL] Parcial atualizado ({len(linhas_relatorio)} empresa(s))")
                except Exception as e_parc:
                    print(f"  [AVISO] Gerencial parcial não salvo: {e_parc}")

                _log_diagnostico.append({
                    "ordem": idx, "empresa": f"{codigo} - {nome}",
                    "resultado": resultado_diag,
                    "observacao": obs_diag,
                })

                # Registrar diagnóstico se houve X em PDF ou XML (leitura de quantidade falhou)
                if pdf_status == "X" or xml_status == "X" or rec_pdf_st == "X" or rec_xml_st == "X":
                    _msg_x = []
                    if pdf_status == "X":
                        _msg_x.append(f"Emitidas PDF: {qtd_pdf} baixados / {validas} esperados")
                    if xml_status == "X":
                        _msg_x.append(f"Emitidas XML: {qtd_xml} baixados / {validas} esperados")
                    if rec_pdf_st == "X":
                        _msg_x.append(f"Recebidas PDF: {pdf_rec} baixados / portal retornou {total_rec or 0} registros")
                    if rec_xml_st == "X":
                        _msg_x.append(f"Recebidas XML: {xml_rec} baixados / portal retornou {total_rec or 0} registros")
                    _fase_x = "EMITIDAS" if (pdf_status == "X" or xml_status == "X") else "RECEBIDAS"
                    registrar_diagnostico_pn(
                        codigo, nome, tipo,
                        fase=_fase_x,
                        erro=Exception(" | ".join(_msg_x)),
                        driver=driver, pasta=pasta_download,
                        tentativa=tentativa, tema_cert=tema,
                        tempo_decorrido=round(time.time() - _tempo_inicio_empresa, 1),
                        status_atual={
                            "PDF Emitidas":  pdf_status,
                            "XML Emitidas":  xml_status,
                            "PDF Recebidas": rec_pdf_st,
                            "XML Recebidas": rec_xml_st,
                        }
                    )

                time.sleep(3)

            except NomeAmigavelInvalidoException as e:
                msg_inv = str(e)
                _resultado_empresa = "NOME AMIGÁVEL INVÁLIDO"
                print(f"  [NOME AMIGÁVEL INVÁLIDO] {codigo} - {nome}: {msg_inv}")
                erros_globais.append({"codigo": codigo, "empresa": nome, "tipo": tipo, "erro": msg_inv})
                _log_diagnostico.append({
                    "ordem": idx, "empresa": f"{codigo} - {nome}",
                    "resultado": "NOME AMIGÁVEL INVÁLIDO", "observacao": msg_inv,
                })
                registrar_diagnostico_pn(codigo, nome, tipo, "OCR_CERTIFICADO", e,
                                          driver=driver, pasta=pasta_download,
                                          tentativa=tentativa, tema_cert=tema,
                                          tempo_decorrido=round(time.time()-_tempo_inicio_empresa,1))
                linhas_relatorio.append({
                    "codigo":             codigo,
                    "inscricao":          "",
                    "empresa":            nome,
                    "rec_qtd":      "", "rec_valor":   "", "rec_pdf":    "",
                    "rec_pdf_st":   "X", "rec_xml":    "", "rec_xml_st": "X",
                    "emit_qtd":     "", "emit_valor":  "", "emit_pdf":   "",
                    "emit_pdf_st":  "X", "emit_xml":   "", "emit_xml_st":"X",
                    "emit_valor_canc":    "",
                    "emit_regular_valor": "",
                    "emit_regular_qtd":   "",
                    "emit_cancel":  "",
                    "login_senha":  "CERTIFICADO",
                    "observacao":   msg_inv,
                })

            except LoginInvalidoException as e:
                msg_login = str(e)
                _resultado_empresa = "Login incorreto"
                # Determina observação com solução embutida
                msg_up = msg_login.upper()
                if "INCOMPLETO" in msg_up:
                    obs_login = msg_login  # já vem formatada da validação pré-browser
                elif "NÃO PREENCHIDO" in msg_up or "NAO PREENCHIDO" in msg_up:
                    obs_login = msg_login
                else:
                    obs_login = (
                        "SENHA INCORRETA — Atualizar senha na planilha"
                        if "senha" in msg_login.lower()
                        else "LOGIN/SENHA INCORRETO — Atualizar credenciais na planilha"
                    )
                print(f"  [LOGIN INVÁLIDO] {codigo} - {nome}: {obs_login}")
                erros_globais.append({"codigo": codigo, "empresa": nome, "tipo": tipo, "erro": obs_login})
                _log_diagnostico.append({
                    "ordem": idx, "empresa": f"{codigo} - {nome}",
                    "resultado": "Login incorreto", "observacao": obs_login,
                })
                registrar_diagnostico_pn(codigo, nome, tipo, "LOGIN", e,
                                          driver=driver, pasta=pasta_download,
                                          tentativa=tentativa,
                                          tempo_decorrido=round(time.time()-_tempo_inicio_empresa,1))
                linhas_relatorio.append({
                    "codigo":             codigo,
                    "inscricao":          "",
                    "empresa":            nome,
                    "emit_qtd":           "Login incorreto",
                    "emit_valor":         "Login incorreto",
                    "emit_valor_canc":    "Login incorreto",
                    "emit_regular_valor": "Login incorreto",
                    "emit_regular_qtd":   "Login incorreto",
                    "emit_pdf":           "Login incorreto",
                    "emit_pdf_st":        "Login incorreto",
                    "emit_xml":           "Login incorreto",
                    "emit_xml_st":        "Login incorreto",
                    "rec_qtd":            "Login incorreto",
                    "rec_valor":          "Login incorreto",
                    "rec_pdf":            "Login incorreto",
                    "rec_pdf_st":         "Login incorreto",
                    "rec_xml":            "Login incorreto",
                    "rec_xml_st":         "Login incorreto",
                    "emit_cancel":        "Login incorreto",
                    "login_senha":        "Login incorreto",
                    "observacao":         obs_login,
                })

            except CertificadoInterferenciaException as e:
                _fechar_popup_certificado()
                msg_ocr = str(e)
                if tentativa == 1:
                    _resultado_empresa = "INTERFERÊNCIA OCR (reprocessar)"
                    obs_ocr = (
                        "INTERFERÊNCIA NA LEITURA DO CERTIFICADO — mouse, foco da janela, AnyDesk ou janela sobreposta. "
                        "Empresa enviada para o final da fila para nova tentativa."
                    )
                    print(f"  [REPROCESSAR] {codigo} - {nome}: {obs_ocr}")
                    erros_globais.append({"codigo": codigo, "empresa": nome, "tipo": tipo, "erro": f"REPROCESSAR OCR (tentativa 1): {msg_ocr}"})
                    _log_diagnostico.append({
                        "ordem": idx, "empresa": f"{codigo} - {nome}",
                        "resultado": "INTERFERÊNCIA OCR", "observacao": msg_ocr[:180],
                    })
                    # Não passa driver aqui: popup nativo/janela sobreposta pode travar screenshot/DOM.
                    registrar_diagnostico_pn(codigo, nome, tipo, "OCR_CERTIFICADO_INTERFERENCIA", e,
                                              driver=None, pasta=pasta_download,
                                              tentativa=tentativa, tema_cert=tema,
                                              tempo_decorrido=round(time.time()-_tempo_inicio_empresa,1))
                    fila.append({"row": row, "tentativa": 2, "idx_original": idx})
                    reprocessar = True
                else:
                    _resultado_empresa = "INTERFERÊNCIA OCR"
                    obs_ocr = (
                        "INTERFERÊNCIA NA LEITURA DO CERTIFICADO TAMBÉM NA 2ª TENTATIVA — "
                        "executar manualmente ou rodar novamente sem mexer no mouse/janela durante o popup."
                    )
                    print(f"  [INTERFERÊNCIA OCR] {codigo} - {nome}: {obs_ocr}")
                    erros_globais.append({"codigo": codigo, "empresa": nome, "tipo": tipo, "erro": obs_ocr})
                    _log_diagnostico.append({
                        "ordem": idx, "empresa": f"{codigo} - {nome}",
                        "resultado": "INTERFERÊNCIA OCR", "observacao": obs_ocr,
                    })
                    # Não passa driver aqui: popup nativo/janela sobreposta pode travar screenshot/DOM.
                    registrar_diagnostico_pn(codigo, nome, tipo, "OCR_CERTIFICADO_INTERFERENCIA", e,
                                              driver=None, pasta=pasta_download,
                                              tentativa=tentativa, tema_cert=tema,
                                              tempo_decorrido=round(time.time()-_tempo_inicio_empresa,1))
                    linhas_relatorio.append({
                        "codigo":       codigo,
                        "empresa":      nome,
                        "emit_qtd":     "OCR",
                        "emit_valor":   "OCR",
                        "emit_pdf":     "OCR",
                        "emit_pdf_st":  "OCR",
                        "emit_xml":     "OCR",
                        "emit_xml_st":  "OCR",
                        "rec_qtd":      "OCR",
                        "rec_valor":    "OCR",
                        "rec_pdf":      "OCR",
                        "rec_pdf_st":   "OCR",
                        "rec_xml":      "OCR",
                        "rec_xml_st":   "OCR",
                        "emit_cancel":  "OCR",
                        "login_senha":  "CERTIFICADO",
                        "observacao":   obs_ocr,
                    })
            except CertificadoNaoLocalizadoException as e:
                msg_cert = str(e)
                _resultado_empresa = "CERTIFICADO NÃO LOCALIZADO"
                msg_up = msg_cert.upper()
                if "VAZIO" in msg_up or "NAO PREENCHIDO" in msg_up or "NÃO PREENCHIDO" in msg_up:
                    obs_cert = msg_cert  # vem formatada da validação pré-browser
                elif "VENCIDO" in msg_up:
                    obs_cert = "CERTIFICADO VENCIDO — Conferir vencimento ou reinstalar no repositório"
                else:
                    obs_cert = (
                        "CERTIFICADO NÃO ENCONTRADO — Conferir vencimento ou se está "
                        "instalado no repositório (certmgr.msc). "
                        f"Tema buscado: '{tema}' — conferir se o Nome Amigável está idêntico"
                    )
                print(f"  [CERTIFICADO NÃO LOCALIZADO] {codigo} - {nome}: {obs_cert}")
                erros_globais.append({"codigo": codigo, "empresa": nome, "tipo": tipo, "erro": obs_cert})
                _log_diagnostico.append({
                    "ordem": idx, "empresa": f"{codigo} - {nome}",
                    "resultado": "CERTIFICADO NÃO LOCALIZADO", "observacao": obs_cert,
                })
                registrar_diagnostico_pn(codigo, nome, tipo, "CERTIFICADO", e,
                                          driver=driver, pasta=pasta_download,
                                          tentativa=tentativa, tema_cert=tema,
                                          tempo_decorrido=round(time.time()-_tempo_inicio_empresa,1))
                linhas_relatorio.append({
                    "codigo":             codigo,
                    "inscricao":          "",
                    "empresa":            nome,
                    "emit_qtd":           "Certificado",
                    "emit_valor":         "Certificado",
                    "emit_valor_canc":    "Certificado",
                    "emit_regular_valor": "Certificado",
                    "emit_regular_qtd":   "Certificado",
                    "emit_pdf":           "Certificado",
                    "emit_pdf_st":        "Certificado",
                    "emit_xml":           "Certificado",
                    "emit_xml_st":        "Certificado",
                    "rec_qtd":            "Certificado",
                    "rec_valor":          "Certificado",
                    "rec_pdf":            "Certificado",
                    "rec_pdf_st":         "Certificado",
                    "rec_xml":            "Certificado",
                    "rec_xml_st":         "Certificado",
                    "emit_cancel":        "Certificado",
                    "login_senha":        "CERTIFICADO",
                    "observacao":         obs_cert,
                })

            except Exception as e:
                msg = str(e)

                eh_cert_nao_localizado = (
                    "CERTIFICADO NÃO LOCALIZADO" in msg.upper() or
                    "CERTIFICADO NAO LOCALIZADO" in msg.upper()
                )

                # Tentativa 1: qualquer erro que não seja certificado vai para reprocessamento
                if tentativa == 1 and not eh_cert_nao_localizado:
                    _resultado_empresa = "INSTABILIDADE (reprocessar)"
                    print(f"  [REPROCESSAR] {nome} — erro na tentativa 1: {msg[:120]}")
                    erros_globais.append({"codigo": codigo, "empresa": nome, "tipo": tipo, "erro": f"REPROCESSAR (tentativa 1): {msg}"})
                    _log_diagnostico.append({
                        "ordem": idx, "empresa": f"{codigo} - {nome}",
                        "resultado": "INSTABILIDADE", "observacao": msg[:120],
                    })

                    # Se o ChromeDriver perdeu comunicação depois do certificado, não usar driver
                    # em diagnóstico/screenshot. Isso era o que travava a fila antes do finally.
                    if _driver_comunicacao_quebrada(msg):
                        print("  [AVISO] Driver instável após certificado — encerrando sem usar DOM/screenshot.")
                        try:
                            _encerrar_driver_sem_travar(driver, timeout=1)
                        except Exception:
                            pass
                        _driver_diag = None
                        driver = None
                    else:
                        _driver_diag = driver

                    registrar_diagnostico_pn(codigo, nome, tipo, _fase_pn, e,
                                              driver=_driver_diag, pasta=pasta_download,
                                              tentativa=tentativa, tema_cert=tema,
                                              tempo_decorrido=round(time.time()-_tempo_inicio_empresa,1))
                    fila.append({"row": row, "tentativa": 2, "idx_original": idx})
                    reprocessar = True
                else:
                    _resultado_empresa = "ERRO OPERACIONAL"
                    if "CERTIFICADO NÃO LOCALIZADO" in msg.upper() or "CERTIFICADO NAO LOCALIZADO" in msg.upper():
                        print(f"  [CERTIFICADO NÃO LOCALIZADO] {codigo} - {nome}")
                    else:
                        print(f"  [ERRO OPERACIONAL] {codigo} - {nome}: {msg[:120]}")
                    erros_globais.append({"codigo": codigo, "empresa": nome, "tipo": tipo, "erro": msg})
                    _log_diagnostico.append({
                        "ordem": idx, "empresa": f"{codigo} - {nome}",
                        "resultado": "ERRO OPERACIONAL", "observacao": msg[:120],
                    })
                    if _driver_comunicacao_quebrada(msg):
                        print("  [AVISO] Driver instável — diagnóstico sem DOM/screenshot para não travar.")
                        try:
                            _encerrar_driver_sem_travar(driver, timeout=1)
                        except Exception:
                            pass
                        _driver_diag = None
                        driver = None
                    else:
                        _driver_diag = driver

                    registrar_diagnostico_pn(codigo, nome, tipo, _fase_pn, e,
                                              driver=_driver_diag, pasta=pasta_download,
                                              tentativa=tentativa, tema_cert=tema,
                                              tempo_decorrido=round(time.time()-_tempo_inicio_empresa,1))
                    qtd_pdf_err = 0
                    qtd_xml_err = 0
                    try:
                        qtd_pdf_err = contar_arquivos(pasta_download, ".pdf")
                        qtd_xml_err = contar_arquivos(pasta_download, ".xml")
                    except Exception:
                        pass
                    linhas_relatorio.append({
                        "codigo":             codigo,
                        "inscricao":          "",
                        "empresa":            nome,
                        "rec_qtd":      "", "rec_valor":   "", "rec_pdf":    "",
                        "rec_pdf_st":   "X", "rec_xml":    "", "rec_xml_st": "X",
                        "emit_qtd":     "", "emit_valor":  "", "emit_pdf":   qtd_pdf_err,
                        "emit_pdf_st":  "X", "emit_xml":   qtd_xml_err, "emit_xml_st": "X",
                        "emit_valor_canc":    "",
                        "emit_regular_valor": "",
                        "emit_regular_qtd":   "",
                        "emit_cancel":  "",
                        "login_senha":  "✓",
                        "observacao":   f"ERRO: {msg[:180]}",
                    })

            finally:
                tempo_empresa = time.time() - _tempo_inicio_empresa
                mins = int(tempo_empresa // 60)
                segs = int(tempo_empresa % 60)
                _tempos_empresa_pn[str(codigo)] = tempo_empresa
                print(f"  [TEMPO] EMPRESA {idx:>03}/{total_empresas} — {mins}min{segs:02d}s")
                if driver:
                    _encerrar_driver_sem_travar(driver, timeout=3)
                print(f"  [FIM] {nome}\n")

            if reprocessar:
                continue

            # ── Email de 50% — disparado uma única vez quando atinge a metade ──
            _empresas_concluidas_pn = len(linhas_relatorio) - len(linhas_ignoradas)
            if (not _email_50_enviado
                    and total_empresas >= 2
                    and _empresas_concluidas_pn * 2 >= total_empresas):
                _email_50_enviado = True
                try:
                    _rwa_email_destino  = os.environ.get("RWA_EMAIL_DESTINO",   "").strip()
                    _rwa_email_conta    = os.environ.get("RWA_EMAIL_CONTA",     "").strip()
                    _rwa_email_senha    = os.environ.get("RWA_EMAIL_SENHA_APP", "").strip()
                    if _rwa_email_destino and _rwa_email_conta and _rwa_email_senha:
                        _comp_fmt3 = f"{COMPETENCIA[:2]}/{COMPETENCIA[2:]}"
                        _t50 = int(time.time() - _tempo_inicio_processo)
                        _m50 = _t50 // 60
                        _s50 = _t50 % 60
                        _tempo50_str = f"{_m50}min{_s50:02d}s"
                        _progresso50 = f"{_empresas_concluidas_pn} de {total_empresas} empresas"
                        # ── plain text (fallback) ──
                        _corpo50_plain = (
                            f"RWA Soluções — NFS-e Padrão Nacional\n"
                            f"{'=' * 38}\n\n"
                            f"Status      : EM ANDAMENTO — 50% concluída\n"
                            f"Competência : {_comp_fmt3}\n"
                            f"Progresso   : {_empresas_concluidas_pn} de {total_empresas} empresas processadas\n"
                            f"Tempo decorrido: {_tempo50_str}\n\n"
                            f"Automação rodando com estabilidade.\n\n"
                            f"Metade da fila concluída. Você receberá um email quando a automação for concluída.\n\n"
                            f"RWA Soluções"
                        )
                        # ── HTML (visual) ──
                        _corpo50_html = _render_html_email_rwa(
                            tag_texto="EM ANDAMENTO",
                            tag_bg="#BA7517", tag_fg="#FAEEDA",
                            titulo="50% concluída",
                            intro="Automação rodando com estabilidade.",
                            intro_destaque=True,
                            secoes=[
                                (None, [
                                    ("Competência",     _comp_fmt3,    False),
                                    ("Progresso",       _progresso50,  False),
                                    ("Tempo decorrido", _tempo50_str,  False),
                                ]),
                            ],
                            callout=(
                                "#FAEEDA", "#BA7517", "#633806",
                                "Metade da fila concluída. Você receberá um email quando a automação for concluída.",
                            ),
                        )
                        _enviar_email_rwa(
                            _rwa_email_conta, _rwa_email_senha, _rwa_email_destino,
                            f"RWA Soluções — NFS-e Padrão Nacional 50% concluída — {_comp_fmt3}",
                            _corpo50_html, _corpo50_plain,
                        )
                        print("[EMAIL] Email de 50% enviado.")
                except Exception as _e50:
                    print(f"[EMAIL] Falha ao enviar email de 50%: {_e50}")


    except Exception as e_global:
        print(f"\n[ERRO CRÍTICO] Script interrompido inesperadamente: {e_global}")
        print("[RELATÓRIO PARCIAL] Salvando o que foi processado até agora...")
        try:
            caminho_parcial = os.path.join(PASTA_BASE, f"PARCIAL_{RELATORIO_NOME}")
            criar_relatorio_gerencial(caminho_parcial, linhas_relatorio)
            salvar_relatorio_erros(erros_globais)
            print(f"[RELATÓRIO PARCIAL] Salvo em: {caminho_parcial}")
        except Exception as e_salvar:
            print(f"[RELATÓRIO PARCIAL] Falha ao salvar: {e_salvar}")
        _disparar_email_parou()
        raise

    caminho_relatorio = os.path.join(PASTA_BASE, RELATORIO_NOME)
    criar_relatorio_gerencial(caminho_relatorio, linhas_relatorio)
    salvar_relatorio_erros(erros_globais)

    # deleta o parcial — processo concluído, só o final fica
    caminho_parcial = os.path.join(PASTA_BASE, f"PARCIAL_{RELATORIO_NOME}")
    try:
        if os.path.exists(caminho_parcial):
            os.remove(caminho_parcial)
    except Exception:
        pass

    # ── RELATÓRIO DE CONCLUSÃO E DIAGNÓSTICO ──────────────────────────────────
    tempo_total_s = time.time() - _tempo_inicio_processo
    gerar_relatorio_conclusao_pn(
        linhas_relatorio,
        tempo_total_s=tempo_total_s,
        total_original=total_empresas + len(linhas_ignoradas),
        tempos_empresa=_tempos_empresa_pn,
    )
    finalizar_diagnostico_pn()

    # ── Email de conclusão — disparado aqui, sem depender de clique do usuário ──
    try:
        _rwa_email_destino  = os.environ.get("RWA_EMAIL_DESTINO",   "").strip()
        _rwa_email_conta    = os.environ.get("RWA_EMAIL_CONTA",     "").strip()
        _rwa_email_senha    = os.environ.get("RWA_EMAIL_SENHA_APP", "").strip()

        if _rwa_email_destino and _rwa_email_conta and _rwa_email_senha:
            _comp_fmt   = f"{COMPETENCIA[:2]}/{COMPETENCIA[2:]}"
            _mins_total = int(tempo_total_s) // 60
            _segs_total = int(tempo_total_s) % 60
            _tot_pn     = len(linhas_relatorio)
            _med_s_pn   = int(tempo_total_s / _tot_pn) if _tot_pn > 0 else 0
            _med_m_pn   = _med_s_pn // 60
            _med_s2_pn  = _med_s_pn % 60

            _mov_pn  = sum(1 for r in linhas_relatorio
                           if str(r.get("login_senha","")) == "✓"
                           and str(r.get("emit_qtd","")) not in ("Sem movimento",""))
            _smov_pn = sum(1 for r in linhas_relatorio
                           if str(r.get("emit_qtd","")) == "Sem movimento"
                           and str(r.get("rec_qtd","")) == "Sem movimento")
            _cert_pn = sum(1 for r in linhas_relatorio
                           if str(r.get("login_senha","")) == "CERTIFICADO")
            _log_pn  = sum(1 for r in linhas_relatorio
                           if str(r.get("login_senha","")).lower() == "login incorreto")
            _div_pn  = sum(1 for r in linhas_relatorio
                           if str(r.get("emit_pdf_st","")) == "X"
                           or str(r.get("emit_xml_st","")) == "X")

            _tempo_total_str = f"{_mins_total}min{_segs_total:02d}s"
            _media_str       = f"{_med_m_pn}min{_med_s2_pn:02d}s"

            # ── plain text (fallback) ──
            _sep_pn = "  " + "-" * 32
            _corpo_fim_plain = (
                f"Prezado(a),\n\n"
                f"A automação NFS-e Padrão Nacional foi concluída.\n\n"
                f"  Competência   : {_comp_fmt}\n"
                f"  Tempo total   : {_tempo_total_str}\n"
                f"  Média/empresa : {_media_str}\n\n"
                f"{_sep_pn}\n"
                f"  Total processadas          : {_tot_pn}\n"
                f"  Com movimento              : {_mov_pn}\n"
                f"  Sem movimento              : {_smov_pn}\n"
                f"  Certificado não localizado : {_cert_pn}\n"
                f"  Login ou senha incorreto   : {_log_pn}\n"
                f"  Divergentes                : {_div_pn}\n"
                f"{_sep_pn}\n\n"
                f"Verifique os relatórios na pasta da competência.\n\n"
                f"RWA Soluções"
            )

            # ── HTML (visual) ──
            _corpo_fim_html = _render_html_email_rwa(
                tag_texto="CONCLUÍDA",
                tag_bg="#3B6D11", tag_fg="#EAF3DE",
                titulo="Automação concluída",
                intro="Prezado(a), a automação foi concluída.",
                secoes=[
                    ("Tempos", [
                        ("Competência",     _comp_fmt,        False),
                        ("Tempo total",     _tempo_total_str, False),
                        ("Média / empresa", _media_str,       False),
                    ]),
                    ("Resultado", [
                        ("Total processadas",          str(_tot_pn),  False),
                        ("Com movimento",              str(_mov_pn),  False),
                        ("Sem movimento",              str(_smov_pn), False),
                        ("Certificado não localizado", str(_cert_pn), _cert_pn > 0),
                        ("Login ou senha incorreto",   str(_log_pn),  _log_pn  > 0),
                        ("Divergentes",                str(_div_pn),  _div_pn  > 0),
                    ]),
                ],
                callout=(
                    "#EAF3DE", "#639922", "#173404",
                    "Verifique os relatórios na pasta da competência.",
                ),
            )

            _enviar_email_rwa(
                _rwa_email_conta, _rwa_email_senha, _rwa_email_destino,
                f"RWA Soluções — NFS-e Padrão Nacional concluída — {_comp_fmt}",
                _corpo_fim_html, _corpo_fim_plain,
            )
            print("[EMAIL] Email de conclusao enviado.")
    except Exception as _e_mail_fim:
        print(f"[EMAIL] Falha ao enviar email de conclusao: {_e_mail_fim}")


if __name__ == "__main__":
    try:
        main()
    except KeyboardInterrupt:
        print("\n\n[INTERROMPIDO] Execucao cancelada pelo usuario.")
        print("[DIAGNOSTICO] O log de execucao foi registrado durante o processo.")
    except Exception as _e_critico_inicial:
        # ❌ ERRO CRÍTICO NO INÍCIO (antes do loop de empresas)
        _msg_erro_critico = f"ERRO CRÍTICO INICIAL: {str(_e_critico_inicial)}\n\n{traceback.format_exc()}"
        print(f"\n❌ {_msg_erro_critico}")
        
        # Registra no DIAGNOSTICO
        try:
            _diag_file = os.path.join(PASTA_DIAG, f"DIAGNOSTICO_ERROS_PADRAO_NACIONAL_{_comp_atual}.txt")
            with open(_diag_file, "a", encoding="utf-8") as _fd:
                _fd.write("\n" + "=" * 90 + "\n")
                _fd.write(f"Data/Hora  : {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}\n")
                _fd.write(_msg_erro_critico + "\n")
                _fd.write("=" * 90 + "\n")
        except Exception:
            pass
        
        # Manda EMAIL DE NÃO INICIADA
        try:
            _disparar_email_nao_iniciada(_e_critico_inicial)
        except Exception as _e_email_erro:
            print(f"[EMAIL] Falha ao enviar email de nao iniciada: {_e_email_erro}")
        
        import sys as _sys_exit
        _sys_exit.exit(1)
