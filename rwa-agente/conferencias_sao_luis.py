# -*- coding: utf-8 -*-
import os
import re
import sys
import queue   as _q_mod
import tkinter as _tk_mod
import threading
import shutil
import traceback
from datetime import datetime
from pathlib import Path

import pdfplumber
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ─────────────────────────────────────────────────────────────────────────────
# JANELA DE PROGRESSO — RWA Soluções
# ─────────────────────────────────────────────────────────────────────────────
class RWAProgressWindow:
    def __init__(self):
        self._queue   = _q_mod.Queue()
        self._pronto  = threading.Event()
        self._fechado = threading.Event()
        self._inicio  = None
        self._animar  = True
        self._thread  = threading.Thread(target=self._run, daemon=True)
        self._thread.start()
        self._pronto.wait(timeout=4)

    def write(self, texto):
        if texto:
            self._queue.put(("LOG", texto))
            m = re.search(r'Processando codigo\s+(\S+)', texto)
            if m:
                self._queue.put(("EMP", m.group(1).strip()))

    def flush(self):
        pass

    def finalizar(self):
        self._animar = False
        self._queue.put(("FIM",))
        self._fechado.wait()

    def _run(self):
        import time as _time
        root = _tk_mod.Tk()
        root.title("RWA Soluções — Conferencias Dominio X Automações")
        root.resizable(False, False)
        root.configure(bg="#13131f")
        W, H = 660, 560
        root.geometry(f"{W}x{H}+{(root.winfo_screenwidth()-W)//2}+{(root.winfo_screenheight()-H)//2}")

        _tk_mod.Frame(root, bg="#4f46e5", height=5).pack(fill="x")
        _hdr = _tk_mod.Frame(root, bg="#13131f")
        _hdr.pack(fill="x", padx=30, pady=(18, 0))
        _tk_mod.Label(_hdr, text="📊", font=("Arial", 22), bg="#13131f", fg="#4f46e5").pack(side="left", padx=(0, 12))
        _col = _tk_mod.Frame(_hdr, bg="#13131f")
        _col.pack(side="left")
        _tk_mod.Label(_col, text="Conferencias Dominio X Automações", font=("Arial", 15, "bold"), bg="#13131f", fg="#ffffff", anchor="w").pack(anchor="w")
        self._lbl_timer = _tk_mod.Label(_hdr, text="00:00:00", font=("Consolas", 13, "bold"), bg="#13131f", fg="#333355")
        self._lbl_timer.pack(side="right")

        _tk_mod.Frame(root, bg="#1e1e30", height=1).pack(fill="x", padx=30, pady=(14, 0))
        _card = _tk_mod.Frame(root, bg="#1a1a2e")
        _card.pack(fill="x", padx=30, pady=(14, 0))
        _tk_mod.Label(_card, text="PROCESSANDO", font=("Arial", 7, "bold"), bg="#1a1a2e", fg="#333366", anchor="w").pack(anchor="w", padx=14, pady=(10, 2))
        self._lbl_emp = _tk_mod.Label(_card, text="Iniciando...", font=("Arial", 11, "bold"), bg="#1a1a2e", fg="#aaaadd", anchor="w")
        self._lbl_emp.pack(anchor="w", padx=14, pady=(0, 10))

        _fp = _tk_mod.Frame(root, bg="#13131f")
        _fp.pack(fill="x", padx=30, pady=(14, 0))
        _tk_mod.Label(_fp, text="PROGRESSO", font=("Arial", 7, "bold"), bg="#13131f", fg="#333366").pack(anchor="w")
        _tk_mod.Frame(_fp, bg="#1e1e30", height=1).pack(fill="x", pady=(4, 6))
        self._canvas = _tk_mod.Canvas(_fp, height=10, bg="#1a1a2e", highlightthickness=0, bd=0)
        self._canvas.pack(fill="x")
        self._barra       = self._canvas.create_rectangle(0, 0, 0, 10, fill="#4f46e5", outline="")
        self._barra_shine = self._canvas.create_rectangle(0, 0, 0,  3, fill="#7c73f5", outline="")

        _fl = _tk_mod.Frame(root, bg="#13131f")
        _fl.pack(fill="both", expand=True, padx=30, pady=(14, 0))
        _tk_mod.Label(_fl, text="LOG DE EXECUÇÃO", font=("Arial", 7, "bold"), bg="#13131f", fg="#333366", anchor="w").pack(anchor="w")
        _tk_mod.Frame(_fl, bg="#1e1e30", height=1).pack(fill="x", pady=(4, 6))
        _fr_txt = _tk_mod.Frame(_fl, bg="#0d0d1a")
        _fr_txt.pack(fill="both", expand=True)
        self._txt = _tk_mod.Text(_fr_txt, bg="#0d0d1a", fg="#7777aa", font=("Consolas", 8), relief="flat",
                                  state="disabled", wrap="none", height=12, selectbackground="#2a2a4a")
        self._txt.tag_config("ok",    foreground="#22c55e")
        self._txt.tag_config("erro",  foreground="#ef4444")
        self._txt.tag_config("aviso", foreground="#f59e0b")
        self._txt.tag_config("info",  foreground="#6366f1")
        _sb = _tk_mod.Scrollbar(_fr_txt, bg="#1a1a2e", troughcolor="#13131f", command=self._txt.yview)
        self._txt.configure(yscrollcommand=_sb.set)
        _sb.pack(side="right", fill="y")
        self._txt.pack(fill="both", expand=True, padx=6, pady=6)

        _rod = _tk_mod.Frame(root, bg="#13131f")
        _rod.pack(fill="x", padx=30, pady=(10, 14))
        _tk_mod.Label(_rod, text="Desenvolvido por RWA Soluções", font=("Arial", 7), bg="#13131f", fg="#252540").pack(side="left")
        self._btn = _tk_mod.Button(_rod, text="⏳  Processando...", font=("Arial", 9, "bold"),
                                    bg="#252538", fg="#444466", relief="flat", padx=18, pady=5,
                                    state="disabled", cursor="arrow")
        self._btn.pack(side="right")

        self._root  = root
        self._inicio = _time.time()
        self._pos    = 0
        self._dir    = 1
        self._pronto.set()
        root.after(50,   self._poll)
        root.after(1000, self._tick)
        root.after(40,   self._anim)
        root.mainloop()
        self._fechado.set()

    def _tick(self):
        import time as _time
        if self._inicio and not self._fechado.is_set():
            s = int(_time.time() - self._inicio)
            self._lbl_timer.configure(text=f"{s//3600:02d}:{(s%3600)//60:02d}:{s%60:02d}", fg="#4f46e5")
            self._root.after(1000, self._tick)

    def _anim(self):
        if self._animar and not self._fechado.is_set():
            B = 80
            x1, x2 = self._pos, min(self._pos + B, 598)
            self._canvas.coords(self._barra,       x1, 0, x2, 10)
            self._canvas.coords(self._barra_shine, x1, 0, x2,  3)
            self._pos += self._dir * 6
            if self._pos + B >= 598: self._dir = -1
            if self._pos <= 0:       self._dir =  1
            self._root.after(40, self._anim)

    def _poll(self):
        try:
            while True:
                item = self._queue.get_nowait()
                if item[0] == "LOG":
                    linha = item[1]
                    self._txt.configure(state="normal")
                    tag = ("ok"    if any(x in linha for x in ["OK -", "✓", "Concluido", "consistente"]) else
                           "erro"  if any(x in linha for x in ["ERRO", "✗", "nao encontrad"]) else
                           "aviso" if any(x in linha for x in ["AVISO", "DIVERGENCIA", "PENDENCIA"]) else
                           "info"  if any(x in linha for x in ["Processando", "CONFERENCIAS", "EMPRESAS"]) else
                           None)
                    self._txt.insert("end", linha, tag) if tag else self._txt.insert("end", linha)
                    self._txt.see("end")
                    self._txt.configure(state="disabled")
                elif item[0] == "EMP":
                    self._lbl_emp.configure(text=f"Código {item[1]}", fg="#ffffff")
                elif item[0] == "FIM":
                    self._animar = False
                    self._canvas.coords(self._barra,       0, 0, 598, 10)
                    self._canvas.coords(self._barra_shine, 0, 0, 598,  3)
                    self._canvas.itemconfig(self._barra,       fill="#22c55e")
                    self._canvas.itemconfig(self._barra_shine, fill="#4ade80")
                    self._lbl_emp.configure(text="✅  Concluído com sucesso!", fg="#22c55e")
                    self._lbl_timer.configure(fg="#22c55e")
                    self._inicio = None
                    self._btn.configure(text="Concluído", state="normal",
                                        bg="#22c55e", fg="#0a0a0a",
                                        cursor="hand2", command=self._root.destroy)
        except _q_mod.Empty:
            pass
        if not self._fechado.is_set():
            self._root.after(50, self._poll)

# ==========================================================
# CONFERENCIA SIMPLES NACIONAL
# Compara: STM Prestado x Dominio Servicos x Dominio Simples
# ==========================================================

def _ler_paths_launcher():
    try:
        import json as _json
        p = Path(os.environ.get("LOCALAPPDATA", "")) / "RWA_AUTOMACOES" / "config" / "paths.json"
        if p.exists():
            return _json.loads(p.read_text("utf-8"))
    except Exception:
        pass
    return {}

_paths_cfg = _ler_paths_launcher()

PASTA_CONFERENCIAS = Path(
    os.environ.get("RWA_PASTA_CONF")
    or _paths_cfg.get("pasta_conf_sl")
    or r"G:\Meu Drive\AUTOMACOES\CONFERENCIAS"
)
PASTA_EMPRESAS = Path(
    os.environ.get("RWA_PASTA_EMPRESAS")
    or _paths_cfg.get("pasta_emp_sl")
    or r"G:\Meu Drive\AUTOMACOES\SAO LUIS"
)
ARQUIVO_SAIDA = PASTA_CONFERENCIAS / "CONFERENCIAS_DOMINIO_X_AUTOMACOES.xlsx"
SENHAS_XLSX   = Path(
    _paths_cfg.get("senhas_sl")
    or (Path(__file__).parent / "SENHAS SAO LUIS.xlsx")
)

TOLERANCIA = 0.01

CFOP_DEVOLUCAO = {"5201","5202","5411","5553","5556","5918","6201","6202","6411","6553","6556","6918"}

PADRAO_SERVICOS  = re.compile(r"^Servi[cç]os_(\d+)_(\d{6})(?:\.pdf)?$", re.IGNORECASE)
PADRAO_SIMPLES   = re.compile(r"^Simples Nacional_(\d+)_(\d{6})(?:\.pdf)?$", re.IGNORECASE)
PADRAO_STM       = re.compile(r"^servico_prestado_saoluis_(\d+)_(\d{6})(?:\.pdf)?$", re.IGNORECASE)
PADRAO_SAIDAS    = re.compile(r"^Sa[ií]das_(\d+)_(\d{6})(?:\.pdf)?$", re.IGNORECASE)
PADRAO_GERENCIAL = re.compile(r"^Relatorio_Gerencial_Sao_Luis_(\d{6})\.xlsx$", re.IGNORECASE)

_linhas_log      = []
_competencia_log = ""


def log(msg):
    print(msg, flush=True)
    _linhas_log.append(str(msg))


def salvar_log():
    try:
        sufixo = _competencia_log if _competencia_log else datetime.now().strftime("%d%m%Y_%H%M%S")
        caminho = PASTA_CONFERENCIAS / f"CONCLUSAO_CONFERENCIA_{sufixo}.txt"
        with open(str(caminho), "w", encoding="utf-8") as f:
            f.write("\n".join(_linhas_log))
        print(f"\n[LOG] Conclusao salva em: {caminho}", flush=True)
    except Exception as e:
        print(f"\n[AVISO] Nao foi possivel salvar o log: {e}", flush=True)


def texto_pdf(caminho):
    texto = ""
    with pdfplumber.open(str(caminho)) as pdf:
        for pagina in pdf.pages:
            texto += "\n" + (pagina.extract_text() or "")
    return texto


def pegar_valores_monetarios(texto):
    return re.findall(r"\d{1,3}(?:\.\d{3})*,\d{2}|\d+,\d{2}", texto)


def para_numero(valor):
    if not valor: return None
    return float(valor.replace(".", "").replace(",", "."))


def formatar(valor):
    if valor is None: return ""
    return f"{valor:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")


def normalizar_cnpj(texto):
    """Remove tudo exceto dígitos."""
    return re.sub(r"\D", "", texto or "")


def formatar_cnpj(digitos):
    """Formata 14 dígitos como XX.XXX.XXX/XXXX-XX."""
    c = re.sub(r"\D", "", digitos or "")
    if len(c) == 14:
        return f"{c[:2]}.{c[2:5]}.{c[5:8]}/{c[8:12]}-{c[12:14]}"
    return digitos


def normalizar_competencia(texto):
    """Normaliza qualquer formato para MM/AAAA."""
    if not texto:
        return ""
    m = re.search(r"\b(\d{2})/(\d{4})\b", texto)
    if m:
        return f"{m.group(1)}/{m.group(2)}"
    m = re.search(r"\d{2}/(\d{2})/(\d{4})", texto)
    if m:
        return f"{m.group(1)}/{m.group(2)}"
    return texto.strip()


def extrair_cnpj_competencia_stm(pdf):
    """STM: CNPJ na linha Contribuinte | Competencia na linha TOTAIS/Periodo Consulta."""
    linhas = texto_pdf(pdf).splitlines()
    cnpj = comp = ""
    for linha in linhas:
        if not cnpj:
            m = re.search(r"Contribuinte:\s*([\d./-]{14,22})", linha)
            if m:
                cnpj = normalizar_cnpj(m.group(1))
        if not comp:
            if "TOTAIS" in linha.upper() or "Periodo Consulta" in linha:
                m = re.search(r"(\d{2}/\d{4})", linha)
                if m:
                    comp = m.group(1)
        if cnpj and comp:
            break
    return cnpj, comp


def extrair_cnpj_competencia_servicos(pdf):
    """Servicos: CNPJ na linha CNPJ | Competencia na linha Periodo DD/MM/AAAA."""
    linhas = texto_pdf(pdf).splitlines()
    cnpj = comp = ""
    for linha in linhas:
        if not cnpj:
            m = re.search(r"CNPJ:\s*([\d./-]+)", linha)
            if m:
                cnpj = normalizar_cnpj(m.group(1))
        if not comp:
            m = re.search(r"Per[íi]odo:\s*(\d{2}/\d{2}/\d{4})", linha)
            if m:
                comp = normalizar_competencia(m.group(1))
        if cnpj and comp:
            break
    return cnpj, comp


def extrair_cnpj_competencia_simples(pdf):
    """Simples Nacional: CNPJ na linha CNPJ | Competencia na linha Periodo MM/AAAA."""
    linhas = texto_pdf(pdf).splitlines()
    cnpj = comp = ""
    for linha in linhas:
        if not cnpj:
            m = re.search(r"CNPJ:\s*([\d./-]+)", linha)
            if m:
                cnpj = normalizar_cnpj(m.group(1))
        if not comp:
            m = re.search(r"Per[íi]odo:\s*(\d{2}/\d{4})", linha)
            if m:
                comp = m.group(1)
        if cnpj and comp:
            break
    return cnpj, comp


def encontrar_pasta_empresa(codigo):
    padrao = re.compile(rf"^{re.escape(str(codigo))}\s*[-]\s*", re.IGNORECASE)
    for pasta in PASTA_EMPRESAS.iterdir():
        if pasta.is_dir() and padrao.match(pasta.name):
            return pasta
    return None


def ler_regime_senhas(codigo):
    try:
        import openpyxl as _opx
        wb = _opx.load_workbook(str(SENHAS_XLSX), read_only=True, data_only=True)
        ws = wb.active
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row and str(row[0]).strip() == str(codigo).strip():
                return str(row[5]).strip() if len(row) > 5 and row[5] else ""
        return ""
    except Exception:
        return ""


def ler_gerencial_slz(path):
    sem_mov = {}
    try:
        import openpyxl as _opx
        wb = _opx.load_workbook(str(path), read_only=True, data_only=True)
        ws = wb.active
        for row in ws.iter_rows(min_row=3, values_only=True):
            if not row or row[0] is None:
                continue
            codigo = str(row[0]).strip()
            if not codigo.isdigit():
                continue
            qtd = str(row[3] or "").strip()
            if "sem movimento" in qtd.lower():
                sem_mov[codigo] = {
                    "empresa": str(row[1] or "").strip(),
                    "regime":  str(row[2] or "").strip(),
                }
    except Exception:
        pass
    return sem_mov


def extrair_stm_vl_serv(pdf):
    linhas = texto_pdf(pdf).splitlines()
    for i, linha in enumerate(linhas):
        if "TOTAIS" in linha.upper():
            bloco = linha + (" " + linhas[i + 1] if i + 1 < len(linhas) else "")
            valores = pegar_valores_monetarios(bloco)
            if len(valores) >= 3:
                return para_numero(valores[0]), para_numero(valores[1]), para_numero(valores[2])
    return None, None, None


def extrair_servicos_total_geral(pdf):
    linhas = texto_pdf(pdf).splitlines()
    for i, linha in enumerate(linhas):
        if "Total Geral" in linha:
            valores = pegar_valores_monetarios(linha)
            if valores:
                return para_numero(valores[0])
            for prox in linhas[i + 1:i + 8]:
                valores = pegar_valores_monetarios(prox)
                if valores:
                    return para_numero(valores[0])
    return None


def extrair_saidas_total_geral(pdf):
    total_geral      = None
    devolucoes       = 0.0
    pegar_prox_total = False
    try:
        with pdfplumber.open(str(pdf)) as _doc:
            for page in _doc.pages:
                words   = page.extract_words()
                cfop_x  = next((w["x0"] for w in words if w["text"].upper() == "CFOP"), 271.0)
                vcon_x  = next((w["x0"] for w in words if "ntábil" in w["text"] or "ntabil" in w["text"].lower()), 335.0)
                linhas_pos = {}
                for w in words:
                    linhas_pos.setdefault(round(w["top"]), []).append(w)
                for top in sorted(linhas_pos):
                    row      = linhas_pos[top]
                    txt      = " ".join(w["text"] for w in row)
                    if pegar_prox_total:
                        vv = pegar_valores_monetarios(txt)
                        if vv:
                            total_geral = para_numero(vv[0])
                        pegar_prox_total = False
                        continue
                    if "Total Geral" in txt and "Total Dia" not in txt:
                        pegar_prox_total = True
                        continue
                    cfop_words = [w for w in row if cfop_x - 15 <= w["x0"] <= cfop_x + 35]
                    for cw in cfop_words:
                        m = re.search(r"(\d)[A-Za-z]?-(\d{3})", cw["text"])
                        if m:
                            cfop = m.group(1) + m.group(2)
                            if cfop in CFOP_DEVOLUCAO:
                                for vw in sorted([w for w in row if w["x0"] >= vcon_x - 10], key=lambda w: w["x0"]):
                                    vv = pegar_valores_monetarios(vw["text"])
                                    if vv:
                                        devolucoes = round(devolucoes + para_numero(vv[0]), 2)
                                        break
    except Exception:
        pass
    if total_geral is None:
        return None, None
    net = round(total_geral - devolucoes, 2)
    dev = round(devolucoes, 2) if devolucoes > 0 else None
    return net, dev


def extrair_simples_competencia(pdf):
    linhas = texto_pdf(pdf).splitlines()
    for i, linha in enumerate(linhas):
        if "Regime de Compet" in linha:
            valores = pegar_valores_monetarios(linha)
            if valores:
                return para_numero(valores[-1])
            for prox in linhas[i + 1:i + 8]:
                valores = pegar_valores_monetarios(prox)
                if valores:
                    return para_numero(valores[-1])
    return None


def dentro_tolerancia(valor):
    if valor is None: return False
    return abs(valor) <= TOLERANCIA


def definir_status(vl_serv, dom_servicos, dif_servicos, observacoes, regime="", dom_simples=None, dom_saidas=None):
    if observacoes:
        return "PENDENCIA"
    if vl_serv is None or dom_servicos is None:
        return "PENDENCIA"

    # Frente 1: STM x Serviços Domínio
    if not dentro_tolerancia(dif_servicos):
        return "DIVERGENTE"

    # NORMAL: apenas frente 1
    if regime and "NORMAL" in regime.upper():
        return "OK"

    # SIMPLES: frente 2 — (Serviços + Saídas) x Simples
    if dom_simples is None:
        return "PENDENCIA"

    soma_dominio     = round((dom_servicos or 0) + (dom_saidas or 0), 2)
    dif_simples_calc = round(soma_dominio - dom_simples, 2)

    if dentro_tolerancia(dif_simples_calc):
        return "OK"

    if dom_simples > soma_dominio:
        return "PARCIAL"

    return "DIVERGENTE"


def gerar_xlsx(resultados, resultados_sem_mov):
    COR_HEADER   = "1F3864"
    COR_GRUPO_E  = "375623"
    COR_GRUPO_D  = "2E75B6"
    COR_ZEBRA    = "F2F2F2"
    COR_BRANCA   = "FFFFFF"
    COR_VERDE    = "C6EFCE"
    COR_VERMELHO = "FFC7CE"
    COR_AMARELO  = "FFEB9C"
    COR_LARANJA  = "FCE4D6"
    COR_VERM_LIN = "FFE0E0"
    FONT_VERDE   = "276221"
    FONT_VERM    = "9C0006"
    FONT_AMAR    = "9C5700"
    FONT_LARAN   = "843C0C"

    ACAO_MAP = {
        "Pasta da empresa nao encontrada":                    "Localizar pasta da empresa no sistema",
        "Relatorio STM (Prefeitura) nao encontrado na pasta": "Confirmar importação STM (prefeitura)",
        "CNPJ divergente":                                    "Empresas não são as mesmas",
        "Diferenca de valores":                               "Verificar lançamentos no Domínio",
        "Simples maior que servicos+saidas":                  "Verificar saídas não lançadas no Domínio",
        "Empresa com regime normal":                           "Conferir relatorio dominio salvos manualmente.",
        "relatorio do simples ausente":                       "Verificar ausencia do relatorio do simples",
    }

    def _acao(obs_list):
        for obs in obs_list:
            for chave, acao in ACAO_MAP.items():
                if chave in obs:
                    return acao
        return ""

    def _bate(dif):
        if dif is None: return None
        return abs(dif) <= TOLERANCIA

    thin  = Side(style="thin",   color="D9D9D9")
    thick = Side(style="medium", color="AAAAAA")
    borda = Border(left=thin, right=thin, top=thin, bottom=thin)

    wb = Workbook()
    ws = wb.active
    ws.title = "CONFERENCIA FINAL"
    ws.sheet_view.showGridLines = False
    ws.row_dimensions[1].height = 22
    ws.row_dimensions[2].height = 6
    ws.row_dimensions[3].height = 26

    ws.merge_cells("F1:G1")
    ws.merge_cells("H1:M1")
    ws.merge_cells("N1:P1")

    for coord, texto, cor in [("F1","PRESTADOS PREFEITURA",COR_GRUPO_E),
                               ("H1","PRESTADOS DOMINIO",   COR_GRUPO_D),
                               ("N1","RESULTADO",           COR_HEADER)]:
        c = ws[coord]; c.value=texto
        c.font      = Font(bold=True, color="FFFFFF", name="Calibri", size=10)
        c.fill      = PatternFill("solid", fgColor=cor)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border    = borda

    colunas = [
        ("A","Código",         8), ("B","Empresa",        36), ("C","CNPJ",           20),
        ("D","Regime",        18), ("E","Competência",    13), ("F","Valor",           15),
        ("G","Deduções",      13), ("H","Rel. Prestados", 16), ("I","Rel. Saídas",     15),
        ("J","Devoluções",    13), ("K","Rel. Simples",   15), ("L","Dif. Prest.",     13),
        ("M","Dif. Simples",  13), ("N","Status",         16), ("O","Observação",      44),
        ("P","Ação",          38),
    ]
    for letra, titulo, larg in colunas:
        c = ws[f"{letra}3"]; c.value=titulo
        c.font      = Font(bold=True, color="FFFFFF", name="Calibri", size=10)
        c.fill      = PatternFill("solid", fgColor=COR_HEADER)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=False)
        c.border    = borda
        ws.column_dimensions[letra].width = larg

    for idx, r in enumerate(resultados, 4):
        ws.row_dimensions[idx].height = 18
        status    = r["status"]
        obs_list  = r.get("observacoes") or []
        sem_pasta = "Pasta da empresa nao encontrada" in obs_list
        cor_bg    = COR_VERM_LIN if sem_pasta else (COR_ZEBRA if idx % 2 == 0 else COR_BRANCA)

        dif_serv      = r.get("diferenca")
        dif_simp      = r.get("dif_simples")
        serv_ok       = _bate(dif_serv)
        simp_ok       = _bate(dif_simp)
        obs_txt       = " | ".join(obs_list)
        acao_txt      = _acao(obs_list)
        regime        = r.get("regime", "")
        eh_normal_emp = regime and "NORMAL" in regime.upper()

        # A B C D E — fundo base
        for letra, val, aln in [("A",r["codigo"],"center"),("B",r["empresa"],"left"),
                                  ("C",r["cnpj"],"center"),  ("D",regime,"center"),
                                  ("E",r["competencia"],"center")]:
            c = ws[f"{letra}{idx}"]; c.value=val
            c.fill      = PatternFill("solid", fgColor=cor_bg)
            c.font      = Font(name="Calibri", size=10)
            c.border    = borda
            c.alignment = Alignment(horizontal=aln, vertical="center",
                                    indent=1 if aln=="left" else 0)

        # F — Valor Prefeitura
        c = ws[f"F{idx}"]; c.value=r["vl_nf"]; c.border=borda
        c.alignment = Alignment(horizontal="right", vertical="center")
        if r["vl_nf"] is not None:
            c.fill=PatternFill("solid",fgColor=COR_VERDE)
            c.font=Font(name="Calibri",size=10,bold=True,color=FONT_VERDE)
            c.number_format="#,##0.00"
        else:
            c.fill=PatternFill("solid",fgColor=COR_VERMELHO if sem_pasta else COR_BRANCA)
            c.font=Font(name="Calibri",size=10)

        # G — Deduções
        c = ws[f"G{idx}"]; c.value=r["deducao"]; c.border=borda
        c.alignment = Alignment(horizontal="right", vertical="center")
        if r["deducao"] is None:
            c.fill=PatternFill("solid",fgColor=COR_VERMELHO if sem_pasta else COR_BRANCA)
            c.font=Font(name="Calibri",size=10)
        elif r["deducao"] == 0:
            c.fill=PatternFill("solid",fgColor=COR_VERDE)
            c.font=Font(name="Calibri",size=10,color=FONT_VERDE)
            c.number_format="#,##0.00"
        else:
            c.fill=PatternFill("solid",fgColor=cor_bg)
            c.font=Font(name="Calibri",size=10)
            c.number_format="#,##0.00"

        # H — Rel. Prestados
        c = ws[f"H{idx}"]; c.border=borda
        c.alignment = Alignment(horizontal="right", vertical="center")
        val = r["dom_servicos"]
        if val is None:
            c.value=None
            c.fill=PatternFill("solid",fgColor=COR_VERMELHO if sem_pasta else COR_BRANCA)
            c.font=Font(name="Calibri",size=10)
        elif serv_ok is True:
            c.value=val; c.fill=PatternFill("solid",fgColor=COR_VERDE)
            c.font=Font(name="Calibri",size=10,color=FONT_VERDE); c.number_format="#,##0.00"
        elif serv_ok is False:
            c.value=val; c.fill=PatternFill("solid",fgColor=COR_VERMELHO)
            c.font=Font(name="Calibri",size=10,color=FONT_VERM); c.number_format="#,##0.00"
        else:
            c.value=val; c.fill=PatternFill("solid",fgColor=cor_bg)
            c.font=Font(name="Calibri",size=10); c.number_format="#,##0.00"

        # I — Rel. Saídas
        c = ws[f"I{idx}"]; c.border=borda
        dom_saidas_val = r.get("dom_saidas")
        if dom_saidas_val is None:
            c.value="X"
            c.fill=PatternFill("solid",fgColor=COR_ZEBRA)
            c.font=Font(name="Calibri",size=10,color="AAAAAA")
            c.alignment=Alignment(horizontal="center",vertical="center")
        else:
            c.value=dom_saidas_val
            c.fill=PatternFill("solid",fgColor=COR_VERDE)
            c.font=Font(name="Calibri",size=10,color=FONT_VERDE)
            c.number_format="#,##0.00"
            c.alignment=Alignment(horizontal="right",vertical="center")

        # J — Devoluções
        c = ws[f"J{idx}"]; c.border=borda
        dom_dev_val = r.get("dom_devolucoes")
        if dom_dev_val is None:
            c.value=None
            c.fill=PatternFill("solid",fgColor=COR_BRANCA)
            c.font=Font(name="Calibri",size=10)
            c.alignment=Alignment(horizontal="center",vertical="center")
        else:
            c.value=dom_dev_val
            c.fill=PatternFill("solid",fgColor=COR_AMARELO)
            c.font=Font(name="Calibri",size=10,bold=True,color=FONT_AMAR)
            c.number_format="#,##0.00"
            c.alignment=Alignment(horizontal="right",vertical="center")

        # K — Rel. Simples
        c = ws[f"K{idx}"]; c.border=borda
        c.alignment=Alignment(horizontal="right",vertical="center")
        if eh_normal_emp:
            c.value="dispensado"
            c.fill=PatternFill("solid",fgColor=COR_ZEBRA)
            c.font=Font(name="Calibri",size=9,color="AAAAAA",italic=True)
            c.alignment=Alignment(horizontal="center",vertical="center")
        else:
            val = r["dom_simples"]
            if val is None:
                c.value="ausente"
                c.fill=PatternFill("solid",fgColor=COR_VERMELHO)
                c.font=Font(name="Calibri",size=9,color=FONT_VERM,italic=True)
                c.alignment=Alignment(horizontal="center",vertical="center")
            elif simp_ok is True:
                c.value=val; c.fill=PatternFill("solid",fgColor=COR_VERDE)
                c.font=Font(name="Calibri",size=10,color=FONT_VERDE); c.number_format="#,##0.00"
            elif simp_ok is False:
                c.value=val; c.fill=PatternFill("solid",fgColor=COR_VERMELHO)
                c.font=Font(name="Calibri",size=10,color=FONT_VERM); c.number_format="#,##0.00"
            else:
                c.value=val; c.fill=PatternFill("solid",fgColor=cor_bg)
                c.font=Font(name="Calibri",size=10); c.number_format="#,##0.00"

        # L M — Diferenças
        for col, val, ok in [("L",dif_serv,serv_ok),("M",dif_simp,simp_ok)]:
            c=ws[f"{col}{idx}"]; c.border=borda
            c.alignment=Alignment(horizontal="right",vertical="center")
            if ok is True:
                c.value=None; c.fill=PatternFill("solid",fgColor=COR_BRANCA)
                c.font=Font(name="Calibri",size=10)
            elif ok is False:
                c.value=val; c.fill=PatternFill("solid",fgColor=COR_AMARELO)
                c.font=Font(name="Calibri",size=10,bold=True,color=FONT_AMAR)
                c.number_format="#,##0.00"
            else:
                c.value=None
                c.fill=PatternFill("solid",fgColor=COR_VERMELHO if sem_pasta else COR_BRANCA)
                c.font=Font(name="Calibri",size=10)

        # N — Status
        c=ws[f"N{idx}"]; c.border=borda
        c.alignment=Alignment(horizontal="center",vertical="center")
        if status=="OK":
            c.value="✓ CONSISTENTE"; c.fill=PatternFill("solid",fgColor=COR_VERDE)
            c.font=Font(name="Calibri",size=10,bold=True,color=FONT_VERDE)
        elif status=="DIVERGENTE":
            c.value="✕ DIVERGENTE";  c.fill=PatternFill("solid",fgColor=COR_VERMELHO)
            c.font=Font(name="Calibri",size=10,bold=True,color=FONT_VERM)
        elif status=="PARCIAL":
            c.value="⚠ PARCIAL";     c.fill=PatternFill("solid",fgColor=COR_LARANJA)
            c.font=Font(name="Calibri",size=10,bold=True,color=FONT_LARAN)
        else:
            c.value="⚠ PENDÊNCIA";   c.fill=PatternFill("solid",fgColor=COR_AMARELO)
            c.font=Font(name="Calibri",size=10,bold=True,color=FONT_AMAR)

        # O — Observação
        c=ws[f"O{idx}"]; c.value=obs_txt; c.border=borda
        c.fill=PatternFill("solid",fgColor=COR_VERM_LIN if sem_pasta else (cor_bg if obs_txt else COR_BRANCA))
        c.font=Font(name="Calibri",size=9,italic=bool(obs_txt),color="595959")
        c.alignment=Alignment(horizontal="left",vertical="center",indent=1,wrap_text=False)

        # P — Ação
        c=ws[f"P{idx}"]; c.value=acao_txt; c.border=borda
        c.fill=PatternFill("solid",fgColor=COR_VERM_LIN if sem_pasta else (cor_bg if acao_txt else COR_BRANCA))
        c.font=Font(name="Calibri",size=9,color="1F3864",bold=bool(acao_txt))
        c.alignment=Alignment(horizontal="left",vertical="center",indent=1,wrap_text=False)

    # Resumo analisados
    linha_res  = len(resultados) + 4
    ws.row_dimensions[linha_res].height = 20
    ok_ct      = sum(1 for r in resultados if r["status"] == "OK")
    div_ct     = sum(1 for r in resultados if r["status"] == "DIVERGENTE")
    pend_ct    = sum(1 for r in resultados if r["status"] == "PENDENCIA")
    parcial_ct = sum(1 for r in resultados if r["status"] == "PARCIAL")
    ws.merge_cells(f"A{linha_res}:P{linha_res}")
    c = ws.cell(linha_res, 1)
    c.value     = (f"Total: {len(resultados)}  |  ✓ Consistentes: {ok_ct}  |"
                   f"  ✕ Divergentes: {div_ct}  |  ⚠ Parciais: {parcial_ct}  |  ⚠ Pendentes: {pend_ct}")
    c.fill      = PatternFill("solid", fgColor=COR_HEADER)
    c.font      = Font(bold=True, color="FFFFFF", name="Calibri", size=10)
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border    = Border(left=thick, right=thick, top=thick, bottom=thick)

    # ── Seção: SEM MOVIMENTO ─────────────────────────────────────────────────
    linha_sm = linha_res + 2
    ws.row_dimensions[linha_sm].height = 22
    ws.merge_cells(f"A{linha_sm}:P{linha_sm}")
    c = ws.cell(linha_sm, 1)
    c.value     = f"EMPRESAS SEM MOVIMENTO ({len(resultados_sem_mov)}) — não analisadas"
    c.fill      = PatternFill("solid", fgColor="7F7F7F")
    c.font      = Font(bold=True, color="FFFFFF", name="Calibri", size=10)
    c.alignment = Alignment(horizontal="center", vertical="center")
    linha_sm += 1

    if resultados_sem_mov:
        for r in resultados_sem_mov:
            ws.row_dimensions[linha_sm].height = 18
            ws.cell(linha_sm, 1).value     = r["codigo"]
            ws.cell(linha_sm, 1).alignment = Alignment(horizontal="center", vertical="center")
            ws.merge_cells(start_row=linha_sm, start_column=2, end_row=linha_sm, end_column=16)
            ws.cell(linha_sm, 2).value     = r["empresa"]
            ws.cell(linha_sm, 2).alignment = Alignment(horizontal="left", indent=1, vertical="center")
            for col in range(1, 17):
                ws.cell(linha_sm, col).font = Font(name="Calibri", size=10, bold=True, color=FONT_AMAR)
                ws.cell(linha_sm, col).fill = PatternFill("solid", fgColor=COR_AMARELO)
            linha_sm += 1
    else:
        ws.merge_cells(f"A{linha_sm}:P{linha_sm}")
        c = ws.cell(linha_sm, 1)
        c.value     = "(nenhuma)"
        c.font      = Font(name="Calibri", size=9, italic=True, color="888888")
        c.alignment = Alignment(horizontal="center", vertical="center")

    ws.auto_filter.ref = "A3:P3"
    ws.freeze_panes    = "A4"
    wb.save(str(ARQUIVO_SAIDA))
    log(f"\nPlanilha gerada: {ARQUIVO_SAIDA}")


def main():
    _pw = RWAProgressWindow()
    _stdout_original = sys.stdout
    sys.stdout = _pw
    try:
        _main_corpo()
    except Exception:
        log("")
        log("ERRO FATAL:")
        log(traceback.format_exc())
    finally:
        sys.stdout = _stdout_original
        try:
            salvar_log()
        except Exception:
            pass
        _pw.finalizar()


def _main_corpo():
    log("=" * 70)
    log("CONFERENCIAS DOMINIO X AUTOMACOES")
    log("=" * 70)
    log(f"CONFERENCIAS : {PASTA_CONFERENCIAS}")
    log(f"EMPRESAS     : {PASTA_EMPRESAS}")
    log("")

    if not PASTA_CONFERENCIAS.exists():
        raise FileNotFoundError(f"Pasta nao encontrada: {PASTA_CONFERENCIAS}")
    if not PASTA_EMPRESAS.exists():
        raise FileNotFoundError(f"Pasta nao encontrada: {PASTA_EMPRESAS}")

    gerencial_path = None
    for arq in PASTA_EMPRESAS.iterdir():
        if arq.is_file() and PADRAO_GERENCIAL.match(arq.name):
            gerencial_path = arq
            break
    sem_mov_ger = ler_gerencial_slz(gerencial_path) if gerencial_path else {}
    if sem_mov_ger:
        log(f"  Gerencial: {len(sem_mov_ger)} empresa(s) sem movimento identificada(s).")

    grupos = {}
    for arquivo in PASTA_CONFERENCIAS.iterdir():
        if not arquivo.is_file(): continue
        m = PADRAO_SERVICOS.match(arquivo.name)
        if m:
            grupos.setdefault((m.group(1), m.group(2)), {})["servicos"] = arquivo
            continue
        m = PADRAO_SIMPLES.match(arquivo.name)
        if m:
            grupos.setdefault((m.group(1), m.group(2)), {})["simples"] = arquivo
            continue
        m = PADRAO_SAIDAS.match(arquivo.name)
        if m:
            grupos.setdefault((m.group(1), m.group(2)), {})["saidas"] = arquivo

    if not grupos:
        log("Nenhum PDF Dominio encontrado.")
        log("Esperado: Servicos_[cod]_[mmaaaa].pdf | Simples Nacional_[cod]_[mmaaaa].pdf")
        return

    resultados         = []
    resultados_sem_mov = []

    for (codigo, competencia), docs in sorted(grupos.items(), key=lambda x: int(x[0][0])):
        global _competencia_log
        if not _competencia_log:
            _competencia_log = competencia
        log("-" * 70)
        log(f"Processando codigo {codigo} | competencia {competencia}")

        if codigo in sem_mov_ger:
            r_sm = sem_mov_ger[codigo]
            resultados_sem_mov.append({
                "codigo": codigo, "empresa": r_sm["empresa"],
                "cnpj": "", "regime": r_sm["regime"], "competencia": competencia,
            })
            log(f"  {r_sm['empresa']}: SEM MOVIMENTO — descartada da análise.")
            continue

        observacoes  = []
        erro_critico = False

        regime_empresa = ler_regime_senhas(codigo)
        eh_normal      = regime_empresa and "NORMAL" in regime_empresa.upper()

        pasta_empresa = encontrar_pasta_empresa(codigo)
        if not pasta_empresa:
            msg = "Pasta da empresa nao encontrada"
            log(f"  ERRO: {msg}")
            resultados.append({
                "codigo": codigo, "empresa": "", "cnpj": "", "regime": regime_empresa,
                "competencia": competencia, "vl_nf": None, "deducao": None, "vl_serv": None,
                "dom_servicos": None, "dom_saidas": None, "dom_devolucoes": None, "dom_simples": None,
                "diferenca": None, "dif_simples": None,
                "status": "PENDENCIA", "observacoes": ["Pasta da empresa nao encontrada"],
            })
            continue

        pasta_competencia = pasta_empresa / competencia
        pasta_competencia.mkdir(parents=True, exist_ok=True)

        # STM Prestado
        stm_pdf = None
        for arquivo in pasta_competencia.iterdir():
            if arquivo.is_file() and PADRAO_STM.match(arquivo.name):
                stm_pdf = arquivo
                break

        vl_nf = deducao = vl_serv = None
        if stm_pdf:
            vl_nf, deducao, vl_serv = extrair_stm_vl_serv(stm_pdf)
            if vl_serv is None:
                observacoes.append("Vl. Serv. STM nao localizado")
                erro_critico = True
        else:
            observacoes.append("Relatorio STM (Prefeitura) nao encontrado na pasta")
            erro_critico = True

        # Servicos Dominio
        dom_servicos = None
        pdf_servicos = docs.get("servicos")
        if pdf_servicos:
            dom_servicos = extrair_servicos_total_geral(pdf_servicos)
            if dom_servicos is None:
                observacoes.append("Total Geral Servicos nao localizado")
                erro_critico = True
        else:
            observacoes.append("Arquivo Servicos ausente")
            erro_critico = True

        # Saidas Dominio (opcional)
        dom_saidas     = None
        dom_devolucoes = None
        pdf_saidas = docs.get("saidas")
        if pdf_saidas:
            dom_saidas, dom_devolucoes = extrair_saidas_total_geral(pdf_saidas)

        # Simples Nacional — apenas para SIMPLES
        dom_simples = None
        pdf_simples = docs.get("simples")
        if eh_normal:
            dom_simples = None
        elif pdf_simples:
            dom_simples = extrair_simples_competencia(pdf_simples)
            if dom_simples is None:
                observacoes.append("Regime Competencia Simples nao localizado")
                erro_critico = True
        else:
            observacoes.append("empresa do simples, relatorio do simples ausente")
            erro_critico = True

        # ── Validação CNPJ e Competência ─────────────────────────────────────
        cnpj_stm,  comp_stm  = extrair_cnpj_competencia_stm(stm_pdf)          if stm_pdf      else ("", "")
        cnpj_serv, comp_serv = extrair_cnpj_competencia_servicos(pdf_servicos) if pdf_servicos else ("", "")
        cnpj_simp, comp_simp = extrair_cnpj_competencia_simples(pdf_simples)   if pdf_simples  else ("", "")

        cnpj_display = formatar_cnpj(cnpj_stm or cnpj_serv or cnpj_simp)

        cnpjs_presentes = [c for c in [cnpj_stm, cnpj_serv, cnpj_simp] if c]
        if len(set(cnpjs_presentes)) > 1:
            observacoes.append(
                f"CNPJ divergente — STM:{formatar_cnpj(cnpj_stm)} "
                f"Serv:{formatar_cnpj(cnpj_serv)} SN:{formatar_cnpj(cnpj_simp)}"
            )

        comp_stm_n  = normalizar_competencia(comp_stm)
        comp_serv_n = normalizar_competencia(comp_serv)
        comp_simp_n = normalizar_competencia(comp_simp)
        comps_presentes = [c for c in [comp_stm_n, comp_serv_n, comp_simp_n] if c]
        if len(set(comps_presentes)) > 1:
            observacoes.append(
                f"Competencia divergente — STM:{comp_stm_n} "
                f"Serv:{comp_serv_n} SN:{comp_simp_n}"
            )

        dif_servicos = round(vl_serv - dom_servicos, 2) if vl_serv is not None and dom_servicos is not None else None
        soma_dominio = round((dom_servicos or 0) + (dom_saidas or 0), 2) if dom_servicos is not None else None
        dif_simples  = round(soma_dominio - dom_simples, 2) if soma_dominio is not None and dom_simples is not None else None

        status = definir_status(vl_serv, dom_servicos, dif_servicos, observacoes,
                                regime=regime_empresa, dom_simples=dom_simples, dom_saidas=dom_saidas)

        if not erro_critico and status == "DIVERGENTE":
            observacoes.append("Diferenca de valores")

        if not erro_critico and status == "PARCIAL":
            observacoes.append("Simples maior que servicos+saidas — verificar saidas")

        if eh_normal:
            observacoes.append("Empresa com regime normal")

        if not erro_critico and status == "OK":
            for pdf in docs.values():
                destino = pasta_competencia / pdf.name
                if destino.exists(): destino.unlink()
                shutil.move(str(pdf), str(destino))
            observacoes.append("PDFs Dominio movidos para pasta da empresa.")
            acao = "PDFs Dominio movidos para pasta da empresa - Autorizado."
        else:
            acao = "PDFs NAO movidos - pendencia ou divergencia"

        # ── Log rico por empresa ──────────────────────────────────────────────
        def _ok(v): return "✓" if v else "✗"
        comp_ok   = len(set(comps_presentes)) <= 1 and bool(comps_presentes)
        cnpj_ok   = len(set(cnpjs_presentes)) <= 1 and bool(cnpjs_presentes)
        dif_s_ok  = dentro_tolerancia(dif_servicos)
        dif_si_ok = dentro_tolerancia(dif_simples)

        log(f"  Empresa : {pasta_empresa.name} - CNPJ: {cnpj_display} - Regime: {regime_empresa or 'NAO LOCALIZADO'}")
        log("")
        log(f"  {'CNPJ Prefeitura':<28}: {cnpj_stm or 'NAO LOCALIZADO'} {_ok(cnpj_ok)}")
        log(f"  {'CNPJ Dominio':<28}: {cnpj_serv or cnpj_simp or 'NAO LOCALIZADO'} {_ok(cnpj_ok)}")
        log(f"  {'Compentencia Prefeitura':<28}: {comp_stm_n or 'NAO LOCALIZADO'} {_ok(comp_ok)}")
        log(f"  {'Compentencia Dominio':<28}: {comp_serv_n or comp_simp_n or 'NAO LOCALIZADO'} {_ok(comp_ok)}")
        log(f"  {'Relatorio Prefeitura':<28}: {stm_pdf.name if stm_pdf else 'NAO ENCONTRADO'} {_ok(stm_pdf)}")
        log(f"  {'Relatorio dominio servico':<28}: {pdf_servicos.name if pdf_servicos else 'AUSENTE'} {_ok(pdf_servicos)}")
        log(f"  {'Relatorio dominio Saidas':<28}: {pdf_saidas.name if pdf_saidas else 'AUSENTE (X)'}")
        log(f"  {'Relatorio dominio Simples':<28}: {pdf_simples.name if pdf_simples else ('DISPENSADO' if eh_normal else 'AUSENTE')} {_ok(pdf_simples or eh_normal)}")
        log(f"  {'Total Prefeitura':<28}: {formatar(vl_serv) or 'NAO LOCALIZADO'} {_ok(vl_serv is not None)}")
        log(f"  {'Total dominio servico':<28}: {formatar(dom_servicos) or 'NAO LOCALIZADO'} {_ok(dom_servicos is not None)}")
        log(f"  {'Total dominio Saidas':<28}: {formatar(dom_saidas) or 'AUSENTE'}{f' (dev: {formatar(dom_devolucoes)})' if dom_devolucoes else ''}")
        log(f"  {'Total dominio Simples':<28}: {formatar(dom_simples) or ('DISPENSADO' if eh_normal else 'NAO LOCALIZADO')} {_ok(dom_simples is not None or eh_normal)}")
        log(f"  {'Dif. Servicos':<28}: {formatar(dif_servicos) or 'NAO CALCULADA'} {_ok(dif_s_ok) if dif_servicos is not None else ''}")
        log(f"  {'Dif. Simples':<28}: {formatar(dif_simples) or ('DISPENSADO' if eh_normal else 'NAO CALCULADA')} {_ok(dif_si_ok) if dif_simples is not None else ''}")
        log(f"  Status         : {status}")
        log(f"  Acao           : {acao}")
        if observacoes:
            log(f"  Observacao     : {' | '.join(observacoes)}")

        resultados.append({
            "codigo":       codigo,
            "empresa":      pasta_empresa.name,
            "cnpj":         cnpj_display,
            "regime":       regime_empresa,
            "competencia":  competencia,
            "vl_nf":        vl_nf,
            "deducao":      deducao,
            "vl_serv":      vl_serv,
            "dom_servicos": dom_servicos,
            "dom_saidas":   dom_saidas,
            "dom_devolucoes": dom_devolucoes,
            "dom_simples":  dom_simples,
            "diferenca":    dif_servicos,
            "dif_simples":  dif_simples,
            "status":       status,
            "observacoes":  list(observacoes),
        })

    codigos_processados = {str(r["codigo"]) for r in resultados + resultados_sem_mov}
    for codigo_sm, r_sm in sem_mov_ger.items():
        if codigo_sm not in codigos_processados:
            resultados_sem_mov.append({
                "codigo": codigo_sm, "empresa": r_sm["empresa"],
                "cnpj": "", "regime": r_sm["regime"], "competencia": _competencia_log,
            })
            log(f"  {r_sm['empresa']}: SEM MOVIMENTO (sem arquivos na pasta) — adicionada via gerencial.")

    gerar_xlsx(resultados, resultados_sem_mov)

    log("")
    log("=" * 70)
    log(f"Total: {len(resultados)} registro(s) | Sem movimento: {len(resultados_sem_mov)} | Concluido.")
    log("=" * 70)


if __name__ == "__main__":
    main()
