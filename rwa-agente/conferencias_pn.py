# -*- coding: utf-8 -*-
import os
import re
import sys
import queue   as _q_mod
import tkinter as _tk_mod
import threading
import shutil
import traceback
from datetime import datetime
from pathlib import Path

import openpyxl
import pdfplumber
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ─────────────────────────────────────────────────────────────────────────────
# JANELA DE PROGRESSO — RWA Soluções
# ─────────────────────────────────────────────────────────────────────────────
class RWAProgressWindow:
    def __init__(self):
        self._queue   = _q_mod.Queue()
        self._pronto  = threading.Event()
        self._fechado = threading.Event()
        self._inicio  = None
        self._animar  = True
        self._thread  = threading.Thread(target=self._run, daemon=True)
        self._thread.start()
        self._pronto.wait(timeout=4)

    def write(self, texto):
        if texto:
            self._queue.put(("LOG", texto))
            m = re.search(r'Processando codigo\s+(\S+)', texto)
            if m:
                self._queue.put(("EMP", m.group(1).strip()))

    def flush(self):
        pass

    def finalizar(self):
        self._animar = False
        self._queue.put(("FIM",))
        self._fechado.wait()

    def _run(self):
        import time as _time
        root = _tk_mod.Tk()
        root.title("RWA Soluções — Conferencias Dominio X Automações — Padrão Nacional")
        root.resizable(False, False)
        root.configure(bg="#13131f")
        W, H = 660, 560
        root.geometry(f"{W}x{H}+{(root.winfo_screenwidth()-W)//2}+{(root.winfo_screenheight()-H)//2}")

        _tk_mod.Frame(root, bg="#4f46e5", height=5).pack(fill="x")
        _hdr = _tk_mod.Frame(root, bg="#13131f")
        _hdr.pack(fill="x", padx=30, pady=(18, 0))
        _tk_mod.Label(_hdr, text="📊", font=("Arial", 22), bg="#13131f", fg="#4f46e5").pack(side="left", padx=(0, 12))
        _col = _tk_mod.Frame(_hdr, bg="#13131f")
        _col.pack(side="left")
        _tk_mod.Label(_col, text="Conferencias — Padrão Nacional", font=("Arial", 15, "bold"), bg="#13131f", fg="#ffffff", anchor="w").pack(anchor="w")
        self._lbl_timer = _tk_mod.Label(_hdr, text="00:00:00", font=("Consolas", 13, "bold"), bg="#13131f", fg="#333355")
        self._lbl_timer.pack(side="right")

        _tk_mod.Frame(root, bg="#1e1e30", height=1).pack(fill="x", padx=30, pady=(14, 0))
        _card = _tk_mod.Frame(root, bg="#1a1a2e")
        _card.pack(fill="x", padx=30, pady=(14, 0))
        _tk_mod.Label(_card, text="PROCESSANDO", font=("Arial", 7, "bold"), bg="#1a1a2e", fg="#333366", anchor="w").pack(anchor="w", padx=14, pady=(10, 2))
        self._lbl_emp = _tk_mod.Label(_card, text="Iniciando...", font=("Arial", 11, "bold"), bg="#1a1a2e", fg="#aaaadd", anchor="w")
        self._lbl_emp.pack(anchor="w", padx=14, pady=(0, 10))

        _fp = _tk_mod.Frame(root, bg="#13131f")
        _fp.pack(fill="x", padx=30, pady=(14, 0))
        _tk_mod.Label(_fp, text="PROGRESSO", font=("Arial", 7, "bold"), bg="#13131f", fg="#333366").pack(anchor="w")
        _tk_mod.Frame(_fp, bg="#1e1e30", height=1).pack(fill="x", pady=(4, 6))
        self._canvas = _tk_mod.Canvas(_fp, height=10, bg="#1a1a2e", highlightthickness=0, bd=0)
        self._canvas.pack(fill="x")
        self._barra       = self._canvas.create_rectangle(0, 0, 0, 10, fill="#4f46e5", outline="")
        self._barra_shine = self._canvas.create_rectangle(0, 0, 0,  3, fill="#7c73f5", outline="")

        _fl = _tk_mod.Frame(root, bg="#13131f")
        _fl.pack(fill="both", expand=True, padx=30, pady=(14, 0))
        _tk_mod.Label(_fl, text="LOG DE EXECUÇÃO", font=("Arial", 7, "bold"), bg="#13131f", fg="#333366", anchor="w").pack(anchor="w")
        _tk_mod.Frame(_fl, bg="#1e1e30", height=1).pack(fill="x", pady=(4, 6))
        _fr_txt = _tk_mod.Frame(_fl, bg="#0d0d1a")
        _fr_txt.pack(fill="both", expand=True)
        self._txt = _tk_mod.Text(_fr_txt, bg="#0d0d1a", fg="#7777aa", font=("Consolas", 8), relief="flat",
                                  state="disabled", wrap="none", height=12, selectbackground="#2a2a4a")
        self._txt.tag_config("ok",    foreground="#22c55e")
        self._txt.tag_config("erro",  foreground="#ef4444")
        self._txt.tag_config("aviso", foreground="#f59e0b")
        self._txt.tag_config("info",  foreground="#6366f1")
        _sb = _tk_mod.Scrollbar(_fr_txt, bg="#1a1a2e", troughcolor="#13131f", command=self._txt.yview)
        self._txt.configure(yscrollcommand=_sb.set)
        _sb.pack(side="right", fill="y")
        self._txt.pack(fill="both", expand=True, padx=6, pady=6)

        _rod = _tk_mod.Frame(root, bg="#13131f")
        _rod.pack(fill="x", padx=30, pady=(10, 14))
        _tk_mod.Label(_rod, text="Desenvolvido por RWA Soluções", font=("Arial", 7), bg="#13131f", fg="#252540").pack(side="left")
        self._btn = _tk_mod.Button(_rod, text="⏳  Processando...", font=("Arial", 9, "bold"),
                                    bg="#252538", fg="#444466", relief="flat", padx=18, pady=5,
                                    state="disabled", cursor="arrow")
        self._btn.pack(side="right")

        self._root  = root
        self._inicio = _time.time()
        self._pos    = 0
        self._dir    = 1
        self._pronto.set()
        root.after(50,   self._poll)
        root.after(1000, self._tick)
        root.after(40,   self._anim)
        root.mainloop()
        self._fechado.set()

    def _tick(self):
        import time as _time
        if self._inicio and not self._fechado.is_set():
            s = int(_time.time() - self._inicio)
            self._lbl_timer.configure(text=f"{s//3600:02d}:{(s%3600)//60:02d}:{s%60:02d}", fg="#4f46e5")
            self._root.after(1000, self._tick)

    def _anim(self):
        if self._animar and not self._fechado.is_set():
            B = 80
            x1, x2 = self._pos, min(self._pos + B, 598)
            self._canvas.coords(self._barra,       x1, 0, x2, 10)
            self._canvas.coords(self._barra_shine, x1, 0, x2,  3)
            self._pos += self._dir * 6
            if self._pos + B >= 598: self._dir = -1
            if self._pos <= 0:       self._dir =  1
            self._root.after(40, self._anim)

    def _poll(self):
        try:
            while True:
                item = self._queue.get_nowait()
                if item[0] == "LOG":
                    linha = item[1]
                    self._txt.configure(state="normal")
                    tag = ("ok"    if any(x in linha for x in ["OK -", "✓", "Concluido", "consistente"]) else
                           "erro"  if any(x in linha for x in ["ERRO", "✗", "nao encontrad"]) else
                           "aviso" if any(x in linha for x in ["AVISO", "DIVERGENCIA", "PENDENCIA"]) else
                           "info"  if any(x in linha for x in ["Processando", "CONFERENCIAS", "EMPRESAS"]) else
                           None)
                    self._txt.insert("end", linha, tag) if tag else self._txt.insert("end", linha)
                    self._txt.see("end")
                    self._txt.configure(state="disabled")
                elif item[0] == "EMP":
                    self._lbl_emp.configure(text=f"Código {item[1]}", fg="#ffffff")
                elif item[0] == "FIM":
                    self._animar = False
                    self._canvas.coords(self._barra,       0, 0, 598, 10)
                    self._canvas.coords(self._barra_shine, 0, 0, 598,  3)
                    self._canvas.itemconfig(self._barra,       fill="#22c55e")
                    self._canvas.itemconfig(self._barra_shine, fill="#4ade80")
                    self._lbl_emp.configure(text="✅  Concluído com sucesso!", fg="#22c55e")
                    self._lbl_timer.configure(fg="#22c55e")
                    self._inicio = None
                    self._btn.configure(text="Concluído", state="normal",
                                        bg="#22c55e", fg="#0a0a0a",
                                        cursor="hand2", command=self._root.destroy)
        except _q_mod.Empty:
            pass
        if not self._fechado.is_set():
            self._root.after(50, self._poll)

# ==========================================================
# CONFERENCIA PADRÃO NACIONAL
# Compara:
#   FASE 1 (Gerencial PN x CONCLUSAO PN):
#     1. Cliente: gerencial col A = código CONCLUSAO E col C = nome CONCLUSAO (EXATO)
#     2. Quantidade: gerencial col D = (Emitidas PDF N) = (Emitidas XML N)
#   FASE 2 (Detalhamento NFSe x PDFs Domínio):
#     3. Período: Detalhamento linha 1 = PDF Serviços = PDF Simples Nacional
#     4. CNPJ:    Detalhamento linha 1 = PDF Serviços = PDF Simples Nacional (só dígitos)
#     5. Valor:   Total Geral Serviços = Regime Comp Total Simples = col G linha TOTAL Det
#
# Se todas baterem → move PDFs Domínio (Serviços + Simples Nacional) da pasta
# CONFERENCIAS para a raiz da competência da empresa.
# ==========================================================

def _ler_paths_launcher():
    try:
        import json as _json
        p = Path(os.environ.get("LOCALAPPDATA", "")) / "RWA_AUTOMACOES" / "config" / "paths.json"
        if p.exists():
            return _json.loads(p.read_text("utf-8"))
    except Exception:
        pass
    return {}

_paths_cfg = _ler_paths_launcher()

PASTA_CONFERENCIAS = Path(
    os.environ.get("RWA_PASTA_CONF")
    or _paths_cfg.get("pasta_conf_pn")
    or r"G:\Meu Drive\AUTOMACOES\CENTRAL\CONFERENCIAS"
)
PASTA_EMPRESAS = Path(
    os.environ.get("RWA_PASTA_EMPRESAS")
    or _paths_cfg.get("pasta_emp_pn")
    or r"G:\Meu Drive\AUTOMACOES\CENTRAL\PADRAO NACIONAL"
)
ARQUIVO_SAIDA = PASTA_CONFERENCIAS / "CONFERENCIAS_PADRAO_NACIONAL_DOMINIO_X_AUTOMACOES.xlsx"
SENHAS_XLSX   = Path(
    _paths_cfg.get("senhas_pn")
    or (Path(__file__).parent / "SENHAS PADRAO NACIONAL.xlsx")
)

TOLERANCIA = 0.01

CFOP_DEVOLUCAO = {"5201","5202","5411","5553","5556","5918","6201","6202","6411","6553","6556","6918"}
PADRAO_SAIDAS  = re.compile(r"^Sa[ií]das_(\d+)_(\d{6})(?:\.pdf)?$", re.IGNORECASE)

# Padrões dos arquivos PDF Domínio na pasta CONFERENCIAS
PADRAO_SERVICOS = re.compile(r"^Servi[cç]os_(\d+)_(\d{6})(?:\.pdf)?$", re.IGNORECASE)
PADRAO_SIMPLES  = re.compile(r"^Simples Nacional_(\d+)_(\d{6})(?:\.pdf)?$", re.IGNORECASE)

# Padrão dos arquivos Gerencial e CONCLUSAO da automação PN
PADRAO_GERENCIAL = re.compile(r"^Relatorio_Gerencial_Padrao_Nacional_(\d{6})\.xlsx$", re.IGNORECASE)
PADRAO_CONCLUSAO = re.compile(r"^CONCLUSAO_PADRAO_NACIONAL_(\d{6})\.txt$", re.IGNORECASE)

# Padrão do Detalhamento dentro da subpasta {codigo}_{mmaaaa}
PADRAO_DETALHAMENTO = re.compile(r"^Detalhamento_de_NFSe_.*\.xlsx$", re.IGNORECASE)

_linhas_log      = []
_competencia_log = ""


def log(msg):
    print(msg, flush=True)
    _linhas_log.append(str(msg))


def salvar_log():
    try:
        sufixo = _competencia_log if _competencia_log else datetime.now().strftime("%d%m%Y_%H%M%S")
        caminho = PASTA_CONFERENCIAS / f"CONCLUSAO_CONFERENCIA_PN_{sufixo}.txt"
        with open(str(caminho), "w", encoding="utf-8") as f:
            f.write("\n".join(_linhas_log))
        print(f"\n[LOG] Conclusao salva em: {caminho}", flush=True)
    except Exception as e:
        print(f"\n[AVISO] Nao foi possivel salvar o log: {e}", flush=True)


# ─────────────────────────────────────────────────────────────────────────────
# UTILITARIOS
# ─────────────────────────────────────────────────────────────────────────────

def texto_pdf(caminho):
    texto = ""
    with pdfplumber.open(str(caminho)) as pdf:
        for pagina in pdf.pages:
            texto += "\n" + (pagina.extract_text() or "")
    return texto


def pegar_valores_monetarios(texto):
    return re.findall(r"\d{1,3}(?:\.\d{3})*,\d{2}|\d+,\d{2}", texto)


def para_numero(valor):
    if valor is None:
        return None
    if isinstance(valor, (int, float)):
        return float(valor)

    s = str(valor).strip()
    if not s:
        return None

    # Proteção: o gerencial pode trazer textos como "⚠ Sem movimento"
    # em colunas que normalmente seriam numéricas. Isso não pode derrubar
    # a conferência inteira; deve virar ausência de número e ser tratado
    # como pendência/sem movimento no fluxo.
    s_limpo = re.sub(r"[^0-9,.-]", "", s)
    if not re.search(r"\d", s_limpo):
        return None

    try:
        if "," in s_limpo:
            return float(s_limpo.replace(".", "").replace(",", "."))
        return float(s_limpo)
    except Exception:
        return None


def formatar(valor):
    if valor is None: return ""
    try:
        return f"{float(valor):,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
    except Exception:
        return str(valor)


def normalizar_cnpj(texto):
    return re.sub(r"\D", "", str(texto or ""))


def formatar_cnpj(digitos):
    c = re.sub(r"\D", "", str(digitos or ""))
    if len(c) == 14:
        return f"{c[:2]}.{c[2:5]}.{c[5:8]}/{c[8:12]}-{c[12:14]}"
    return digitos


def encontrar_pasta_empresa(codigo):
    padrao = re.compile(rf"^{re.escape(str(codigo))}\s*[-]\s*", re.IGNORECASE)
    for pasta in PASTA_EMPRESAS.iterdir():
        if pasta.is_dir() and padrao.match(pasta.name):
            return pasta
    return None


def ler_regime_senhas(codigo):
    try:
        import openpyxl as _opx
        wb = _opx.load_workbook(str(SENHAS_XLSX), read_only=True, data_only=True)
        ws = wb.active
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row and str(row[0]).strip() == str(codigo).strip():
                return str(row[2]).strip() if len(row) > 2 and row[2] else ""
        return ""
    except Exception:
        return ""


def extrair_saidas_total_geral(pdf):
    total_geral      = None
    devolucoes       = 0.0
    pegar_prox_total = False
    try:
        with pdfplumber.open(str(pdf)) as _doc:
            for page in _doc.pages:
                words   = page.extract_words()
                cfop_x  = next((w["x0"] for w in words if w["text"].upper() == "CFOP"), 271.0)
                vcon_x  = next((w["x0"] for w in words if "ntábil" in w["text"] or "ntabil" in w["text"].lower()), 335.0)
                linhas_pos = {}
                for w in words:
                    linhas_pos.setdefault(round(w["top"]), []).append(w)
                for top in sorted(linhas_pos):
                    row      = linhas_pos[top]
                    txt      = " ".join(w["text"] for w in row)
                    if pegar_prox_total:
                        vv = pegar_valores_monetarios(txt)
                        if vv:
                            total_geral = para_numero(vv[0])
                        pegar_prox_total = False
                        continue
                    if "Total Geral" in txt and "Total Dia" not in txt:
                        pegar_prox_total = True
                        continue
                    cfop_words = [w for w in row if cfop_x - 15 <= w["x0"] <= cfop_x + 35]
                    for cw in cfop_words:
                        m = re.search(r"(\d)[A-Za-z]?-(\d{3})", cw["text"])
                        if m:
                            cfop = m.group(1) + m.group(2)
                            if cfop in CFOP_DEVOLUCAO:
                                for vw in sorted([w for w in row if w["x0"] >= vcon_x - 10], key=lambda w: w["x0"]):
                                    vv = pegar_valores_monetarios(vw["text"])
                                    if vv:
                                        devolucoes = round(devolucoes + para_numero(vv[0]), 2)
                                        break
    except Exception:
        pass
    if total_geral is None:
        return None, None
    net = round(total_geral - devolucoes, 2)
    dev = round(devolucoes, 2) if devolucoes > 0 else None
    return net, dev


# ─────────────────────────────────────────────────────────────────────────────
# LEITURA DO GERENCIAL PADRÃO NACIONAL
# Estrutura:
#   Linha 1: cabeçalho de grupo (EMITIDAS / RECEBIDAS)
#   Linha 2: cabeçalhos das colunas
#   Linha 3+: dados das empresas
# Colunas: A=Código, B=Inscrição, C=Empresa, D=Emitidas Qtd, E=Regular,
#          F=Canceladas, G=PDF Baixados, H=XML Baixados, ...
# ─────────────────────────────────────────────────────────────────────────────

def ler_gerencial(path):
    wb = openpyxl.load_workbook(str(path), data_only=True)
    ws = wb.active
    empresas = {}
    for row_idx in range(3, ws.max_row + 1):
        cod = ws.cell(row_idx, 1).value
        if cod is None or str(cod).strip() == "":
            continue
        codigo_str = str(cod).strip()
        # Pula linhas inválidas (ex: total/rodapé)
        if not re.match(r"^\d+$", codigo_str):
            continue
        empresas[codigo_str] = {
            "codigo":       codigo_str,
            "inscricao":    str(ws.cell(row_idx, 2).value or "").strip(),
            "empresa":      str(ws.cell(row_idx, 3).value or "").strip(),
            "qtd_emitidas": ws.cell(row_idx, 4).value,
            "pdf_baixados": ws.cell(row_idx, 7).value,
            "xml_baixados": ws.cell(row_idx, 8).value,
        }
    return empresas


# ─────────────────────────────────────────────────────────────────────────────
# LEITURA DO CONCLUSAO PADRÃO NACIONAL
# Identifica seções e classifica empresas em:
#   ANALISAR        -> "EMPRESAS PROCESSADAS COM SUCESSO" (sem o "SEM MOVIMENTO")
#   SEM_MOVIMENTO   -> "PROCESSADAS COM SUCESSO SEM MOVIMENTO NA COMPETENCIA"
#   IGNORADA_CERT   -> "CERTIFICADO NAO LOCALIZADO"
#   IGNORADA_LOGIN  -> "LOGIN OU SENHA INCORRETO"
#   IGNORADA_PORTAL -> "PORTAL INSTAVEL"
# Para cada empresa em ANALISAR captura pdf_qtd e xml_qtd das linhas:
#   "Emitidas PDF                 : ✓ (N)"
#   "Emitidas XML                 : ✓ (N)"
# ─────────────────────────────────────────────────────────────────────────────

def ler_conclusao(path):
    with open(str(path), "r", encoding="utf-8") as f:
        texto = f.read()

    empresas = {}
    secao_atual = None
    ultimo_cod = None

    re_secao = [
        (re.compile(r"EMPRESAS PROCESSADAS COM SUCESSO SEM MOVIMENTO", re.IGNORECASE), "SEM_MOVIMENTO"),
        (re.compile(r"EMPRESAS PROCESSADAS COM SUCESSO",               re.IGNORECASE), "ANALISAR"),
        (re.compile(r"CERTIFICADO NAO LOCALIZADO",                     re.IGNORECASE), "IGNORADA_CERT"),
        (re.compile(r"LOGIN OU SENHA INCORRETO",                       re.IGNORECASE), "IGNORADA_LOGIN"),
        (re.compile(r"PORTAL INSTAVEL",                                re.IGNORECASE), "IGNORADA_PORTAL"),
    ]

    re_empresa = re.compile(r"^\s*Empresa\s*:\s*(\d+)\s*-\s*(.+?)\s*$")
    re_pdf     = re.compile(r"^\s*Emitidas PDF\s*:\s*([✓✗])\s*\((\d+)\)\s*$")
    re_xml     = re.compile(r"^\s*Emitidas XML\s*:\s*([✓✗])\s*\((\d+)\)\s*$")
    re_motivo  = re.compile(r"^\s*Motivo\s*:\s*(.+?)\s*$")

    for linha in texto.splitlines():
        # Detecta mudança de seção (importante: "SEM MOVIMENTO" antes de "PROCESSADAS COM SUCESSO")
        for rgx, grupo in re_secao:
            if rgx.search(linha):
                secao_atual = grupo
                break

        if secao_atual is None:
            continue

        m = re_empresa.match(linha)
        if m:
            codigo  = m.group(1).strip()
            empresa = m.group(2).strip()
            empresas[codigo] = {
                "codigo":   codigo,
                "empresa":  empresa,
                "grupo":    secao_atual,
                "pdf_qtd":  None,
                "xml_qtd":  None,
                "motivo":   "",
            }
            ultimo_cod = codigo
            continue

        if ultimo_cod is None:
            continue

        m = re_pdf.match(linha)
        if m:
            ok  = m.group(1) == "✓"
            qtd = int(m.group(2))
            empresas[ultimo_cod]["pdf_qtd"] = qtd if ok else None
            continue

        m = re_xml.match(linha)
        if m:
            ok  = m.group(1) == "✓"
            qtd = int(m.group(2))
            empresas[ultimo_cod]["xml_qtd"] = qtd if ok else None
            continue

        m = re_motivo.match(linha)
        if m:
            empresas[ultimo_cod]["motivo"] = m.group(1).strip()
            continue

    return empresas


# ─────────────────────────────────────────────────────────────────────────────
# LEITURA DO DETALHAMENTO DE NFSe
# Linha 1: "DETALHAMENTO DE NFS-e — NOME - CNPJ — MM/AAAA - REGIME: ..."
# Procura linha que começa com "TOTAL" e captura col G (VALOR TOTAL NFSe)
# ─────────────────────────────────────────────────────────────────────────────

def ler_detalhamento(path):
    wb = openpyxl.load_workbook(str(path), data_only=True)
    # Preferir aba PRESTADO se existir, senão a primeira
    if "PRESTADO" in wb.sheetnames:
        ws = wb["PRESTADO"]
    else:
        ws = wb.active

    titulo = ws.cell(1, 1).value or ""
    titulo = str(titulo)

    nome = cnpj = comp = ""

    # Competência: primeira ocorrência de MM/AAAA no título
    m = re.search(r"(\d{2}/\d{4})", titulo)
    if m:
        comp = m.group(1)

    # CNPJ: 14 dígitos consecutivos
    m = re.search(r"(\d{14})", titulo)
    if m:
        cnpj = m.group(1)

    # Nome: entre "NFS-e" + separador (— – -) e o CNPJ
    # Aceita em-dash (U+2014), en-dash (U+2013) e hífen normal como separadores
    m = re.search(r"NFS\-?e\s*[\u2014\u2013\-]\s*(.+?)\s*[\u2014\u2013\-]\s*\d{14}", titulo)
    if m:
        nome = m.group(1).strip()

    valor_total = None
    for row_idx in range(2, ws.max_row + 1):
        cel_a = ws.cell(row_idx, 1).value
        if isinstance(cel_a, str) and cel_a.strip().upper().startswith("TOTAL"):
            v = ws.cell(row_idx, 7).value  # col G
            valor_total = para_numero(v) if v is not None else None
            break

    return {
        "nome":         nome,
        "cnpj":         cnpj,
        "competencia":  comp,
        "valor_total":  valor_total,
    }


def encontrar_detalhamento(pasta_empresa, codigo, competencia):
    """
    Procura o Detalhamento_de_NFSe_*.xlsx em duas localizações, nessa ordem:
      1) Raiz da competência:        pasta_empresa/{mmaaaa}/                      (novo padrão)
      2) Subpasta interna (legado):  pasta_empresa/{mmaaaa}/{codigo}_{mmaaaa}/    (fallback)
    Retorna o primeiro arquivo encontrado, ou None se nenhum existir.
    """
    # 1) Raiz da competência (novo padrão)
    pasta_raiz = pasta_empresa / competencia
    if pasta_raiz.exists():
        for arq in pasta_raiz.iterdir():
            if arq.is_file() and PADRAO_DETALHAMENTO.match(arq.name):
                return arq

    # 2) Subpasta interna {codigo}_{mmaaaa} (fallback - execuções antigas)
    pasta_sub = pasta_empresa / competencia / f"{codigo}_{competencia}"
    if pasta_sub.exists():
        for arq in pasta_sub.iterdir():
            if arq.is_file() and PADRAO_DETALHAMENTO.match(arq.name):
                return arq

    return None


# ─────────────────────────────────────────────────────────────────────────────
# EXTRAÇÃO PDF DOMÍNIO — SERVIÇOS
# Captura: CNPJ, Período (MM/AAAA), Total Geral
# Periodo no PDF Serviços vem como "Período: DD/MM/AAAA até DD/MM/AAAA"
# ─────────────────────────────────────────────────────────────────────────────

def extrair_pdf_servicos(path):
    texto = texto_pdf(path)

    cnpj = ""
    m = re.search(r"CNPJ:\s*([\d./-]+)", texto)
    if m:
        cnpj = normalizar_cnpj(m.group(1))

    competencia = ""
    m = re.search(r"Per[íi]odo:\s*(\d{2})/(\d{2})/(\d{4})", texto)
    if m:
        # Captura o MÊS e ANO da data inicial
        competencia = f"{m.group(2)}/{m.group(3)}"

    total_geral = None
    linhas = texto.splitlines()
    for i, linha in enumerate(linhas):
        if "Total Geral" in linha:
            valores = pegar_valores_monetarios(linha)
            if valores:
                total_geral = para_numero(valores[0])
                break
            # Pode estar na próxima linha
            for prox in linhas[i + 1:i + 8]:
                valores = pegar_valores_monetarios(prox)
                if valores:
                    total_geral = para_numero(valores[0])
                    break
            break

    return {"cnpj": cnpj, "competencia": competencia, "total_geral": total_geral}


# ─────────────────────────────────────────────────────────────────────────────
# EXTRAÇÃO PDF DOMÍNIO — SIMPLES NACIONAL
# Captura: CNPJ, Período (MM/AAAA), Regime de Competência (Total)
# Periodo no PDF Simples vem direto "Período: MM/AAAA"
# Linha "Regime de Competência" tem 3 valores: Merc.Interno, Merc.Externo, Total.
# O Total é o ÚLTIMO valor da linha.
# ─────────────────────────────────────────────────────────────────────────────

def extrair_pdf_simples(path):
    texto = texto_pdf(path)

    cnpj = ""
    m = re.search(r"CNPJ:\s*([\d./-]+)", texto)
    if m:
        cnpj = normalizar_cnpj(m.group(1))

    competencia = ""
    m = re.search(r"Per[íi]odo:\s*(\d{2}/\d{4})", texto)
    if m:
        competencia = m.group(1)

    regime_comp_total = None
    linhas = texto.splitlines()
    for linha in linhas:
        if "Regime de Compet" in linha:
            valores = pegar_valores_monetarios(linha)
            if valores:
                regime_comp_total = para_numero(valores[-1])  # ÚLTIMO valor = Total
                break

    return {"cnpj": cnpj, "competencia": competencia, "regime_comp_total": regime_comp_total}


# ─────────────────────────────────────────────────────────────────────────────
# CORE — EXECUTAR AS 5 CONFERÊNCIAS PARA UMA EMPRESA
# ─────────────────────────────────────────────────────────────────────────────

def executar_conferencia_empresa(codigo, dados_gerencial, info_concl,
                                  pasta_emp, pdfs_dominio, competencia,
                                  regime="", dom_saidas=None, dom_devolucoes=None):
    """
    Retorna um dict com o resultado da conferência completa de uma empresa.
    Status possíveis: CONSISTENTE, DIVERGENTE, PENDENCIA, NAO_ANALISADA

    O resultado inclui:
      - "checks": dict com True/False/None para cada conferência
                  (codigo, nome, qtd, competencia, cnpj, valor)
      - "valores": dict com os valores brutos extraídos de cada fonte,
                   para imprimir no log com ✓/✗ em cada linha.
    """
    eh_normal = regime and "NORMAL" in regime.upper()

    resultado = {
        "codigo":            codigo,
        "empresa_concl":     info_concl["empresa"],
        "empresa_ger":       "",
        "cnpj":              "",
        "competencia":       "",
        "regime":            regime,
        "qtd_ger":           None,
        "qtd_pdf":           None,
        "qtd_xml":           None,
        "valor_det":         None,
        "valor_serv":        None,
        "valor_saidas":      None,
        "valor_devolucoes":  None,
        "valor_simp":        None,
        "grupo":             info_concl["grupo"],
        "motivo":            info_concl.get("motivo", ""),
        "status":            "PENDENCIA",
        "observacoes":       [],
        "acao":              "",
        # Checks individuais por conferência (True = bateu, False = divergiu, None = não verificado)
        "checks": {
            "codigo":      None,
            "nome":        None,
            "qtd":         None,
            "competencia": None,
            "cnpj":        None,
            "valor":       None,
        },
        # Valores brutos por fonte (para imprimir no log com ✓/✗)
        "valores": {
            "cod_ger":   "", "cod_cnc":   "",
            "nome_ger":  "", "nome_cnc":  "",
            "qtd_ger":   None, "qtd_pdf":   None, "qtd_xml":   None,
            "comp_det":  "", "comp_serv": "", "comp_simp": "",
            "cnpj_det":  "", "cnpj_serv": "", "cnpj_simp": "",
            "val_det":   None, "val_serv":  None, "val_simp":  None,
        },
    }

    # ─── EMPRESAS NÃO ANALISADAS (SEM_MOVIMENTO ou IGNORADAS) ───
    if info_concl["grupo"] != "ANALISAR":
        resultado["status"] = "NAO_ANALISADA"
        return resultado

    info_ger = dados_gerencial.get(codigo)
    if not info_ger:
        resultado["observacoes"].append("Empresa não localizada no gerencial")
        resultado["status"] = "PENDENCIA"
        return resultado

    resultado["empresa_ger"] = info_ger["empresa"]
    resultado["qtd_ger"]     = info_ger["qtd_emitidas"]
    resultado["qtd_pdf"]     = info_concl["pdf_qtd"]
    resultado["qtd_xml"]     = info_concl["xml_qtd"]

    # ─── CONFERÊNCIA 1: CLIENTE (EXATO, sem normalização) ───
    cod_ger  = info_ger["codigo"]
    nome_ger = info_ger["empresa"]
    cod_cnc  = info_concl["codigo"]
    nome_cnc = info_concl["empresa"]

    resultado["valores"]["cod_ger"]  = cod_ger
    resultado["valores"]["cod_cnc"]  = cod_cnc
    resultado["valores"]["nome_ger"] = nome_ger
    resultado["valores"]["nome_cnc"] = nome_cnc

    resultado["checks"]["codigo"] = (cod_ger == cod_cnc)
    resultado["checks"]["nome"]   = (nome_ger == nome_cnc)

    if not resultado["checks"]["codigo"]:
        resultado["observacoes"].append(
            f"Código diverge: gerencial={cod_ger} CONCLUSAO={cod_cnc}"
        )
    if not resultado["checks"]["nome"]:
        resultado["observacoes"].append(
            f"Nome diverge: gerencial='{nome_ger}' CONCLUSAO='{nome_cnc}'"
        )

    if resultado["observacoes"]:
        resultado["status"] = "DIVERGENTE"
        return resultado

    # ─── CONFERÊNCIA 2: QUANTIDADE EMITIDAS ───
    qtd_ger_raw = info_ger["qtd_emitidas"]
    qtd_ger = para_numero(qtd_ger_raw)
    qtd_pdf = info_concl["pdf_qtd"]
    qtd_xml = info_concl["xml_qtd"]

    resultado["valores"]["qtd_ger"] = qtd_ger
    resultado["valores"]["qtd_pdf"] = qtd_pdf
    resultado["valores"]["qtd_xml"] = qtd_xml

    # Se o próprio gerencial marcou a empresa como sem movimento, ela não
    # deve quebrar a execução nem entrar na conferência de valores.
    if qtd_ger is None and "sem movimento" in str(qtd_ger_raw).lower() and not pdfs_dominio.get((codigo, "servicos")) and not pdfs_dominio.get((codigo, "simples")):
        resultado["status"] = "NAO_ANALISADA"
        resultado["grupo"] = "SEM_MOVIMENTO"
        resultado["motivo"] = "Gerencial indica sem movimento"
        return resultado

    if qtd_ger is None or qtd_pdf is None or qtd_xml is None:
        resultado["checks"]["qtd"] = False
        resultado["observacoes"].append(
            f"Quantidade não localizada: ger={qtd_ger_raw} PDF={qtd_pdf} XML={qtd_xml}"
        )
        resultado["status"] = "PENDENCIA"
        return resultado

    resultado["checks"]["qtd"] = (int(qtd_ger) == int(qtd_pdf) == int(qtd_xml))

    if not resultado["checks"]["qtd"]:
        resultado["observacoes"].append(
            f"Quantidade diverge: gerencial D={int(qtd_ger)} | "
            f"PDF={int(qtd_pdf)} | XML={int(qtd_xml)}"
        )
        resultado["status"] = "DIVERGENTE"
        return resultado

    # ─── FASE 2: localiza Detalhamento + PDFs Domínio ───
    pdf_serv = pdfs_dominio.get((codigo, "servicos"))
    pdf_simp = pdfs_dominio.get((codigo, "simples"))

    if not pdf_serv:
        resultado["observacoes"].append("PDF Domínio Serviços ausente em CONFERENCIAS")
    if not pdf_simp and not eh_normal:
        resultado["observacoes"].append("PDF Domínio Simples Nacional ausente em CONFERENCIAS")
    if not pasta_emp:
        resultado["observacoes"].append("Pasta da empresa não encontrada em PADRAO NACIONAL")

    # ─── NORMAL: dispensa Simples — apenas Det x Serviços ───
    if eh_normal and pdf_serv and pasta_emp:
        if resultado["observacoes"]:
            resultado["status"] = "PENDENCIA"
            return resultado
        det_path = encontrar_detalhamento(pasta_emp, codigo, competencia)
        if not det_path:
            resultado["observacoes"].append(
                f"Detalhamento_de_NFSe_*.xlsx não encontrado em "
                f"{pasta_emp.name}\\{competencia}\\"
            )
            resultado["status"] = "PENDENCIA"
            return resultado
        try:
            det  = ler_detalhamento(det_path)
            serv = extrair_pdf_servicos(pdf_serv)
        except Exception as e:
            resultado["observacoes"].append(f"Erro ao ler fontes: {type(e).__name__}: {e}")
            resultado["status"] = "PENDENCIA"
            return resultado
        resultado["cnpj"]             = det["cnpj"] or serv["cnpj"]
        resultado["competencia"]      = det["competencia"] or serv["competencia"]
        resultado["valor_det"]        = det["valor_total"]
        resultado["valor_serv"]       = serv["total_geral"]
        resultado["valor_saidas"]     = dom_saidas
        resultado["valor_devolucoes"] = dom_devolucoes
        resultado["valores"]["val_det"]  = det["valor_total"]
        resultado["valores"]["val_serv"] = serv["total_geral"]
        val_det  = det["valor_total"]
        val_serv = serv["total_geral"]
        if val_det is None or val_serv is None or abs(val_det - val_serv) > TOLERANCIA:
            resultado["checks"]["valor"] = False
            resultado["observacoes"].append(
                f"Valor diverge: Det={formatar(val_det)} | Serv={formatar(val_serv)}"
            )
            resultado["status"] = "DIVERGENTE"
            return resultado
        resultado["checks"]["valor"] = True
        pdfs_mover = [pdf_serv]
        pdf_saidas_path = pdfs_dominio.get((codigo, "saidas"))
        if pdf_saidas_path:
            pdfs_mover.append(pdf_saidas_path)
        destino_pasta = pasta_emp / competencia
        destino_pasta.mkdir(parents=True, exist_ok=True)
        try:
            for pdf in pdfs_mover:
                destino = destino_pasta / pdf.name
                if destino.exists():
                    destino.unlink()
                shutil.move(str(pdf), str(destino))
            resultado["observacoes"].append("Empresa com regime normal")
            resultado["observacoes"].append("PDFs Domínio movidos para pasta da empresa.")
            resultado["acao"]   = "PDFs Domínio movidos para pasta da empresa - Autorizado."
            resultado["status"] = "CONSISTENTE"
        except Exception as e:
            resultado["observacoes"].append(f"Erro ao mover PDFs: {type(e).__name__}: {e}")
            resultado["acao"]   = "PDFs NÃO movidos - erro no movimento."
            resultado["status"] = "PENDENCIA"
        return resultado

    if not pdf_simp and pdf_serv and pasta_emp and not eh_normal:
        det_path = encontrar_detalhamento(pasta_emp, codigo, competencia)
        if det_path:
            try:
                det  = ler_detalhamento(det_path)
                serv = extrair_pdf_servicos(pdf_serv)

                val_det  = det["valor_total"]
                val_serv = serv["total_geral"]

                comp_det  = det["competencia"]
                comp_serv = serv["competencia"]
                cnpj_det  = det["cnpj"]
                cnpj_serv = serv["cnpj"]

                comp_ok = (comp_det == comp_serv) if comp_det and comp_serv else bool(comp_det or comp_serv)
                cnpj_ok = (cnpj_det == cnpj_serv) if cnpj_det and cnpj_serv else bool(cnpj_det or cnpj_serv)

                if (
                    comp_ok and
                    cnpj_ok and
                    val_det is not None and
                    val_serv is not None and
                    abs(val_det - val_serv) <= TOLERANCIA
                ):
                    resultado["competencia"] = comp_det or comp_serv or competencia
                    resultado["cnpj"]        = cnpj_det or cnpj_serv or info_ger.get("inscricao", "")
                    resultado["valor_det"]   = val_det
                    resultado["valor_serv"]  = val_serv
                    resultado["status"]      = "CONSISTENTE PARCIAL"
                    resultado["observacoes"] = [
                        "Relatorio serviços bateu, porém relatorio simples nacional ausente, confirmar se a empresa é do Simples ou normal."
                    ]
                    return resultado
            except Exception:
                pass

    if resultado["observacoes"]:
        resultado["status"] = "PENDENCIA"
        return resultado

    det_path = encontrar_detalhamento(pasta_emp, codigo, competencia)
    if not det_path:
        resultado["observacoes"].append(
            f"Detalhamento_de_NFSe_*.xlsx não encontrado em "
            f"{pasta_emp.name}\\{competencia}\\"
        )
        resultado["status"] = "PENDENCIA"
        return resultado

    try:
        det  = ler_detalhamento(det_path)
        serv = extrair_pdf_servicos(pdf_serv)
        simp = extrair_pdf_simples(pdf_simp)
    except Exception as e:
        resultado["observacoes"].append(f"Erro ao ler fontes Fase 2: {type(e).__name__}: {e}")
        resultado["status"] = "PENDENCIA"
        return resultado

    resultado["cnpj"]        = det["cnpj"] or serv["cnpj"] or simp["cnpj"]
    resultado["competencia"] = det["competencia"] or serv["competencia"] or simp["competencia"]
    resultado["valor_det"]   = det["valor_total"]
    resultado["valor_serv"]  = serv["total_geral"]
    resultado["valor_simp"]  = simp["regime_comp_total"]

    resultado["valores"]["comp_det"]  = det["competencia"]
    resultado["valores"]["comp_serv"] = serv["competencia"]
    resultado["valores"]["comp_simp"] = simp["competencia"]
    resultado["valores"]["cnpj_det"]  = det["cnpj"]
    resultado["valores"]["cnpj_serv"] = serv["cnpj"]
    resultado["valores"]["cnpj_simp"] = simp["cnpj"]
    resultado["valores"]["val_det"]   = det["valor_total"]
    resultado["valores"]["val_serv"]  = serv["total_geral"]
    resultado["valores"]["val_simp"]  = simp["regime_comp_total"]

    # ─── CONFERÊNCIA 3: PERÍODO ───
    comps = [det["competencia"], serv["competencia"], simp["competencia"]]
    comps_validos = [c for c in comps if c]
    resultado["checks"]["competencia"] = (len(set(comps_validos)) == 1 and len(comps_validos) == 3)
    if not resultado["checks"]["competencia"]:
        resultado["observacoes"].append(
            f"Período diverge: Det={det['competencia']} | "
            f"Serv={serv['competencia']} | Simp={simp['competencia']}"
        )

    # ─── CONFERÊNCIA 4: CNPJ ───
    cnpjs = [det["cnpj"], serv["cnpj"], simp["cnpj"]]
    cnpjs_validos = [c for c in cnpjs if c]
    resultado["checks"]["cnpj"] = (len(set(cnpjs_validos)) == 1 and len(cnpjs_validos) == 3)
    if not resultado["checks"]["cnpj"]:
        resultado["observacoes"].append(
            f"CNPJ diverge: Det={det['cnpj']} | "
            f"Serv={serv['cnpj']} | Simp={simp['cnpj']}"
        )

    # ─── CONFERÊNCIA 5: VALOR ───
    val_det  = det["valor_total"]
    val_serv = serv["total_geral"]
    val_simp = simp["regime_comp_total"]

    resultado["valor_saidas"]     = dom_saidas
    resultado["valor_devolucoes"] = dom_devolucoes

    soma_dominio = round((val_serv or 0) + (dom_saidas or 0), 2) if val_serv is not None else None
    dif_simples  = round(soma_dominio - val_simp, 2) if soma_dominio is not None and val_simp is not None else None

    det_serv_ok = abs(val_det - val_serv) <= TOLERANCIA if val_det is not None and val_serv is not None else False
    simp_ok     = abs(dif_simples) <= TOLERANCIA if dif_simples is not None else False

    if val_det is None or val_serv is None or val_simp is None:
        resultado["checks"]["valor"] = False
        resultado["observacoes"].append(
            f"Valor faltante: Det={formatar(val_det)} | "
            f"Serv={formatar(val_serv)} | Simp={formatar(val_simp)}"
        )
    else:
        resultado["checks"]["valor"] = det_serv_ok and simp_ok
        if not det_serv_ok:
            resultado["observacoes"].append(
                f"Valor diverge: Det={formatar(val_det)} | Serv={formatar(val_serv)}"
            )
        if not simp_ok:
            if val_simp > soma_dominio:
                resultado["observacoes"].append("Simples maior que servicos+saidas — verificar saidas")
            else:
                resultado["observacoes"].append(
                    f"Simples diverge: Serv+Saidas={formatar(soma_dominio)} | Simp={formatar(val_simp)}"
                )

    if resultado["observacoes"]:
        if det_serv_ok and not simp_ok and val_simp is not None and soma_dominio is not None and val_simp > soma_dominio:
            resultado["status"] = "PARCIAL"
        else:
            resultado["status"] = "DIVERGENTE"
        return resultado

    # ─── TUDO OK → MOVER PDFs Domínio para raiz da competência ───
    pdfs_mover = [pdf_serv, pdf_simp]
    pdf_saidas_path = pdfs_dominio.get((codigo, "saidas"))
    if pdf_saidas_path:
        pdfs_mover.append(pdf_saidas_path)
    destino_pasta = pasta_emp / competencia
    destino_pasta.mkdir(parents=True, exist_ok=True)
    try:
        for pdf in pdfs_mover:
            destino = destino_pasta / pdf.name
            if destino.exists():
                destino.unlink()
            shutil.move(str(pdf), str(destino))
        resultado["observacoes"].append("PDFs Domínio movidos para pasta da empresa.")
        resultado["acao"]   = "PDFs Domínio movidos para pasta da empresa - Autorizado."
        resultado["status"] = "CONSISTENTE"
    except Exception as e:
        resultado["observacoes"].append(f"Erro ao mover PDFs: {type(e).__name__}: {e}")
        resultado["acao"]   = "PDFs NÃO movidos - erro no movimento."
        resultado["status"] = "PENDENCIA"

    return resultado


# ─────────────────────────────────────────────────────────────────────────────
# GERAÇÃO DA PLANILHA DE SAÍDA
# Corpo principal: empresas ANALISADAS (CONSISTENTE/DIVERGENTE/PENDENCIA)
# Seções finais: SEM MOVIMENTO + IGNORADAS (com motivo)
# ─────────────────────────────────────────────────────────────────────────────

COR_HEADER     = "1F3864"
COR_GRUPO_PN   = "2E75B6"
COR_GRUPO_DOM  = "548235"
COR_ZEBRA      = "F2F2F2"
COR_BRANCA     = "FFFFFF"
COR_VERDE      = "C6EFCE"
COR_VERMELHO   = "FFC7CE"
COR_AMARELO    = "FFEB9C"
COR_LARANJA    = "FCE4D6"
COR_CINZA      = "D9D9D9"
FONT_VERDE     = "006100"
FONT_VERM      = "9C0006"
FONT_AMAR      = "9C5700"
FONT_LARAN     = "843C0C"


def gerar_xlsx(resultados_analise, resultados_sem_mov, resultados_ignoradas, competencia):
    thin  = Side(style="thin",   color="D9D9D9")
    borda = Border(left=thin, right=thin, top=thin, bottom=thin)

    wb = Workbook()
    ws = wb.active
    ws.title = "CONFERENCIA PN"
    ws.sheet_view.showGridLines = False
    ws.row_dimensions[1].height = 22
    ws.row_dimensions[2].height = 6
    ws.row_dimensions[3].height = 32

    # ── Cabeçalho de grupo (linha 1) ───────────────────────────────────────
    grupos = [
        ("A1:D1", "A1", "DADOS CLIENTE",     COR_HEADER),
        ("E1:I1", "E1", "AUTOMAÇÃO (PN)",    COR_GRUPO_PN),
        ("J1:J1", "J1", "PORTAL NACIONAL",   COR_GRUPO_DOM),
        ("K1:N1", "K1", "DOMINIO SISTEMAS",  COR_GRUPO_DOM),
        ("O1:Q1", "O1", "RESULTADO",         COR_HEADER),
    ]

    for merge_ref, coord, texto, cor in grupos:
        ws.merge_cells(merge_ref)
        c = ws[coord]
        c.value     = texto
        c.font      = Font(bold=True, color="FFFFFF", name="Calibri", size=10)
        c.fill      = PatternFill("solid", fgColor=cor)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border    = borda

    # ── Cabeçalho de colunas (linha 3) ─────────────────────────────────────
    colunas = [
        ("A", "Código",                          8),
        ("B", "Empresa",                         32),
        ("C", "CNPJ",                            20),
        ("D", "Regime",                          18),
        ("E", "Competência",                     13),
        ("F", "Qtd Ger.",                        10),
        ("G", "Qtd PDF",                         10),
        ("H", "Qtd XML",                         10),
        ("I", "Status",                          18),
        ("J", "Vlr. Relatorio Detalhamento",     23.6),
        ("K", "Vlr. Relatorio Serviço Dominio",  25.0),
        ("L", "Vlr. Saídas",                     15),
        ("M", "Devoluções",                      13),
        ("N", "Vlr. Relatorio Simples Nacional", 25.6),
        ("O", "Status",                          16),
        ("P", "Ação",                            38),
        ("Q", "Observação",                      50),
    ]
    for letra, titulo, larg in colunas:
        c = ws[f"{letra}3"]
        c.value     = titulo
        c.font      = Font(bold=True, color="FFFFFF", name="Calibri", size=10)
        c.fill      = PatternFill("solid", fgColor=COR_HEADER)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border    = borda
        ws.column_dimensions[letra].width = larg

    def _status_texto(status):
        if status == "CONSISTENTE":
            return "✓ CONSISTENTE"
        if status == "CONSISTENTE PARCIAL":
            return "◐ CONSISTENTE PARCIAL"
        if status == "PARCIAL":
            return "⚠ PARCIAL"
        if status == "DIVERGENTE":
            return "✕ DIVERGENTE"
        return "⚠ PENDÊNCIA"

    def _formatar_status_cell(c, status):
        c.border = borda
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.value = _status_texto(status)
        if status == "CONSISTENTE":
            c.fill = PatternFill("solid", fgColor=COR_VERDE)
            c.font = Font(name="Calibri", size=10, bold=True, color=FONT_VERDE)
        elif status == "CONSISTENTE PARCIAL":
            c.fill = PatternFill("solid", fgColor=COR_AMARELO)
            c.font = Font(name="Calibri", size=10, bold=True, color=FONT_AMAR)
        elif status == "PARCIAL":
            c.fill = PatternFill("solid", fgColor=COR_LARANJA)
            c.font = Font(name="Calibri", size=10, bold=True, color=FONT_LARAN)
        elif status == "DIVERGENTE":
            c.fill = PatternFill("solid", fgColor=COR_VERMELHO)
            c.font = Font(name="Calibri", size=10, bold=True, color=FONT_VERM)
        else:
            c.fill = PatternFill("solid", fgColor=COR_AMARELO)
            c.font = Font(name="Calibri", size=10, bold=True, color=FONT_AMAR)

    # ── Corpo: empresas analisadas ─────────────────────────────────────────
    linha = 4
    for r in resultados_analise:
        ws.row_dimensions[linha].height = 18
        cor_bg = COR_ZEBRA if linha % 2 == 0 else COR_BRANCA
        obs_txt = " | ".join(r.get("observacoes") or [])

        # A:H — dados cliente + regime + competência + quantidades
        for letra, val, aln in [
            ("A", r["codigo"],                            "center"),
            ("B", r["empresa_ger"] or r["empresa_concl"], "left"),
            ("C", formatar_cnpj(r["cnpj"]),               "center"),
            ("D", r.get("regime", ""),                    "center"),
            ("E", r["competencia"] or competencia,        "center"),
            ("F", r["qtd_ger"],                           "center"),
            ("G", r["qtd_pdf"],                           "center"),
            ("H", r["qtd_xml"],                           "center"),
        ]:
            c = ws[f"{letra}{linha}"]
            c.value     = val
            c.border    = borda
            c.font      = Font(name="Calibri", size=10)
            c.fill      = PatternFill("solid", fgColor=cor_bg)
            c.alignment = Alignment(horizontal=aln, vertical="center", indent=1 if aln == "left" else 0)

        # I — Status da automação PN
        _formatar_status_cell(ws[f"I{linha}"], r["status"])

        # J:K — Vlr. Det e Vlr. Serviço
        for letra, val in [("J", r["valor_det"]),
                           ("K", r["valor_serv"])]:
            c = ws[f"{letra}{linha}"]
            c.border    = borda
            c.font      = Font(name="Calibri", size=10)
            c.alignment = Alignment(horizontal="right", vertical="center")
            if val is None:
                c.value = None
                c.fill  = PatternFill("solid", fgColor=cor_bg)
            else:
                c.value         = float(val)
                c.number_format = "#,##0.00"
                if r["status"] == "CONSISTENTE":
                    c.fill = PatternFill("solid", fgColor=COR_VERDE)
                    c.font = Font(name="Calibri", size=10, bold=True, color=FONT_VERDE)
                elif r["status"] == "DIVERGENTE":
                    c.fill = PatternFill("solid", fgColor=COR_VERMELHO)
                    c.font = Font(name="Calibri", size=10, bold=True, color=FONT_VERM)
                else:
                    c.fill = PatternFill("solid", fgColor=cor_bg)

        # L — Vlr. Saídas
        c = ws[f"L{linha}"]; c.border=borda
        dom_saidas_val = r.get("valor_saidas")
        if dom_saidas_val is None:
            c.value="X"; c.fill=PatternFill("solid",fgColor=COR_CINZA)
            c.font=Font(name="Calibri",size=10,color="AAAAAA")
            c.alignment=Alignment(horizontal="center",vertical="center")
        else:
            c.value=float(dom_saidas_val); c.number_format="#,##0.00"
            c.fill=PatternFill("solid",fgColor=COR_VERDE)
            c.font=Font(name="Calibri",size=10,bold=True,color=FONT_VERDE)
            c.alignment=Alignment(horizontal="right",vertical="center")

        # M — Devoluções
        c = ws[f"M{linha}"]; c.border=borda
        dom_dev_val = r.get("valor_devolucoes")
        if dom_dev_val is None:
            c.value=None; c.fill=PatternFill("solid",fgColor=COR_BRANCA)
            c.font=Font(name="Calibri",size=10)
            c.alignment=Alignment(horizontal="center",vertical="center")
        else:
            c.value=float(dom_dev_val); c.number_format="#,##0.00"
            c.fill=PatternFill("solid",fgColor=COR_AMARELO)
            c.font=Font(name="Calibri",size=10,bold=True,color=FONT_AMAR)
            c.alignment=Alignment(horizontal="right",vertical="center")

        # N — Vlr. Simples
        c = ws[f"N{linha}"]; c.border=borda
        eh_normal_emp = r.get("regime","") and "NORMAL" in str(r.get("regime","")).upper()
        val_simp = r.get("valor_simp")
        if eh_normal_emp:
            c.value="dispensado"; c.fill=PatternFill("solid",fgColor=COR_CINZA)
            c.font=Font(name="Calibri",size=9,color="AAAAAA",italic=True)
            c.alignment=Alignment(horizontal="center",vertical="center")
        elif val_simp is None:
            c.value=None; c.fill=PatternFill("solid",fgColor=cor_bg)
            c.font=Font(name="Calibri",size=10)
            c.alignment=Alignment(horizontal="right",vertical="center")
        else:
            c.value=float(val_simp); c.number_format="#,##0.00"
            if r["status"]=="CONSISTENTE":
                c.fill=PatternFill("solid",fgColor=COR_VERDE)
                c.font=Font(name="Calibri",size=10,bold=True,color=FONT_VERDE)
            elif r["status"]=="DIVERGENTE":
                c.fill=PatternFill("solid",fgColor=COR_VERMELHO)
                c.font=Font(name="Calibri",size=10,bold=True,color=FONT_VERM)
            else:
                c.fill=PatternFill("solid",fgColor=cor_bg)
                c.font=Font(name="Calibri",size=10)
            c.alignment=Alignment(horizontal="right",vertical="center")

        # O — Status do resultado final
        _formatar_status_cell(ws[f"O{linha}"], r["status"])

        # P — Ação
        c = ws[f"P{linha}"]
        c.value     = r.get("acao", "")
        c.border    = borda
        c.fill      = PatternFill("solid", fgColor=cor_bg if r.get("acao") else COR_BRANCA)
        c.font      = Font(name="Calibri", size=9, color="1F3864", bold=bool(r.get("acao")))
        c.alignment = Alignment(horizontal="left", vertical="center", indent=1)

        # Q — Observação
        c = ws[f"Q{linha}"]
        c.value     = obs_txt
        c.border    = borda
        c.fill      = PatternFill("solid", fgColor=cor_bg if obs_txt else COR_BRANCA)
        c.font      = Font(name="Calibri", size=9, italic=bool(obs_txt), color="595959")
        c.alignment = Alignment(horizontal="left", vertical="center", indent=1)

        linha += 1

    # ── Resumo do corpo principal ──────────────────────────────────────────
    linha += 1
    consistentes = sum(1 for r in resultados_analise if r["status"] == "CONSISTENTE")
    divergentes  = sum(1 for r in resultados_analise if r["status"] == "DIVERGENTE")
    pendentes    = sum(1 for r in resultados_analise if r["status"] == "PENDENCIA")
    parciais     = sum(1 for r in resultados_analise if r["status"] == "PARCIAL")
    ws.merge_cells(f"A{linha}:Q{linha}")
    c = ws[f"A{linha}"]
    c.value = (f"Analisadas: {len(resultados_analise)}  |  "
               f"✓ Consistentes: {consistentes}  |  "
               f"✕ Divergentes: {divergentes}  |  "
               f"⚠ Parciais: {parciais}  |  "
               f"⚠ Pendentes: {pendentes}")
    c.font      = Font(bold=True, color="FFFFFF", name="Calibri", size=10)
    c.fill      = PatternFill("solid", fgColor=COR_HEADER)
    c.alignment = Alignment(horizontal="center", vertical="center")
    linha += 2

    # ── Seção: SEM MOVIMENTO ───────────────────────────────────────────────
    ws.merge_cells(f"A{linha}:Q{linha}")
    c = ws[f"A{linha}"]
    c.value     = f"EMPRESAS SEM MOVIMENTO ({len(resultados_sem_mov)}) — não analisadas"
    c.font      = Font(bold=True, color="FFFFFF", name="Calibri", size=10)
    c.fill      = PatternFill("solid", fgColor="7F7F7F")
    c.alignment = Alignment(horizontal="center", vertical="center")
    linha += 1

    if resultados_sem_mov:
        for r in resultados_sem_mov:
            ws.cell(linha, 1).value = r["codigo"]
            ws.cell(linha, 1).alignment = Alignment(horizontal="center")
            ws.merge_cells(start_row=linha, start_column=2, end_row=linha, end_column=17)
            ws.cell(linha, 2).value     = r["empresa_concl"]
            ws.cell(linha, 2).alignment = Alignment(horizontal="left", indent=1)
            for col in range(1, 18):
                ws.cell(linha, col).font = Font(name="Calibri", size=10, bold=True, color=FONT_AMAR)
                ws.cell(linha, col).fill = PatternFill("solid", fgColor=COR_AMARELO)
            linha += 1
    else:
        ws.merge_cells(f"A{linha}:Q{linha}")
        c = ws[f"A{linha}"]
        c.value     = "(nenhuma)"
        c.font      = Font(name="Calibri", size=9, italic=True, color="888888")
        c.alignment = Alignment(horizontal="center")
        linha += 1

    linha += 1

    # ── Seção: IGNORADAS ───────────────────────────────────────────────────
    ws.merge_cells(f"A{linha}:Q{linha}")
    c = ws[f"A{linha}"]
    c.value     = (f"EMPRESAS IGNORADAS ({len(resultados_ignoradas)}) — "
                   f"certificado/login/portal — não analisadas")
    c.font      = Font(bold=True, color="FFFFFF", name="Calibri", size=10)
    c.fill      = PatternFill("solid", fgColor="7F7F7F")
    c.alignment = Alignment(horizontal="center", vertical="center")
    linha += 1

    if resultados_ignoradas:
        for r in resultados_ignoradas:
            ws.cell(linha, 1).value = r["codigo"]
            ws.cell(linha, 1).alignment = Alignment(horizontal="center")
            ws.merge_cells(start_row=linha, start_column=2, end_row=linha, end_column=4)
            ws.cell(linha, 2).value = r["empresa_concl"]
            ws.cell(linha, 2).alignment = Alignment(horizontal="left", indent=1)
            ws.merge_cells(start_row=linha, start_column=5, end_row=linha, end_column=17)
            motivo_label = {
                "IGNORADA_CERT":   "Certificado não localizado",
                "IGNORADA_LOGIN":  "Login/senha incorreto",
                "IGNORADA_PORTAL": "Portal instável",
            }.get(r["grupo"], r["grupo"])
            ws.cell(linha, 5).value = f"{motivo_label} — {r.get('motivo', '')}"
            ws.cell(linha, 5).alignment = Alignment(horizontal="left", indent=1)
            for col in range(1, 18):
                ws.cell(linha, col).font = Font(name="Calibri", size=10, color="595959")
                ws.cell(linha, col).fill = PatternFill("solid", fgColor=COR_CINZA)
            linha += 1
    else:
        ws.merge_cells(f"A{linha}:Q{linha}")
        c = ws[f"A{linha}"]
        c.value     = "(nenhuma)"
        c.font      = Font(name="Calibri", size=9, italic=True, color="888888")
        c.alignment = Alignment(horizontal="center")
        linha += 1

    ws.auto_filter.ref = "A3:Q3"
    ws.freeze_panes    = "A4"
    wb.save(str(ARQUIVO_SAIDA))
    log(f"Planilha gerada: {ARQUIVO_SAIDA}")


# ─────────────────────────────────────────────────────────────────────────────
# CORPO PRINCIPAL
# ─────────────────────────────────────────────────────────────────────────────

def _main_corpo():
    log("=" * 70)
    log("CONFERENCIAS DOMINIO X AUTOMACOES — PADRÃO NACIONAL")
    log("=" * 70)
    log(f"CONFERENCIAS : {PASTA_CONFERENCIAS}")
    log(f"EMPRESAS     : {PASTA_EMPRESAS}")
    log("")

    if not PASTA_CONFERENCIAS.exists():
        raise FileNotFoundError(f"Pasta não encontrada: {PASTA_CONFERENCIAS}")
    if not PASTA_EMPRESAS.exists():
        raise FileNotFoundError(f"Pasta não encontrada: {PASTA_EMPRESAS}")

    # ── Localiza Gerencial e CONCLUSAO em PASTA_EMPRESAS ────────────────────
    gerencial_path = None
    conclusao_path = None
    for arq in PASTA_EMPRESAS.iterdir():
        if not arq.is_file():
            continue
        if PADRAO_GERENCIAL.match(arq.name):
            gerencial_path = arq
        if PADRAO_CONCLUSAO.match(arq.name):
            conclusao_path = arq

    if not gerencial_path:
        log("ERRO: Relatorio_Gerencial_Padrao_Nacional_*.xlsx não encontrado em EMPRESAS.")
        return
    if not conclusao_path:
        log("ERRO: CONCLUSAO_PADRAO_NACIONAL_*.txt não encontrado em EMPRESAS.")
        return

    # Extrai competência do nome do arquivo do gerencial
    m = PADRAO_GERENCIAL.match(gerencial_path.name)
    competencia = m.group(1) if m else ""
    global _competencia_log
    _competencia_log = competencia

    log(f"Gerencial   : {gerencial_path.name}")
    log(f"Conclusao   : {conclusao_path.name}")
    log(f"Competencia : {competencia}")
    log("")

    # ── Lê dados ────────────────────────────────────────────────────────────
    dados_gerencial = ler_gerencial(gerencial_path)
    dados_conclusao = ler_conclusao(conclusao_path)
    log(f"Empresas no Gerencial : {len(dados_gerencial)}")
    log(f"Empresas no CONCLUSAO : {len(dados_conclusao)}")
    log("")

    # ── Localiza PDFs Domínio na PASTA_CONFERENCIAS ─────────────────────────
    pdfs_dominio = {}
    for arq in PASTA_CONFERENCIAS.iterdir():
        if not arq.is_file():
            continue
        m = PADRAO_SERVICOS.match(arq.name)
        if m and m.group(2) == competencia:
            pdfs_dominio[(m.group(1), "servicos")] = arq
            continue
        m = PADRAO_SIMPLES.match(arq.name)
        if m and m.group(2) == competencia:
            pdfs_dominio[(m.group(1), "simples")] = arq
            continue
        m = PADRAO_SAIDAS.match(arq.name)
        if m and m.group(2) == competencia:
            pdfs_dominio[(m.group(1), "saidas")] = arq

    # ── Processa empresa por empresa do CONCLUSAO ───────────────────────────
    resultados_analise   = []
    resultados_sem_mov   = []
    resultados_ignoradas = []

    codigos_ordenados = sorted(dados_conclusao.keys(), key=lambda x: int(x))

    for codigo in codigos_ordenados:
        log("-" * 70)
        log(f"Processando codigo {codigo} | competencia {competencia}")

        info_concl     = dados_conclusao[codigo]
        pasta_emp      = encontrar_pasta_empresa(codigo)
        regime_empresa = ler_regime_senhas(codigo)
        dom_saidas     = None
        dom_devolucoes = None
        pdf_saidas     = pdfs_dominio.get((codigo, "saidas"))
        if pdf_saidas:
            dom_saidas, dom_devolucoes = extrair_saidas_total_geral(pdf_saidas)

        resultado = executar_conferencia_empresa(
            codigo, dados_gerencial, info_concl, pasta_emp,
            pdfs_dominio, competencia,
            regime=regime_empresa, dom_saidas=dom_saidas, dom_devolucoes=dom_devolucoes
        )

        # Roteamento por status
        if resultado["status"] == "NAO_ANALISADA":
            if resultado["grupo"] == "SEM_MOVIMENTO":
                resultados_sem_mov.append(resultado)
                log(f"  {info_concl['empresa']}: SEM MOVIMENTO — descartada da análise.")
            else:
                resultados_ignoradas.append(resultado)
                motivo_lbl = {
                    "IGNORADA_CERT":   "Certificado não localizado",
                    "IGNORADA_LOGIN":  "Login/senha incorreto",
                    "IGNORADA_PORTAL": "Portal instável",
                }.get(resultado["grupo"], resultado["grupo"])
                log(f"  {info_concl['empresa']}: IGNORADA — {motivo_lbl}.")
            continue

        # Analisada — log rico com ✓/✗ em cada linha (padrão São Luís)
        resultados_analise.append(resultado)
        chk = resultado["checks"]
        val = resultado["valores"]

        def _marca(c):
            """Retorna ✓ se bateu, ✗ se diverge, vazio se não verificado."""
            if c is True:  return "✓"
            if c is False: return "✗"
            return ""

        nome_log = resultado.get("empresa_ger") or resultado["empresa_concl"]
        log(f"  Empresa : {codigo} - {nome_log} - CNPJ: {formatar_cnpj(resultado['cnpj'])}")
        log("")

        # CONFERÊNCIA 1 — Cliente (código + nome)
        log(f"  {'Codigo Gerencial':<28}: {val['cod_ger']} {_marca(chk['codigo'])}")
        log(f"  {'Codigo CONCLUSAO':<28}: {val['cod_cnc']} {_marca(chk['codigo'])}")
        log(f"  {'Nome Gerencial':<28}: {val['nome_ger']} {_marca(chk['nome'])}")
        log(f"  {'Nome CONCLUSAO':<28}: {val['nome_cnc']} {_marca(chk['nome'])}")

        # CONFERÊNCIA 2 — Quantidade emitidas
        qtd_ger_fmt = "" if val['qtd_ger'] is None else f"{int(val['qtd_ger'])}"
        qtd_pdf_fmt = "" if val['qtd_pdf'] is None else f"{int(val['qtd_pdf'])}"
        qtd_xml_fmt = "" if val['qtd_xml'] is None else f"{int(val['qtd_xml'])}"
        log(f"  {'Qtd Gerencial (col D)':<28}: {qtd_ger_fmt} {_marca(chk['qtd'])}")
        log(f"  {'Emitidas PDF':<28}: {qtd_pdf_fmt} {_marca(chk['qtd'])}")
        log(f"  {'Emitidas XML':<28}: {qtd_xml_fmt} {_marca(chk['qtd'])}")

        # CONFERÊNCIA 3 — Período
        log(f"  {'Competencia Detalhamento':<28}: {val['comp_det']} {_marca(chk['competencia'])}")
        log(f"  {'Competencia PDF Servicos':<28}: {val['comp_serv']} {_marca(chk['competencia'])}")
        log(f"  {'Competencia PDF Simples':<28}: {val['comp_simp']} {_marca(chk['competencia'])}")

        # CONFERÊNCIA 4 — CNPJ
        log(f"  {'CNPJ Detalhamento':<28}: {val['cnpj_det']} {_marca(chk['cnpj'])}")
        log(f"  {'CNPJ PDF Servicos':<28}: {val['cnpj_serv']} {_marca(chk['cnpj'])}")
        log(f"  {'CNPJ PDF Simples':<28}: {val['cnpj_simp']} {_marca(chk['cnpj'])}")

        # CONFERÊNCIA 5 — Valor
        log(f"  {'Valor Detalhamento (col G)':<28}: {formatar(val['val_det'])} {_marca(chk['valor'])}")
        log(f"  {'Valor PDF Servicos':<28}: {formatar(val['val_serv'])} {_marca(chk['valor'])}")
        log(f"  {'Valor PDF Simples':<28}: {formatar(val['val_simp'])} {_marca(chk['valor'])}")

        log(f"  Status         : {resultado['status']}")
        if resultado.get("observacoes"):
            log(f"  Observacao     : {' | '.join(resultado['observacoes'])}")
        if resultado.get("acao"):
            log(f"  Acao           : {resultado['acao']}")

    # ── Gera planilha ───────────────────────────────────────────────────────
    log("")
    codigos_processados = {r["codigo"] for r in resultados_analise + resultados_sem_mov + resultados_ignoradas}
    for cod_sm, info_sm in dados_gerencial.items():
        if cod_sm not in codigos_processados:
            qtd = str(info_sm.get("qtd_emitidas") or "").strip()
            if "sem movimento" in qtd.lower():
                resultados_sem_mov.append({
                    "codigo": cod_sm,
                    "empresa_concl": info_sm["empresa"],
                    "grupo": "SEM_MOVIMENTO",
                })
                log(f"  {info_sm['empresa']}: SEM MOVIMENTO (gerencial) — sem arquivos na pasta.")

    gerar_xlsx(resultados_analise, resultados_sem_mov, resultados_ignoradas, competencia)

    # ── Resumo final ────────────────────────────────────────────────────────
    consistentes = sum(1 for r in resultados_analise if r["status"] == "CONSISTENTE")
    divergentes  = sum(1 for r in resultados_analise if r["status"] == "DIVERGENTE")
    pendentes    = sum(1 for r in resultados_analise if r["status"] == "PENDENCIA")
    log("")
    log("=" * 70)
    log(f"Analisadas    : {len(resultados_analise)}  |  "
        f"Consistentes: {consistentes}  |  Divergentes: {divergentes}  |  "
        f"Pendentes: {pendentes}")
    log(f"Sem movimento : {len(resultados_sem_mov)}")
    log(f"Ignoradas     : {len(resultados_ignoradas)}")
    log("=" * 70)


def main():
    _pw = RWAProgressWindow()
    _stdout_original = sys.stdout
    sys.stdout = _pw
    try:
        _main_corpo()
    except Exception:
        log("")
        log("ERRO FATAL:")
        log(traceback.format_exc())
    finally:
        sys.stdout = _stdout_original
        try:
            salvar_log()
        except Exception:
            pass
        _pw.finalizar()


if __name__ == "__main__":
    main()
