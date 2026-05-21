# -*- coding: utf-8 -*-
"""
RWA Tecnologia Operacional — Agente v2.0
Motor headless. Sem interface. Roda em segundo plano.
Busca tarefas no servidor e executa as automações localmente.
"""

import os
import sys
import json
import time
import uuid
import socket
import hashlib
import getpass
import platform
import threading
import subprocess
import logging
from datetime import datetime
from pathlib import Path

import pystray
from PIL import Image, ImageDraw
import urllib.request
import urllib.error

# ── Versão e servidor ────────────────────────────────────────────────
_VERSAO       = "2.0"
_SERVIDOR_URL = "https://web-production-31152.up.railway.app"
_INTERVALO_S  = 1   # polling a cada 1 segundo

_SCRIPT_DIR = os.path.dirname(os.path.abspath(
    sys.executable if getattr(sys, "frozen", False) else __file__
))
if not os.path.isdir(_SCRIPT_DIR):
    _SCRIPT_DIR = os.getcwd()

_PASTA_CONFIG = os.path.join(os.environ.get("LOCALAPPDATA", _SCRIPT_DIR), "RWA_AUTOMACOES", "config")
_PASTA_LOGS   = os.path.join(os.environ.get("LOCALAPPDATA", _SCRIPT_DIR), "RWA_AUTOMACOES", "logs")
_ARQUIVO_CRED  = os.path.join(_PASTA_CONFIG, "credenciais.json")
_ARQUIVO_PATH  = os.path.join(_PASTA_CONFIG, "paths.json")
_ARQUIVO_LOG   = os.path.join(_PASTA_LOGS,   "agente.log")
_ARQUIVO_SINAL = os.path.join(_PASTA_CONFIG, "parar.signal")

os.makedirs(_PASTA_CONFIG, exist_ok=True)
os.makedirs(_PASTA_LOGS,   exist_ok=True)

# ── Log ──────────────────────────────────────────────────────────────
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    datefmt="%d/%m/%Y %H:%M:%S",
    handlers=[
        logging.FileHandler(_ARQUIVO_LOG, encoding="utf-8"),
        logging.StreamHandler(sys.stdout),
    ]
)
log = logging.getLogger("RWA_AGENTE")

# ── Estado global ────────────────────────────────────────────────────
_EMAIL_SESSAO   = ""
_FINGERPRINT    = ""
_loop_ativo     = False
_tarefa_rodando = False
_tray_icon      = None


# ─────────────────────────────────────────────────────────────────────
# FINGERPRINT
# ─────────────────────────────────────────────────────────────────────

def _gerar_fingerprint() -> str:
    """Mesmo algoritmo do fp.py — garante fingerprint idêntico."""
    def _serial():
        try:
            r = subprocess.check_output("wmic diskdrive get serialnumber", shell=True, stderr=subprocess.DEVNULL)
            return r.decode(errors="ignore").split()[-1]
        except Exception:
            return ""

    def _uuid_bios():
        try:
            r = subprocess.check_output("wmic csproduct get uuid", shell=True, stderr=subprocess.DEVNULL)
            return r.decode(errors="ignore").split()[-1]
        except Exception:
            return ""

    base = "|".join([
        socket.gethostname(),
        getpass.getuser(),
        platform.platform(),
        str(uuid.getnode()),
        _serial(),
        _uuid_bios(),
    ])
    return hashlib.sha256(base.encode()).hexdigest().upper()[:32]


# ─────────────────────────────────────────────────────────────────────
# CREDENCIAIS E PATHS
# ─────────────────────────────────────────────────────────────────────

def _carregar_credenciais() -> dict:
    try:
        if os.path.exists(_ARQUIVO_CRED):
            with open(_ARQUIVO_CRED, "r", encoding="utf-8") as f:
                return json.load(f)
    except Exception:
        pass
    return {}


def _carregar_paths() -> dict:
    try:
        if os.path.exists(_ARQUIVO_PATH):
            with open(_ARQUIVO_PATH, "r", encoding="utf-8") as f:
                return json.load(f)
    except Exception:
        pass
    return {"senhas_sl": "", "senhas_pn": ""}


def _normalizar_cabecalho_email(valor):
    texto = str(valor or "").strip().upper()
    for origem, destino in [("Á", "A"), ("À", "A"), ("Â", "A"), ("Ã", "A"),
                            ("É", "E"), ("Ê", "E"), ("Í", "I"),
                            ("Ó", "O"), ("Ô", "O"), ("Õ", "O"),
                            ("Ú", "U"), ("Ç", "C")]:
        texto = texto.replace(origem, destino)
    return texto


def _ler_email_destino_planilha(caminho_planilha: str) -> str:
    try:
        if not caminho_planilha or not os.path.exists(caminho_planilha):
            return ""

        import openpyxl
        wb = openpyxl.load_workbook(caminho_planilha, read_only=True, data_only=True)
        ws = wb.active
        cabecalhos = [_normalizar_cabecalho_email(c.value) for c in next(ws.iter_rows(min_row=1, max_row=1))]

        idx_email = next((i for i, c in enumerate(cabecalhos) if c in ("EMAIL", "E-MAIL") or c.startswith("EMAIL")), None)
        idx_receber = next((i for i, c in enumerate(cabecalhos) if "RECEBER" in c and "EMAIL" in c), None)

        if idx_email is None:
            idx_email = 9
        if idx_receber is None:
            idx_receber = 10

        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or len(row) <= max(idx_email, idx_receber):
                continue

            email = str(row[idx_email]).strip() if row[idx_email] else ""
            receber = _normalizar_cabecalho_email(row[idx_receber]) if row[idx_receber] else ""

            if email and "@" in email and receber == "SIM":
                return email
    except Exception as e:
        log.warning(f"[EMAIL] Falha ao ler destinatário da planilha: {e}")
    return ""


def _injetar_paths(modulo: str = ""):
    """
    Injeta os caminhos das planilhas como variáveis de ambiente
    para que as automações possam ler ao iniciar.
    """
    paths = _carregar_paths()
    if paths.get("senhas_sl"):
        os.environ["RWA_SENHAS_SL"] = paths["senhas_sl"]
    if paths.get("senhas_pn"):
        os.environ["RWA_SENHAS_PN"] = paths["senhas_pn"]

    os.environ.pop("RWA_EMAIL_DESTINO", None)
    os.environ["RWA_EMAIL_CONTA"]     = "rwaautomacoes@gmail.com"
    os.environ["RWA_EMAIL_SENHA_APP"] = "zdqjiqwpmchrbcry"
    os.environ["RWA_EMAIL_AGENDADO"]  = "1"

    caminho_planilha = ""
    if modulo == "sao_luis":
        caminho_planilha = paths.get("senhas_sl", "")
    elif modulo == "padrao_nacional":
        caminho_planilha = paths.get("senhas_pn", "")

    if caminho_planilha:
        destino = _ler_email_destino_planilha(caminho_planilha)
        if destino:
            os.environ["RWA_EMAIL_DESTINO"] = destino
            log.info(f"[EMAIL] Destinatário operacional carregado para {modulo}: {destino}")
        else:
            log.info(f"[EMAIL] Nenhum destinatário habilitado para {modulo}.")
    elif modulo in ("sao_luis", "padrao_nacional"):
        log.info(f"[EMAIL] Planilha não configurada para {modulo}; email operacional não será enviado.")


# ─────────────────────────────────────────────────────────────────────
# HTTP
# ─────────────────────────────────────────────────────────────────────

def _post(endpoint: str, payload: dict) -> dict:
    url  = _SERVIDOR_URL + endpoint
    data = json.dumps(payload).encode("utf-8")
    req  = urllib.request.Request(
        url, data=data,
        headers={"Content-Type": "application/json"},
        method="POST")
    try:
        with urllib.request.urlopen(req, timeout=10) as r:
            return json.loads(r.read().decode())
    except urllib.error.HTTPError as e:
        try:
            detail = json.loads(e.read().decode()).get("detail", f"HTTP {e.code}")
        except Exception:
            detail = f"HTTP {e.code}"
        return {"status": "erro", "mensagem": detail}
    except Exception as e:
        return {"status": "erro", "mensagem": str(e)}


def _chamar_login(email: str, senha: str) -> dict:
    return _post("/agente/login", {
        "email":       email.strip().lower(),
        "senha":       senha,
        "fingerprint": _FINGERPRINT,
        "versao":      _VERSAO,
    })


def _chamar_tarefa() -> dict:
    return _post("/agente/tarefa", {
        "email":       _EMAIL_SESSAO,
        "fingerprint": _FINGERPRINT,
    })


def _chamar_status(tarefa_id: str, status: str, obs: str = ""):
    _post("/agente/status", {
        "email":       _EMAIL_SESSAO,
        "fingerprint": _FINGERPRINT,
        "tarefa_id":   tarefa_id,
        "status":      status,
        "observacao":  obs,
    })


# ─────────────────────────────────────────────────────────────────────
# SYSTEM TRAY
# ─────────────────────────────────────────────────────────────────────

def _criar_imagem_tray():
    img = Image.new("RGBA", (64, 64), (0, 0, 0, 0))
    d   = ImageDraw.Draw(img)
    d.ellipse([4, 4, 60, 60], fill="#4f46e5")
    d.rectangle([18, 20, 24, 44], fill="white")
    d.rectangle([29, 20, 35, 44], fill="white")
    d.rectangle([40, 20, 46, 44], fill="white")
    return img


def _on_tray_sair(icon, item):
    global _loop_ativo
    log.info("[TRAY] Encerrando agente.")
    _loop_ativo = False
    icon.stop()
    sys.exit(0)


def _iniciar_tray():
    global _tray_icon
    menu = pystray.Menu(
        pystray.MenuItem("RWA — Agente ativo", lambda i, it: None, enabled=False),
        pystray.Menu.SEPARATOR,
        pystray.MenuItem("Encerrar agente", _on_tray_sair),
    )
    _tray_icon = pystray.Icon(
        "RWA_AGENTE",
        _criar_imagem_tray(),
        "RWA Tecnologia Operacional — Agente",
        menu,
    )
    threading.Thread(target=_tray_icon.run, daemon=True).start()
    log.info("[TRAY] Ícone iniciado na bandeja.")


# ─────────────────────────────────────────────────────────────────────
# PARADA FORÇADA
# ─────────────────────────────────────────────────────────────────────

def _limpar_sinal():
    try:
        if os.path.exists(_ARQUIVO_SINAL):
            os.remove(_ARQUIVO_SINAL)
    except Exception:
        pass


# Script de mapeamento módulo → arquivo
_MODULO_ARG = {
    "sao_luis":              "stm",
    "padrao_nacional":       "pn",
    "conferencia_sao_luis":  "conf_sl",
    "conferencia_pn":        "conf_pn",
}

# Processo ativo da automação
_proc_ativo = {"proc": None}


def _chamar_status_tarefa(tarefa_id: str) -> str:
    """Consulta o status atual de uma tarefa no servidor."""
    try:
        r = _post("/agente/status_tarefa", {
            "email":       _EMAIL_SESSAO,
            "fingerprint": _FINGERPRINT,
            "tarefa_id":   tarefa_id,
        })
        return r.get("status", "")
    except Exception:
        return ""


def _monitorar_parada(tarefa_id: str):
    """
    Fica verificando o arquivo de sinal.
    Se encontrar — mata o subprocesso inteiro (Python + Chrome).
    """
    while True:
        # Sinal local (launcher)
        if os.path.exists(_ARQUIVO_SINAL):
            log.info(f"[PARAR] Sinal local detectado — encerrando tarefa {tarefa_id}")
            proc = _proc_ativo.get("proc")
            if proc and proc.poll() is None:
                try:
                    subprocess.call(
                        ["taskkill", "/F", "/T", "/PID", str(proc.pid)],
                        stdout=subprocess.DEVNULL,
                        stderr=subprocess.DEVNULL)
                    log.info(f"[PARAR] Processo {proc.pid} encerrado.")
                except Exception as e:
                    log.warning(f"[PARAR] Erro ao encerrar processo: {e}")
            _limpar_sinal()
            return

        # Sinal remoto (portal web)
        status = _chamar_status_tarefa(tarefa_id)
        if status == "cancelado":
            log.info(f"[PARAR] Cancelamento remoto detectado — encerrando tarefa {tarefa_id}")
            proc = _proc_ativo.get("proc")
            if proc and proc.poll() is None:
                try:
                    subprocess.call(
                        ["taskkill", "/F", "/T", "/PID", str(proc.pid)],
                        stdout=subprocess.DEVNULL,
                        stderr=subprocess.DEVNULL)
                    log.info(f"[PARAR] Processo {proc.pid} encerrado via portal.")
                except Exception as e:
                    log.warning(f"[PARAR] Erro ao encerrar processo: {e}")
            return

        if _proc_ativo.get("proc") is None:
            return
        time.sleep(2)


# ─────────────────────────────────────────────────────────────────────
# EXECUÇÃO DE TAREFAS
# ─────────────────────────────────────────────────────────────────────

def _executar_tarefa(tarefa: dict):
    global _tarefa_rodando

    tarefa_id = str(tarefa.get("id", ""))
    modulo    = tarefa.get("modulo", "")

    log.info(f"[EXEC] Iniciando — id={tarefa_id} módulo={modulo}")
    _chamar_status(tarefa_id, "em_execucao")
    _limpar_sinal()
    _injetar_paths(modulo)

    arg = _MODULO_ARG.get(modulo)
    if not arg:
        log.error(f"[EXEC] Módulo desconhecido: {modulo}")
        _chamar_status(tarefa_id, "erro", f"Módulo desconhecido: {modulo}")
        _tarefa_rodando = False
        return

    launcher_path = os.path.join(_SCRIPT_DIR, "RWA_LAUNCHER.exe")
    if not os.path.exists(launcher_path):
        log.error(f"[EXEC] RWA_LAUNCHER.exe não encontrado: {launcher_path}")
        _chamar_status(tarefa_id, "erro", "RWA_LAUNCHER.exe não encontrado")
        _tarefa_rodando = False
        return

    log.info(f"[EXEC] Iniciando subprocesso: RWA_LAUNCHER.exe {arg}")

    try:
        # Inicia automação como subprocesso separado
        proc = subprocess.Popen(
            [launcher_path, arg],
            cwd=_SCRIPT_DIR,
            env=os.environ.copy(),
        )
        _proc_ativo["proc"] = proc

        # Monitor de parada apenas para SL e PN
        _MODULOS_COM_PARAR = {"sao_luis", "padrao_nacional"}
        if modulo in _MODULOS_COM_PARAR:
            threading.Thread(
                target=_monitorar_parada,
                args=(tarefa_id,),
                daemon=True
            ).start()

        # Aguarda o processo terminar
        proc.wait()
        _proc_ativo["proc"] = None

        if proc.returncode == 0 or proc.returncode is None:
            log.info(f"[EXEC] Concluído — {modulo}")
            _chamar_status(tarefa_id, "concluido")
        elif proc.returncode == -1 or proc.returncode == 1:
            # Pode ter sido parado pelo usuário ou erro normal
            if not os.path.exists(_ARQUIVO_SINAL):
                log.info(f"[EXEC] Interrompido — {modulo} (código {proc.returncode})")
                _chamar_status(tarefa_id, "cancelado", "Interrompido pelo usuário")
            else:
                _limpar_sinal()
                _chamar_status(tarefa_id, "cancelado", "Interrompido pelo usuário")
        else:
            log.error(f"[EXEC] Erro — {modulo} (código {proc.returncode})")
            _chamar_status(tarefa_id, "erro", f"Código de saída: {proc.returncode}")

    except Exception as e:
        log.error(f"[EXEC] Erro ao executar {modulo}: {e}")
        _chamar_status(tarefa_id, "erro", str(e))
    finally:
        _proc_ativo["proc"] = None
        _tarefa_rodando = False


# ─────────────────────────────────────────────────────────────────────
# LOOP PRINCIPAL
# ─────────────────────────────────────────────────────────────────────

def _loop_agente():
    global _tarefa_rodando

    log.info(f"[AGENTE] Loop iniciado. Intervalo: {_INTERVALO_S}s")

    while _loop_ativo:
        try:
            if not _tarefa_rodando and _EMAIL_SESSAO:
                resultado = _chamar_tarefa()
                tarefa    = resultado.get("tarefa") if resultado else None

                if tarefa:
                    log.info(f"[AGENTE] Tarefa recebida: {tarefa.get('modulo')} id={tarefa.get('id')}")
                    _tarefa_rodando = True
                    threading.Thread(
                        target=_executar_tarefa,
                        args=(tarefa,),
                        daemon=True
                    ).start()

                elif resultado and resultado.get("erro") == "maquina_nao_autorizada":
                    log.warning("[AGENTE] Máquina não autorizada no servidor.")

        except Exception as e:
            log.error(f"[AGENTE] Erro no loop: {e}")

        time.sleep(_INTERVALO_S)


# ─────────────────────────────────────────────────────────────────────
# LOGIN SILENCIOSO
# ─────────────────────────────────────────────────────────────────────

def _login_silencioso(email: str, senha: str) -> bool:
    """Valida credenciais no servidor. Retorna True se autorizado."""
    global _EMAIL_SESSAO

    log.info(f"[LOGIN] Validando — {email}")
    resp = _chamar_login(email, senha)

    if resp.get("status") == "ok":
        _EMAIL_SESSAO = email
        log.info(f"[LOGIN] Autorizado — {resp.get('cliente', '')}")
        return True
    else:
        log.error(f"[LOGIN] Recusado — {resp.get('mensagem', 'erro')}")
        return False


# ─────────────────────────────────────────────────────────────────────
# MAIN
# ─────────────────────────────────────────────────────────────────────

def main():
    global _FINGERPRINT, _loop_ativo

    log.info("=" * 55)
    log.info("  RWA Tecnologia Operacional — Agente v2.0")
    log.info("=" * 55)

    # 1. Fingerprint
    _FINGERPRINT = _gerar_fingerprint()
    log.info(f"[FP] {_FINGERPRINT}")

    # 2. Carregar credenciais salvas pelo launcher
    # Aguarda credenciais serem salvas pelo launcher (loop até existir)
    while True:
        cred = _carregar_credenciais()
        if cred and cred.get("email") and cred.get("senha"):
            break
        log.info("[AGENTE] Aguardando credenciais do launcher...")
        time.sleep(5)

    email = cred["email"]
    senha = cred["senha"]

    # 3. Login no servidor
    if not _login_silencioso(email, senha):
        log.error("[AGENTE] Login falhou. Verifique credenciais.")
        sys.exit(1)

    # 4. Tray
    _iniciar_tray()

    # 5. Loop
    _loop_ativo = True
    log.info("[AGENTE] Aguardando tarefas...")
    _loop_agente()


if __name__ == "__main__":
    main()
